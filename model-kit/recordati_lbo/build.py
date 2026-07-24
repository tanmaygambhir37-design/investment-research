"""
Deterministic reconstruction of the Recordati LBO model (14 sheets).
Auto-generated from the shipped workbook; running it reproduces the model byte-for-formula.
Requires: pip install openpyxl.   Run: python build.py   ->  writes the .xlsx next to this file.
Scenario switch lives at 03 Assumptions!B3  (1=Bear, 2=Base, 3=Bull).
"""
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
wb.remove(wb.active)
try: wb.calculation.fullCalcOnLoad = True
except Exception: pass

ws = wb.create_sheet('00 Executive Summary')
ws.column_dimensions['A'].width = 46.0
ws.column_dimensions['B'].width = 12.0
c = ws['A1']; c.value = 'RECORDATI (BIT: REC) — INDEPENDENT LBO, EXECUTIVE SUMMARY | 17 JULY 2026'
c.font = Font(bold=True, color='FFFFFFFF', size=12.0)
c.fill = PatternFill("solid", fgColor='FF1F3864')
c = ws['A3']; c.value = 'THE ANSWER'
c.font = Font(bold=True, color='FF1F3864', size=14.0)
c = ws['A4']; c.value = 'Base-case IRR (5yr, 11.5x exit):'
c.font = Font(bold=True, color='FF1F3864')
c = ws['B4']; c.value = "='10 Returns'!B11"
c.number_format = '0.0%'
c.font = Font(color='FF008000', size=10.0)
c.fill = PatternFill("solid", fgColor='FFFFFF00')
c = ws['A5']; c.value = 'Base-case MOIC:'
c.font = Font(bold=True, color='FF1F3864')
c = ws['B5']; c.value = "='10 Returns'!B9"
c.number_format = '0.00#'
c.font = Font(color='FF008000', size=10.0)
c.fill = PatternFill("solid", fgColor='FFFFFF00')
c = ws['A6']; c.value = 'Entry: 12.6x on close-date net debt (12.9x on FY25A gross, announced basis) / 5.5x leverage / ~EUR 7.6bn equity'
c.font = Font(size=10.0)
c = ws['A7']; c.value = 'Sponsor equity (model):'
c.font = Font(size=10.0)
c = ws['B7']; c.value = "='06 Sources & Uses'!B18"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF008000', size=10.0)
c = ws['A8']; c.value = 'Debt paydown over hold (gross):'
c.font = Font(size=10.0)
c = ws['B8']; c.value = "='06 Sources & Uses'!B13-'08 Debt Schedule'!K17"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF008000', size=10.0)
c = ws['A10']; c.value = 'WHAT DRIVES THE RETURN'
c.font = Font(bold=True, color='FF1F3864', size=14.0)
c = ws['A11']; c.value = '1. EBITDA growth: organic RD (+14%->9%) + SPC (+2%) to ~EUR 1.46bn by 2031; BD EUR 400m/yr 2027-2030 (exit-yr spend removed; cum. EUR 1.6bn) adds ~EUR 178m acquired EBITDA at 9.0x effective -> ~EUR 1.6bn total.'
c.font = Font(size=10.0)
c = ws['A12']; c.value = '2. Deleveraging: ~EUR 550-900m FCF/yr sweeps TLBs; leverage 5.5x -> ~1.9x by exit (BD-funded base) / ~1.0x (no-BD floor).'
c.font = Font(size=10.0)
c = ws['A13']; c.value = '3. Exit multiple: base 11.5x (COMPRESSION vs 12.9x entry). ~2.0-2.5pts IRR per turn. Rossini precedent: 0 expansion over 8 yrs.'
c.font = Font(size=10.0)
c = ws['A15']; c.value = 'KEY ASSUMPTIONS (speed-read)'
c.font = Font(bold=True, color='FF1F3864', size=14.0)
c = ws['A16']; c.value = 'RD growth y1-3 / y4-5'
c.font = Font(size=10.0)
c = ws['B16']; c.value = "='03 Assumptions'!F6"
c.number_format = '0.0%'
c.font = Font(color='FF008000', size=10.0)
c = ws['A17']; c.value = 'SPC growth'
c.font = Font(size=10.0)
c = ws['B17']; c.value = "='03 Assumptions'!F8"
c.number_format = '0.0%'
c.font = Font(color='FF008000', size=10.0)
c = ws['A18']; c.value = 'Margin path 36.5% →'
c.font = Font(size=10.0)
c = ws['B18']; c.value = "='03 Assumptions'!F9"
c.number_format = '0.0%'
c.font = Font(color='FF008000', size=10.0)
c = ws['A19']; c.value = 'Entry leverage'
c.font = Font(size=10.0)
c = ws['B19']; c.value = "='03 Assumptions'!F10"
c.number_format = '0.00#'
c.font = Font(color='FF008000', size=10.0)
c = ws['A20']; c.value = 'Exit multiple'
c.font = Font(size=10.0)
c = ws['B20']; c.value = "='03 Assumptions'!F11"
c.number_format = '0.00#'
c.font = Font(color='FF008000', size=10.0)
c = ws['A21']; c.value = 'Blended debt cost ~6.3%; tax 24%; maint. capex 1.75% rev'
c.font = Font(size=10.0)
c = ws['A23']; c.value = 'RECOMMENDED ACTION'
c.font = Font(bold=True, color='FF1F3864', size=14.0)
c = ws['A24']; c.value = 'PARTICIPATE selectively (co-invest): V3 base ~15% IRR / ~2.0x MOIC (live 00!B4:B5). No-BD conservative floor ~15% (F13=0) — funding BD is'
c.font = Font(bold=True, size=10.0)
c.fill = PatternFill("solid", fgColor='FFFFF2CC')
c = ws['A25']; c.value = 'roughly IRR-NEUTRAL: the 9x-in / 11.5x-out arbitrage is almost exactly offset by interest drag on un-swept cash + forfeited deleveraging. The'
c.font = Font(size=10.0)
c = ws['A26']; c.value = 'sponsor-consistent ~18% case is exit-multiple MAINTENANCE (entry-flat 12.9x = +2.9pt), NOT BD arbitrage. Bull (BD+12.9x+6.0x lev) ~24%/3.0x. Clears a co-invest bar, not a 20% standalone hurdle.'
c.font = Font(size=10.0)
c = ws['A28']; c.value = 'KEY RISKS (ranked)'
c.font = Font(bold=True, color='FF1F3864', size=14.0)
c = ws['A29']; c.value = '1. Exit multiple compression toward Ipsen 8.4x: at 10x exit, IRR falls to single digits — dominant risk.'
c.font = Font(size=10.0)
c = ws['A30']; c.value = '2. Isturisa vs Corcept relacorilant: RD growth engine concentration (~35-40% of RD est.).'
c.font = Font(size=10.0)
c = ws['A31']; c.value = '3. FY26 margin guide-down proves structural (not investment/FX): EBITDA base erodes.'
c.font = Font(size=10.0)
c = ws['A32']; c.value = '4. US policy (IRA/tariffs) on the RD growth market; SPC decay acceleration (Turkey/Russia).'
c.font = Font(size=10.0)
c = ws['A33']; c.value = '5. Seller = best-informed party (CVC VII → IX continuation): adverse-selection discount warranted.'
c.font = Font(size=10.0)
c = ws['A35']; c.value = "HOW TO USE: scenario switch '03 Assumptions'!B3; BLUE=input, BLACK=formula, GREEN=link, RED=check; yellow=key levers."
c.font = Font(size=10.0)
c = ws['A37']; c.value = 'GLOSSARY'
c.font = Font(bold=True, color='FF1F3864', size=14.0)
c = ws['A38']; c.value = 'IRR / MOIC'
c.font = Font(bold=True, color='FF1F3864')
c = ws['C38']; c.value = 'Internal rate of return / Multiple of invested capital (exit equity ÷ entry equity)'
c.font = Font(size=10.0)
c = ws['A39']; c.value = 'TLB / SSN / RCF'
c.font = Font(bold=True, color='FF1F3864')
c = ws['C39']; c.value = 'Term Loan B (floating, 1% amort, sweepable) / Senior Secured Notes (fixed bullet) / Revolving Credit Facility'
c.font = Font(size=10.0)
c = ws['A40']; c.value = 'Cash sweep'
c.font = Font(bold=True, color='FF1F3864')
c = ws['C40']; c.value = 'Excess FCF mandatorily prepays TLBs (EUR first, then USD)'
c.font = Font(size=10.0)
c = ws['A41']; c.value = 'OID'
c.font = Font(bold=True, color='FF1F3864')
c = ws['C41']; c.value = 'Original issue discount — financing fee reducing net debt proceeds, capitalised'
c.font = Font(size=10.0)
c = ws['A42']; c.value = 'DSCR / Coverage'
c.font = Font(bold=True, color='FF1F3864')
c = ws['C42']; c.value = 'Debt service coverage ratio / EBITDA ÷ cash interest — financeability tests'
c.font = Font(size=10.0)
c = ws['A43']; c.value = 'S&U'
c.font = Font(bold=True, color='FF1F3864')
c = ws['C43']; c.value = 'Sources & Uses — where the money comes from and goes at close'
c.font = Font(size=10.0)
c = ws['B45']; c.value = '2029'
c.font = Font(size=10.0)
c = ws['C45']; c.value = '2030'
c.font = Font(size=10.0)
c = ws['D45']; c.value = '2031'
c.font = Font(size=10.0)
c = ws['A46']; c.value = 'IRR by exit year (live):'
c.font = Font(bold=True, color='FF1F3864')
c = ws['B46']; c.value = "='10 Returns'!B27"
c.number_format = '0.0%'
c.font = Font(color='FF008000', size=10.0)
c = ws['C46']; c.value = "='10 Returns'!C27"
c.number_format = '0.0%'
c.font = Font(color='FF008000', size=10.0)
c = ws['D46']; c.value = "='10 Returns'!D27"
c.number_format = '0.0%'
c.font = Font(color='FF008000', size=10.0)

ws = wb.create_sheet('01 Cover')
ws.column_dimensions['A'].width = 46.0
ws.column_dimensions['B'].width = 12.0
c = ws['A1']; c.value = 'RECORDATI S.p.A. (BIT: REC) — INDEPENDENT LBO MODEL'
c.font = Font(bold=True, color='FFFFFFFF', size=12.0)
c.fill = PatternFill("solid", fgColor='FF1F3864')
c = ws['A3']; c.value = 'Investment question: could an outside investor independently justify the €10.7bn take-private by CVC Fund IX & GBL?'
c.font = Font(size=10.0)
c = ws['A4']; c.value = "Built BLIND to the sponsors' model, then compared against the actual announced transaction (see 13 Dashboard)."
c.font = Font(size=10.0)
c = ws['A6']; c.value = 'Prepared 17-Jul-2026 | Analyst: Tanmay Gambhir | €m | Entry: close 31-Dec-26 | Exit: 31-Dec-31 (base)'
c.font = Font(size=10.0)
c = ws['A8']; c.value = 'Color code: BLUE input | BLACK formula | GREEN cross-sheet | RED check. Yellow = primary levers.'
c.font = Font(size=10.0)
c = ws['A9']; c.value = "Scenario switch: '03 Assumptions'!B3 (1=Bear, 2=Base, 3=Bull)."
c.font = Font(size=10.0)
c = ws['A11']; c.value = '00 Executive Summary — answer first'
c.font = Font(color='FF008000', size=10.0)
c = ws['A12']; c.value = '02 Sources'
c.font = Font(color='FF008000', size=10.0)
c = ws['A13']; c.value = '03 Assumptions'
c.font = Font(color='FF008000', size=10.0)
c = ws['A14']; c.value = '04 Historical Financials'
c.font = Font(color='FF008000', size=10.0)
c = ws['A15']; c.value = '05 Transaction Overview'
c.font = Font(color='FF008000', size=10.0)
c = ws['A16']; c.value = '06 Sources & Uses'
c.font = Font(color='FF008000', size=10.0)
c = ws['A17']; c.value = '07 Purchase Price'
c.font = Font(color='FF008000', size=10.0)
c = ws['A18']; c.value = '08 Debt Schedule (sweep mechanics)'
c.font = Font(color='FF008000', size=10.0)
c = ws['A19']; c.value = '09 Operating Model (RD/SPC split)'
c.font = Font(color='FF008000', size=10.0)
c = ws['A20']; c.value = '10 Returns (IRR, MOIC, bridges)'
c.font = Font(color='FF008000', size=10.0)
c = ws['A21']; c.value = '11 Sensitivities'
c.font = Font(color='FF008000', size=10.0)
c = ws['A22']; c.value = '12 Charts'
c.font = Font(color='FF008000', size=10.0)
c = ws['A23']; c.value = '13 Dashboard + actual-deal comparison'
c.font = Font(color='FF008000', size=10.0)

ws = wb.create_sheet('03 Assumptions')
ws.column_dimensions['A'].width = 46.0
ws.column_dimensions['B'].width = 12.0
c = ws['A1']; c.value = 'ASSUMPTIONS & SCENARIO SWITCH — all inputs blue'
c.font = Font(bold=True, color='FFFFFFFF', size=12.0)
c.fill = PatternFill("solid", fgColor='FF1F3864')
c = ws['A3']; c.value = 'SCENARIO SWITCH (1=Bear, 2=Base, 3=Bull)'
c.font = Font(bold=True, color='FF1F3864')
c = ws['B3']; c.value = 2
c.number_format = '#,##0;\\(#,##0\\);\\-'
c.font = Font(color='FF0000FF', size=10.0)
c.fill = PatternFill("solid", fgColor='FFFFFF00')
c = ws['C3']; c.value = '=IF(OR(B3=1,B3=2,B3=3),"OK","ERROR: 1/2/3 only")'
c.font = Font(bold=True, color='FFFF0000', size=10.0)
c = ws['A5']; c.value = 'Scenario inputs'
c.font = Font(bold=True, color='FF1F3864')
c = ws['C5']; c.value = 'Bear'
c.font = Font(bold=True, color='FF1F3864')
c = ws['D5']; c.value = 'Base'
c.font = Font(bold=True, color='FF1F3864')
c = ws['E5']; c.value = 'Bull'
c.font = Font(bold=True, color='FF1F3864')
c = ws['F5']; c.value = 'ACTIVE'
c.font = Font(bold=True, color='FF1F3864')
c = ws['A6']; c.value = 'Rare Diseases growth 2027-29'
c.font = Font(size=10.0)
c = ws['C6']; c.value = 0.08
c.number_format = '0.0%'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['D6']; c.value = 0.14
c.number_format = '0.0%'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['E6']; c.value = 0.19
c.number_format = '0.0%'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['F6']; c.value = '=CHOOSE($B$3,C6,D6,E6)'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c.fill = PatternFill("solid", fgColor='FFFFFF00')
c = ws['A7']; c.value = 'Rare Diseases growth 2030-31'
c.font = Font(size=10.0)
c = ws['C7']; c.value = 0.05
c.number_format = '0.0%'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['D7']; c.value = 0.09
c.number_format = '0.0%'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['E7']; c.value = 0.12
c.number_format = '0.0%'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['F7']; c.value = '=CHOOSE($B$3,C7,D7,E7)'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c.fill = PatternFill("solid", fgColor='FFFFFF00')
c = ws['A8']; c.value = 'Specialty & Primary Care growth p.a.'
c.font = Font(size=10.0)
c = ws['C8']; c.value = -0.01
c.number_format = '0.0%'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['D8']; c.value = 0.02
c.number_format = '0.0%'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['E8']; c.value = 0.035
c.number_format = '0.0%'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['F8']; c.value = '=CHOOSE($B$3,C8,D8,E8)'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c.fill = PatternFill("solid", fgColor='FFFFFF00')
c = ws['A9']; c.value = 'EBITDA margin endpoint (by 2029)'
c.font = Font(size=10.0)
c = ws['C9']; c.value = 0.36
c.number_format = '0.0%'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['D9']; c.value = 0.3775
c.number_format = '0.0%'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['E9']; c.value = 0.39
c.number_format = '0.0%'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['F9']; c.value = '=CHOOSE($B$3,C9,D9,E9)'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c.fill = PatternFill("solid", fgColor='FFFFFF00')
c = ws['A10']; c.value = 'Entry leverage (x FY25 EBITDA €991m)'
c.font = Font(size=10.0)
c = ws['C10']; c.value = 5
c.number_format = '0.00#'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['D10']; c.value = 5.5
c.number_format = '0.00#'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['E10']; c.value = 6
c.number_format = '0.00#'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['F10']; c.value = '=CHOOSE($B$3,C10,D10,E10)'
c.number_format = '0.00#'
c.font = Font(size=10.0)
c.fill = PatternFill("solid", fgColor='FFFFFF00')
c = ws['A11']; c.value = 'Exit EV/EBITDA multiple'
c.font = Font(size=10.0)
c = ws['C11']; c.value = 10
c.number_format = '0.00#'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['D11']; c.value = 11.5
c.number_format = '0.00#'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['E11']; c.value = 12.9
c.number_format = '0.00#'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['F11']; c.value = '=CHOOSE($B$3,C11,D11,E11)'
c.number_format = '0.00#'
c.font = Font(size=10.0)
c.fill = PatternFill("solid", fgColor='FFFFFF00')
c = ws['A12']; c.value = 'Cash sweep % of excess FCF'
c.font = Font(size=10.0)
c = ws['C12']; c.value = 1
c.number_format = '0.0%'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['D12']; c.value = 1
c.number_format = '0.0%'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['E12']; c.value = 1
c.number_format = '0.0%'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['F12']; c.value = '=CHOOSE($B$3,C12,D12,E12)'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c.fill = PatternFill("solid", fgColor='FFFFFF00')
c = ws['A13']; c.value = 'BD / in-licensing spend per year 2027-31 (EUR m)'
c.font = Font(size=10.0)
c = ws['C13']; c.value = 0
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['D13']; c.value = 400
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['E13']; c.value = 600
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['F13']; c.value = '=CHOOSE($B$3,C13,D13,E13)'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c.fill = PatternFill("solid", fgColor='FFFFFF00')
c = ws['A15']; c.value = 'Fixed inputs'
c.font = Font(bold=True, color='FF1F3864')
c = ws['A16']; c.value = 'FY2025A EBITDA (company-adj, €m)'
c.font = Font(size=10.0)
c = ws['B16']; c.value = 991.1
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['A17']; c.value = 'FY2025A revenue (€m)'
c.font = Font(size=10.0)
c = ws['B17']; c.value = 2618.4
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['A18']; c.value = 'FY2025A Rare Diseases revenue'
c.font = Font(size=10.0)
c = ws['B18']; c.value = 1081.4
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['A19']; c.value = 'FY2025A SPC revenue'
c.font = Font(size=10.0)
c = ws['B19']; c.value = 1537
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['A20']; c.value = 'Offer price €/share (ex-div)'
c.font = Font(size=10.0)
c = ws['B20']; c.value = 51.29
c.number_format = '#,##0.00'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['A21']; c.value = 'Shares (m)'
c.font = Font(size=10.0)
c = ws['B21']; c.value = 209.125
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['A22']; c.value = 'Equity purchase (max disbursement, €m)'
c.font = Font(size=10.0)
c = ws['B22']; c.value = 10726
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['A23']; c.value = 'Existing net debt refinanced (close-date est., =B52)'
c.font = Font(size=10.0)
c = ws['B23']; c.value = '=B52'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['A24']; c.value = 'Advisory/other fees % of EV'
c.font = Font(size=10.0)
c = ws['B24']; c.value = 0.0125
c.number_format = '0.0%'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['A25']; c.value = 'OID & financing fees % of new debt'
c.font = Font(size=10.0)
c = ws['B25']; c.value = 0.025
c.number_format = '0.0%'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['A26']; c.value = 'Euribor (3M, assumed)'
c.font = Font(size=10.0)
c = ws['B26']; c.value = 0.021
c.number_format = '0.0%'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['A27']; c.value = 'TLB-EUR spread'
c.font = Font(size=10.0)
c = ws['B27']; c.value = 0.0375
c.number_format = '0.0%'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['A28']; c.value = 'TLB-USD all-in rate'
c.font = Font(size=10.0)
c = ws['B28']; c.value = 0.0725
c.number_format = '0.0%'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['A29']; c.value = 'Senior Secured Notes coupon'
c.font = Font(size=10.0)
c = ws['B29']; c.value = 0.0675
c.number_format = '0.0%'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['A30']; c.value = 'RCF size (undrawn, €m)'
c.font = Font(size=10.0)
c = ws['B30']; c.value = 500
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['A31']; c.value = 'RCF commitment fee'
c.font = Font(size=10.0)
c = ws['B31']; c.value = 0.004
c.number_format = '0.0%'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['A32']; c.value = 'Maintenance capex % revenue'
c.font = Font(size=10.0)
c = ws['B32']; c.value = 0.0175
c.number_format = '0.0%'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['A33']; c.value = 'D&A % revenue'
c.font = Font(size=10.0)
c = ws['B33']; c.value = 0.079
c.number_format = '0.0%'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['A34']; c.value = 'Cash tax rate'
c.font = Font(size=10.0)
c = ws['B34']; c.value = 0.24
c.number_format = '0.0%'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['A35']; c.value = 'Change in NWC % of revenue change'
c.font = Font(size=10.0)
c = ws['B35']; c.value = 0.15
c.number_format = '0.0%'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['A36']; c.value = 'TLB-EUR % of total debt'
c.font = Font(size=10.0)
c = ws['B36']; c.value = 0.55
c.number_format = '0.0%'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['A37']; c.value = 'TLB-USD % of total debt'
c.font = Font(size=10.0)
c = ws['B37']; c.value = 0.18
c.number_format = '0.0%'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['A38']; c.value = 'TLB mandatory amortisation p.a.'
c.font = Font(size=10.0)
c = ws['B38']; c.value = 0.01
c.number_format = '0.0%'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['A39']; c.value = '2026 RD growth (guidance-implied)'
c.font = Font(size=10.0)
c = ws['B39']; c.value = 0.14
c.number_format = '0.0%'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['A40']; c.value = '2026 SPC growth (guidance-implied)'
c.font = Font(size=10.0)
c = ws['B40']; c.value = 0
c.number_format = '0.0%'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['A41']; c.value = '2026 EBITDA margin (guidance)'
c.font = Font(size=10.0)
c = ws['B41']; c.value = 0.365
c.number_format = '0.0%'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['A42']; c.value = 'Hold period (years, exit end-2031)'
c.font = Font(size=10.0)
c = ws['B42']; c.value = 5
c.number_format = '#,##0;\\(#,##0\\);\\-'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['A43']; c.value = 'BD acquisition multiple (x EBITDA)'
c.font = Font(size=10.0)
c = ws['B43']; c.value = 9
c.number_format = '0.00#'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['A44']; c.value = 'Minimum operating cash at close (EUR m)'
c.font = Font(size=10.0)
c = ws['B44']; c.value = 250
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['A45']; c.value = 'ATAD interest deduction cap (% of EBITDA)'
c.font = Font(size=10.0)
c = ws['B45']; c.value = 0.3
c.number_format = '0.0%'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['A47']; c.value = 'v2: BASE now FUNDS BD (EUR 400m/yr at 9.0x, pre-sweep) per IC direction. Set F13=0 for the conservative no-BD floor.'
c.font = Font(size=10.0)
c = ws['A48']; c.value = 'v2: taxes apply the ATAD 30%-of-EBITDA interest deduction cap; EUR 250m minimum cash funded at close; fees 1.25% EV.'
c.font = Font(size=10.0)
c = ws['A49']; c.value = 'FY25A net debt, gross (announced-deal basis)'
c = ws['B49']; c.value = 2037.3
c = ws['A50']; c.value = '(-) 2026E FCF generated pre-close'
c = ws['B50']; c.value = -550
c = ws['A51']; c.value = '(+) 2026E dividends paid'
c = ws['B51']; c.value = 280
c = ws['A52']; c.value = 'Est. net debt at Q4-26 close (L-confidence)'
c = ws['B52']; c.value = '=B49+B50+B51'
c = ws['A54']; c.value = 'v3: net debt refinanced now uses close-date estimate (B52), not FY25A gross (B49). Lowers uses/equity check & entry EV.'

ws = wb.create_sheet('02 Sources')
ws.column_dimensions['A'].width = 46.0
ws.column_dimensions['B'].width = 12.0
ws.column_dimensions['D'].width = 85.0
ws.column_dimensions['E'].width = 12.0
c = ws['A1']; c.value = 'SOURCE REGISTER'
c.font = Font(bold=True, color='FFFFFFFF', size=12.0)
c.fill = PatternFill("solid", fgColor='FF1F3864')
c = ws['A3']; c.value = 'Item'
c.font = Font(bold=True, color='FF1F3864')
c = ws['B3']; c.value = 'Value'
c.font = Font(bold=True, color='FF1F3864')
c = ws['C3']; c.value = 'Conf.'
c.font = Font(bold=True, color='FF1F3864')
c = ws['D3']; c.value = 'Source / date'
c.font = Font(bold=True, color='FF1F3864')
c = ws['A4']; c.value = 'Offer terms'
c.font = Font(size=10.0)
c = ws['B4']; c.value = '€51.29 ex-div (€52.00 cum); 209,125,156 sh; €10,726.0m'
c.font = Font(size=10.0)
c = ws['C4']; c.value = 'H'
c.font = Font(size=10.0)
c = ws['D4']; c.value = 'Offeror PR (CVC IX & GBL via Respighi BidCo), 22-May-2026'
c.font = Font(size=10.0)
c = ws['A5']; c.value = 'Premium'
c.font = Font(size=10.0)
c = ws['B5']; c.value = '12.89% vs undisturbed 25-Mar-26'
c.font = Font(size=10.0)
c = ws['C5']; c.value = 'H'
c.font = Font(size=10.0)
c = ws['D5']; c.value = 'Offeror PR'
c.font = Font(size=10.0)
c = ws['A6']; c.value = 'FY25 revenue / EBITDA / net debt'
c.font = Font(size=10.0)
c = ws['B6']; c.value = '2,618.4 / 991.1 / 2,037.3'
c.font = Font(size=10.0)
c = ws['C6']; c.value = 'H'
c.font = Font(size=10.0)
c = ws['D6']; c.value = 'Recordati FY25 PR 17-Feb-26; accounts approved 19-Mar-26'
c.font = Font(size=10.0)
c = ws['A7']; c.value = 'FY26 guidance'
c.font = Font(size=10.0)
c = ws['B7']; c.value = 'Rev 2,730-2,800; EBITDA 995-1,030 (36.5%)'
c.font = Font(size=10.0)
c = ws['C7']; c.value = 'H'
c.font = Font(size=10.0)
c = ws['D7']; c.value = 'FY25 PR'
c.font = Font(size=10.0)
c = ws['A8']; c.value = 'Segment FY25'
c.font = Font(size=10.0)
c = ws['B8']; c.value = 'RD 1,081.4 (+29.7%); SPC ~1,537 derived'
c.font = Font(size=10.0)
c = ws['C8']; c.value = 'H/M'
c.font = Font(size=10.0)
c = ws['D8']; c.value = 'FY25 PR / derived'
c.font = Font(size=10.0)
c = ws['A9']; c.value = 'Historicals 2021-25'
c.font = Font(size=10.0)
c = ws['B9']; c.value = 'see 04'
c.font = Font(size=10.0)
c = ws['C9']; c.value = 'H'
c.font = Font(size=10.0)
c = ws['D9']; c.value = 'stockanalysis.com (Fiscal.ai), May-26; EBITDA definition note on 04'
c.font = Font(size=10.0)
c = ws['A10']; c.value = 'Debt package (actual deal)'
c.font = Font(size=10.0)
c = ws['B10']; c.value = '~€5.5-6.0bn sought; up to €8bn pitched'
c.font = Font(size=10.0)
c = ws['C10']; c.value = 'M'
c.font = Font(size=10.0)
c = ws['D10']; c.value = 'Bloomberg 17-Apr-26'
c.font = Font(size=10.0)
c = ws['A11']; c.value = 'Debt pricing assumptions'
c.font = Font(size=10.0)
c = ws['B11']; c.value = 'TLB E+375; USD 7.25%; SSN 6.75%; OID 2.5%'
c.font = Font(size=10.0)
c = ws['C11']; c.value = 'L/M'
c.font = Font(size=10.0)
c = ws['D11']; c.value = 'Market convention; Euribor 2.1% assumed (refresh at close)'
c.font = Font(size=10.0)
c = ws['A12']; c.value = 'Rossini 2018 entry'
c.font = Font(size=10.0)
c = ws['B12']; c.value = '€28.00/sh; 12.9x 2017 EBITDA; €3.03bn for 51.8%'
c.font = Font(size=10.0)
c = ws['C12']; c.value = 'H'
c.font = Font(size=10.0)
c = ws['D12']; c.value = 'Unquote/CVC PR 2018'
c.font = Font(size=10.0)
c = ws['A13']; c.value = 'Isturisa peak sales'
c.font = Font(size=10.0)
c = ws['B13']; c.value = '>€1.2bn (doubled Nov-25)'
c.font = Font(size=10.0)
c = ws['C13']; c.value = 'H'
c.font = Font(size=10.0)
c = ws['D13']; c.value = 'Recordati 9M-25 PR 11-Nov-25'
c.font = Font(size=10.0)
c = ws['A14']; c.value = 'Maintenance capex / tax'
c.font = Font(size=10.0)
c = ws['B14']; c.value = '1.5-2.0% rev / 24%'
c.font = Font(size=10.0)
c = ws['C14']; c.value = 'H'
c.font = Font(size=10.0)
c = ws['D14']; c.value = 'Fiscal.ai 2021-25: capex 22-39m, ETR 20.7-23.9%'
c.font = Font(size=10.0)
c = ws['A15']; c.value = 'Structure & scenario paths'
c.font = Font(size=10.0)
c = ws['B15']; c.value = '03 Assumptions'
c.font = Font(size=10.0)
c = ws['C15']; c.value = 'L/M'
c.font = Font(size=10.0)
c = ws['D15']; c.value = 'Analyst judgment; see project Assumptions Register'
c.font = Font(size=10.0)
c = ws['A18']; c.value = 'Page citations: offer terms — Offeror PR pp.1-2; FY25 figures — FY25 PR headline table; segments — FY25/Q1-25 PRs.'
c.font = Font(size=10.0)
c = ws['A19']; c.value = 'No annotated screenshots (environment limitation); citations are table-level and PDFs are linked in the project Bibliography.'
c.font = Font(size=10.0)

ws = wb.create_sheet('04 Historical Financials')
ws.column_dimensions['A'].width = 46.0
ws.column_dimensions['B'].width = 12.0
c = ws['A1']; c.value = 'HISTORICAL FINANCIALS 2021-2025'
c.font = Font(bold=True, color='FFFFFFFF', size=12.0)
c.fill = PatternFill("solid", fgColor='FF1F3864')
c = ws['A3']; c.value = '€m'
c.font = Font(bold=True, color='FF1F3864')
c = ws['B3']; c.value = '2021'
c.font = Font(bold=True, color='FF1F3864')
c = ws['C3']; c.value = '2022'
c.font = Font(bold=True, color='FF1F3864')
c = ws['D3']; c.value = '2023'
c.font = Font(bold=True, color='FF1F3864')
c = ws['E3']; c.value = '2024'
c.font = Font(bold=True, color='FF1F3864')
c = ws['F3']; c.value = '2025'
c.font = Font(bold=True, color='FF1F3864')
c = ws['A4']; c.value = 'Revenue'
c.font = Font(size=10.0)
c = ws['B4']; c.value = 1580
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['C4']; c.value = 1853
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['D4']; c.value = 2082
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['E4']; c.value = 2342
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['F4']; c.value = 2618.4
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['A5']; c.value = 'Company-adj. EBITDA'
c.font = Font(size=10.0)
c = ws['B5']; c.value = 588
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['C5']; c.value = 692
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['D5']; c.value = 770
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['E5']; c.value = 865.8
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['F5']; c.value = 991.1
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['A6']; c.value = 'R&D expense'
c.font = Font(size=10.0)
c = ws['B6']; c.value = 166
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['C6']; c.value = 220
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['D6']; c.value = 256
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['E6']; c.value = 286
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['F6']; c.value = 341
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['A7']; c.value = 'Operating cash flow'
c.font = Font(size=10.0)
c = ws['B7']; c.value = 492
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['C7']; c.value = 462
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['D7']; c.value = 485
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['E7']; c.value = 570
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['F7']; c.value = 597
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['A8']; c.value = 'PP&E capex'
c.font = Font(size=10.0)
c = ws['B8']; c.value = 22
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['C8']; c.value = 24
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['D8']; c.value = 30
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['E8']; c.value = 37
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['F8']; c.value = 39
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['A9']; c.value = 'FCF (OCF - capex)'
c.font = Font(size=10.0)
c = ws['B9']; c.value = 470
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['C9']; c.value = 438
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['D9']; c.value = 456
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['E9']; c.value = 533
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['F9']; c.value = 557
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['A10']; c.value = 'Intangibles purchased (BD/in-licensing)'
c.font = Font(size=10.0)
c = ws['B10']; c.value = 66
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['C10']; c.value = 72
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['D10']; c.value = 354
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['E10']; c.value = 815
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['F10']; c.value = 46
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['A11']; c.value = 'DPS (€)'
c.font = Font(size=10.0)
c = ws['B11']; c.value = 1.1
c.number_format = '#,##0.00'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['C11']; c.value = 1.15
c.number_format = '#,##0.00'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['D11']; c.value = 1.2
c.number_format = '#,##0.00'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['E11']; c.value = 1.27
c.number_format = '#,##0.00'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['F11']; c.value = 1.34
c.number_format = '#,##0.00'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['A12']; c.value = 'Effective tax rate'
c.font = Font(size=10.0)
c = ws['B12']; c.value = 0.167
c.number_format = '0.0%'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['C12']; c.value = 0.222
c.number_format = '0.0%'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['D12']; c.value = 0.207
c.number_format = '0.0%'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['E12']; c.value = 0.239
c.number_format = '0.0%'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['F12']; c.value = 0.237
c.number_format = '0.0%'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['A13']; c.value = 'Net debt (year-end)'
c.font = Font(size=10.0)
c = ws['E13']; c.value = 1875
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['F13']; c.value = 2037.3
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['A15']; c.value = 'EBITDA margin'
c.font = Font(size=10.0)
c = ws['B15']; c.value = '=B5/B4'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['C15']; c.value = '=C5/C4'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['D15']; c.value = '=D5/D4'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['E15']; c.value = '=E5/E4'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['F15']; c.value = '=F5/F4'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['A16']; c.value = 'Revenue growth'
c.font = Font(size=10.0)
c = ws['C16']; c.value = '=C4/B4-1'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['D16']; c.value = '=D4/C4-1'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['E16']; c.value = '=E4/D4-1'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['F16']; c.value = '=F4/E4-1'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['A17']; c.value = 'FCF / EBITDA conversion'
c.font = Font(size=10.0)
c = ws['B17']; c.value = '=B9/B5'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['C17']; c.value = '=C9/C5'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['D17']; c.value = '=D9/D5'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['E17']; c.value = '=E9/E5'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['F17']; c.value = '=F9/F5'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['A19']; c.value = 'NOTE: 2021-24 EBITDA shown uses company-adjusted definition where available; aggregator-standardised EBITDA differs'
c.font = Font(size=10.0)
c = ws['A20']; c.value = '(2025: 877 standardised vs 991 company-adj). Deal math throughout uses the COMPANY-ADJUSTED figure (offer-consistent).'
c.font = Font(size=10.0)
c = ws['A21']; c.value = 'Growth engine runs through the intangibles line (Enjaymo €815m in 2024) — base case deliberately excludes BD spend.'
c.font = Font(size=10.0)

ws = wb.create_sheet('05 Transaction Overview')
ws.column_dimensions['A'].width = 46.0
ws.column_dimensions['B'].width = 12.0
c = ws['A1']; c.value = 'TRANSACTION OVERVIEW — CVC IX & GBL / RESPIGHI BIDCO'
c.font = Font(bold=True, color='FFFFFFFF', size=12.0)
c.fill = PatternFill("solid", fgColor='FF1F3864')
c = ws['A3']; c.value = 'Announcement'
c.font = Font(bold=True, color='FF1F3864')
c = ws['C3']; c.value = '22 May 2026 (non-binding interest 25-Mar-26)'
c.font = Font(size=10.0)
c = ws['A4']; c.value = 'Structure'
c.font = Font(bold=True, color='FF1F3864')
c = ws['C4']; c.value = 'Voluntary totalitarian cash tender offer; delisting from Euronext Milan'
c.font = Font(size=10.0)
c = ws['A5']; c.value = 'Offer price'
c.font = Font(bold=True, color='FF1F3864')
c = ws['C5']; c.value = '€51.29 ex-div (€52.00 cum 2025 final dividend €0.71)'
c.font = Font(size=10.0)
c = ws['A6']; c.value = 'Premium'
c.font = Font(bold=True, color='FF1F3864')
c = ws['C6']; c.value = '12.89% vs undisturbed 25-Mar-26 (low: Rossini 46.82% pre-committed)'
c.font = Font(size=10.0)
c = ws['A7']; c.value = 'Rossini'
c.font = Font(bold=True, color='FF1F3864')
c = ws['C7']; c.value = 'CVC VII vehicle, 46.82%, irrevocably tendering — effectively CVC VII → CVC IX continuation'
c.font = Font(size=10.0)
c = ws['A8']; c.value = 'Co-investors'
c.font = Font(bold=True, color='FF1F3864')
c = ws['C8']; c.value = 'GBL (≤~€1.3bn), ADIA/Luxinva, CPPIB, PSP, StepStone, AlpInvest, MGG, CapSol, A. Recordati rollover'
c.font = Font(size=10.0)
c = ws['A9']; c.value = 'Financing'
c.font = Font(bold=True, color='FF1F3864')
c = ws['C9']; c.value = 'Fully committed; ~€5.5-6.0bn debt sought (Mediobanca, GS, Jefferies, JPM, DB)'
c.font = Font(size=10.0)
c = ws['A10']; c.value = 'Conditions'
c.font = Font(bold=True, color='FF1F3864')
c = ws['C10']; c.value = '≥66.67% acceptance; antitrust/FDI/FSR'
c.font = Font(size=10.0)
c = ws['A11']; c.value = 'Timeline'
c.font = Font(bold=True, color='FF1F3864')
c = ws['C11']; c.value = 'Closing expected Q4 2026; squeeze-out/merger thereafter'
c.font = Font(size=10.0)
c = ws['A12']; c.value = 'Rationale (offeror)'
c.font = Font(bold=True, color='FF1F3864')
c = ws['C12']; c.value = 'Private ownership for higher-risk BD/R&D phase without public-market earnings scrutiny'
c.font = Font(size=10.0)
c = ws['A15']; c.value = 'THIS MODEL: an INDEPENDENT underwrite at the same price — sponsor economics tested blind, then reconciled (13 Dashboard).'
c.font = Font(size=10.0)
c = ws['A16']; c.value = 'Why the market traded at EUR 46 undisturbed: FY26 margin guide-down (37.8%->36.5%), FX headwind ~-3.5%, US policy'
c.font = Font(size=10.0)
c = ws['A17']; c.value = 'overhang and Isturisa competitive risk - i.e. the public market partially priced the risks our bear case models. The low'
c.font = Font(size=10.0)
c = ws['A18']; c.value = '12.89% premium cleared because Rossini (46.82%) pre-committed; minorities had no effective blocking position.'
c.font = Font(size=10.0)

ws = wb.create_sheet('06 Sources & Uses')
ws.column_dimensions['A'].width = 46.0
ws.column_dimensions['B'].width = 12.0
c = ws['A1']; c.value = 'SOURCES & USES (at close, 31-Dec-2026)'
c.font = Font(bold=True, color='FFFFFFFF', size=12.0)
c.fill = PatternFill("solid", fgColor='FF1F3864')
c = ws['A3']; c.value = 'USES (€m)'
c.font = Font(bold=True, color='FF1F3864')
c = ws['A4']; c.value = 'Equity purchase price'
c.font = Font(size=10.0)
c = ws['B4']; c.value = "='03 Assumptions'!B22"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A5']; c.value = 'Refinance existing net debt'
c.font = Font(size=10.0)
c = ws['B5']; c.value = "='03 Assumptions'!B23"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A6']; c.value = 'Advisory & other fees (1.25% of EV, = 03!B24)'
c.font = Font(size=10.0)
c = ws['B6']; c.value = "='03 Assumptions'!B24*('07 Purchase Price'!B7)"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A7']; c.value = 'OID & financing fees (2.5% of new debt)'
c.font = Font(size=10.0)
c = ws['B7']; c.value = "='03 Assumptions'!B25*B13"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A8']; c.value = 'Minimum operating cash to balance sheet'
c.font = Font(size=10.0)
c = ws['B8']; c.value = "='03 Assumptions'!B44"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF008000', size=10.0)
c = ws['A9']; c.value = 'TOTAL USES'
c.font = Font(bold=True, color='FF1F3864')
c = ws['B9']; c.value = '=SUM(B4:B8)'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c.fill = PatternFill("solid", fgColor='FFFFFF00')
c = ws['A11']; c.value = 'SOURCES (EUR m)'
c.font = Font(bold=True, color='FF1F3864')
c = ws['A12']; c.value = 'Total new debt = leverage x FY25A EBITDA'
c.font = Font(size=10.0)
c = ws['B12']; c.value = "='03 Assumptions'!F10"
c.number_format = '0.00#'
c.font = Font(color='FF008000', size=10.0)
c = ws['A13']; c.value = '  Total debt (EUR m)'
c.font = Font(size=10.0)
c = ws['B13']; c.value = "='03 Assumptions'!F10*'03 Assumptions'!B16"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A14']; c.value = '  Term Loan B (EUR) - E+375, 1% amort'
c.font = Font(size=10.0)
c = ws['B14']; c.value = "=B13*'03 Assumptions'!B36"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A15']; c.value = '  Term Loan B (USD) - 7.25% all-in, 1% amort'
c.font = Font(size=10.0)
c = ws['B15']; c.value = "=B13*'03 Assumptions'!B37"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A16']; c.value = '  Senior Secured Notes - 6.75%, bullet'
c.font = Font(size=10.0)
c = ws['B16']; c.value = '=B13-B14-B15'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A17']; c.value = '  RCF EUR 500m - undrawn at close'
c.font = Font(size=10.0)
c = ws['B17']; c.value = 0
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['A18']; c.value = 'Sponsor & co-invest equity (plug)'
c.font = Font(bold=True, color='FF1F3864')
c = ws['B18']; c.value = '=B9-B13'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c.fill = PatternFill("solid", fgColor='FFFFFF00')
c = ws['A19']; c.value = 'TOTAL SOURCES'
c.font = Font(bold=True, color='FF1F3864')
c = ws['B19']; c.value = '=B13+B17+B18'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A21']; c.value = '=IF(ABS(B19-B9)<0.1,"CHECK OK: sources = uses","ERROR: S&U do not balance")'
c.font = Font(bold=True, color='FFFF0000', size=10.0)
c = ws['A22']; c.value = 'Equity % of capitalisation'
c.font = Font(size=10.0)
c = ws['B22']; c.value = '=B18/B19'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['A23']; c.value = 'USD tranche = natural hedge against US rare-disease revenue (dossier s.12).'
c.font = Font(size=10.0)

ws = wb.create_sheet('07 Purchase Price')
ws.column_dimensions['A'].width = 46.0
ws.column_dimensions['B'].width = 12.0
c = ws['A1']; c.value = 'PURCHASE PRICE & ENTRY MULTIPLES'
c.font = Font(bold=True, color='FFFFFFFF', size=12.0)
c.fill = PatternFill("solid", fgColor='FF1F3864')
c = ws['A3']; c.value = 'Offer price (€/share, ex-div)'
c.font = Font(size=10.0)
c = ws['B3']; c.value = "='03 Assumptions'!B20"
c.number_format = '#,##0.00'
c.font = Font(color='FF008000', size=10.0)
c = ws['A4']; c.value = 'Shares outstanding (m)'
c.font = Font(size=10.0)
c = ws['B4']; c.value = "='03 Assumptions'!B21"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF008000', size=10.0)
c = ws['A5']; c.value = 'Equity purchase price (€m)'
c.font = Font(size=10.0)
c = ws['B5']; c.value = "='03 Assumptions'!B22"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF008000', size=10.0)
c = ws['A6']; c.value = '(+) Existing net debt refinanced'
c.font = Font(size=10.0)
c = ws['B6']; c.value = "='03 Assumptions'!B23"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF008000', size=10.0)
c = ws['A7']; c.value = 'Enterprise value (€m)'
c.font = Font(size=10.0)
c = ws['B7']; c.value = '=B5+B6'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A8']; c.value = 'EV / FY2025A EBITDA (close-date net debt, independent)'
c.font = Font(size=10.0)
c = ws['B8']; c.value = "=B7/'03 Assumptions'!B16"
c.number_format = '0.00#'
c.font = Font(size=10.0)
c = ws['A9']; c.value = 'EV / FY2026E EBITDA (model)'
c.font = Font(size=10.0)
c = ws['B9']; c.value = "=B7/'09 Operating Model'!F18"
c.number_format = '0.00#'
c.font = Font(size=10.0)
c = ws['A10']; c.value = 'Cum-dividend price (€52.00) vs undisturbed €46.06'
c.font = Font(size=10.0)
c = ws['B10']; c.value = '=52/46.06-1'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['A11']; c.value = 'STRESS: EV / standardised (aggregator) EBITDA EUR 877m'
c.font = Font(size=10.0)
c = ws['B11']; c.value = '=B7/877'
c.number_format = '0.00#'
c.font = Font(size=10.0)
c = ws['A12']; c.value = 'Reference: Rossini/CVC VII entered 2018 at €28.00/sh = 12.9x 2017 EBITDA — identical multiple, zero expansion over 8 years.'
c.font = Font(size=10.0)
c = ws['A13']; c.value = '=IF(ABS(B14-12.87)<0.15,"CHECK OK: announced-basis entry ~12.9x ties to deal math","CHECK: entry multiple deviates from deal math")'
c.font = Font(bold=True, color='FFFF0000', size=10.0)
c = ws['A14']; c.value = 'Ref: EV/EBITDA on FY25A gross ND EUR 2,037m (announced-deal basis)'
c = ws['B14']; c.value = "=(B5+'03 Assumptions'!B49)/'03 Assumptions'!B16"

ws = wb.create_sheet('09 Operating Model')
ws.column_dimensions['A'].width = 46.0
ws.column_dimensions['B'].width = 12.0
c = ws['A1']; c.value = 'OPERATING MODEL - RD / SPC SPLIT (organic; row 17-18 add BD-acquired EBITDA)'
c.font = Font(bold=True, color='FFFFFFFF', size=12.0)
c.fill = PatternFill("solid", fgColor='FF1F3864')
c = ws['A3']; c.value = '€m'
c.font = Font(bold=True, color='FF1F3864')
c = ws['F3']; c.value = '2026E'
c.font = Font(bold=True, color='FF1F3864')
c = ws['G3']; c.value = '2027'
c.font = Font(bold=True, color='FF1F3864')
c = ws['H3']; c.value = '2028'
c.font = Font(bold=True, color='FF1F3864')
c = ws['I3']; c.value = '2029'
c.font = Font(bold=True, color='FF1F3864')
c = ws['J3']; c.value = '2030'
c.font = Font(bold=True, color='FF1F3864')
c = ws['K3']; c.value = '2031'
c.font = Font(bold=True, color='FF1F3864')
c = ws['A4']; c.value = 'Year of hold (t)'
c.font = Font(size=10.0)
c = ws['G4']; c.value = 1
c.number_format = '#,##0;\\(#,##0\\);\\-'
c.font = Font(size=10.0)
c = ws['H4']; c.value = 2
c.number_format = '#,##0;\\(#,##0\\);\\-'
c.font = Font(size=10.0)
c = ws['I4']; c.value = 3
c.number_format = '#,##0;\\(#,##0\\);\\-'
c.font = Font(size=10.0)
c = ws['J4']; c.value = 4
c.number_format = '#,##0;\\(#,##0\\);\\-'
c.font = Font(size=10.0)
c = ws['K4']; c.value = 5
c.number_format = '#,##0;\\(#,##0\\);\\-'
c.font = Font(size=10.0)
c = ws['A5']; c.value = 'Rare Diseases revenue'
c.font = Font(size=10.0)
c = ws['F5']; c.value = "='03 Assumptions'!B18*(1+'03 Assumptions'!B39)"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['G5']; c.value = "=F5*(1+IF(G4<=3,'03 Assumptions'!$F$6,'03 Assumptions'!$F$7))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['H5']; c.value = "=G5*(1+IF(H4<=3,'03 Assumptions'!$F$6,'03 Assumptions'!$F$7))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['I5']; c.value = "=H5*(1+IF(I4<=3,'03 Assumptions'!$F$6,'03 Assumptions'!$F$7))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['J5']; c.value = "=I5*(1+IF(J4<=3,'03 Assumptions'!$F$6,'03 Assumptions'!$F$7))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['K5']; c.value = "=J5*(1+IF(K4<=3,'03 Assumptions'!$F$6,'03 Assumptions'!$F$7))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A6']; c.value = 'SPC revenue'
c.font = Font(size=10.0)
c = ws['F6']; c.value = "='03 Assumptions'!B19*(1+'03 Assumptions'!B40)"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['G6']; c.value = "=F6*(1+'03 Assumptions'!$F$8)"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['H6']; c.value = "=G6*(1+'03 Assumptions'!$F$8)"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['I6']; c.value = "=H6*(1+'03 Assumptions'!$F$8)"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['J6']; c.value = "=I6*(1+'03 Assumptions'!$F$8)"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['K6']; c.value = "=J6*(1+'03 Assumptions'!$F$8)"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A7']; c.value = 'Total revenue'
c.font = Font(bold=True, color='FF1F3864')
c = ws['F7']; c.value = '=F5+F6'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['G7']; c.value = '=G5+G6'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['H7']; c.value = '=H5+H6'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['I7']; c.value = '=I5+I6'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['J7']; c.value = '=J5+J6'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['K7']; c.value = '=K5+K6'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A8']; c.value = 'EBITDA'
c.font = Font(bold=True, color='FF1F3864')
c = ws['F8']; c.value = "=F7*'03 Assumptions'!B41"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['G8']; c.value = "=G7*('03 Assumptions'!$B$41+('03 Assumptions'!$F$9-'03 Assumptions'!$B$41)*MIN(G4,3)/3)"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['H8']; c.value = "=H7*('03 Assumptions'!$B$41+('03 Assumptions'!$F$9-'03 Assumptions'!$B$41)*MIN(H4,3)/3)"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['I8']; c.value = "=I7*('03 Assumptions'!$B$41+('03 Assumptions'!$F$9-'03 Assumptions'!$B$41)*MIN(I4,3)/3)"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['J8']; c.value = "=J7*('03 Assumptions'!$B$41+('03 Assumptions'!$F$9-'03 Assumptions'!$B$41)*MIN(J4,3)/3)"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['K8']; c.value = "=K7*('03 Assumptions'!$B$41+('03 Assumptions'!$F$9-'03 Assumptions'!$B$41)*MIN(K4,3)/3)"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A9']; c.value = 'EBITDA margin'
c.font = Font(size=10.0)
c = ws['F9']; c.value = '=F8/F7'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['G9']; c.value = '=G8/G7'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['H9']; c.value = '=H8/H7'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['I9']; c.value = '=I8/I7'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['J9']; c.value = '=J8/J7'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['K9']; c.value = '=K8/K7'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['A10']; c.value = 'D&A'
c.font = Font(size=10.0)
c = ws['F10']; c.value = "=-F7*'03 Assumptions'!$B$33"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['G10']; c.value = "=-G7*'03 Assumptions'!$B$33"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['H10']; c.value = "=-H7*'03 Assumptions'!$B$33"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['I10']; c.value = "=-I7*'03 Assumptions'!$B$33"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['J10']; c.value = "=-J7*'03 Assumptions'!$B$33"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['K10']; c.value = "=-K7*'03 Assumptions'!$B$33"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A11']; c.value = 'EBIT'
c.font = Font(size=10.0)
c = ws['F11']; c.value = '=F8+F10'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['G11']; c.value = '=G8+G10'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['H11']; c.value = '=H8+H10'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['I11']; c.value = '=I8+I10'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['J11']; c.value = '=J8+J10'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['K11']; c.value = '=K8+K10'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A12']; c.value = 'Maintenance capex'
c.font = Font(size=10.0)
c = ws['F12']; c.value = "=-F7*'03 Assumptions'!$B$32"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['G12']; c.value = "=-G7*'03 Assumptions'!$B$32"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['H12']; c.value = "=-H7*'03 Assumptions'!$B$32"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['I12']; c.value = "=-I7*'03 Assumptions'!$B$32"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['J12']; c.value = "=-J7*'03 Assumptions'!$B$32"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['K12']; c.value = "=-K7*'03 Assumptions'!$B$32"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A13']; c.value = 'Change in NWC (cash out)'
c.font = Font(size=10.0)
c = ws['F13']; c.value = "=-(F7-'03 Assumptions'!B17)*'03 Assumptions'!$B$35"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['G13']; c.value = "=-(G7-F7)*'03 Assumptions'!$B$35"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['H13']; c.value = "=-(H7-G7)*'03 Assumptions'!$B$35"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['I13']; c.value = "=-(I7-H7)*'03 Assumptions'!$B$35"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['J13']; c.value = "=-(J7-I7)*'03 Assumptions'!$B$35"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['K13']; c.value = "=-(K7-J7)*'03 Assumptions'!$B$35"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A15']; c.value = 'RD % of revenue'
c.font = Font(size=10.0)
c = ws['F15']; c.value = '=F5/F7'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['G15']; c.value = '=G5/G7'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['H15']; c.value = '=H5/H7'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['I15']; c.value = '=I5/I7'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['J15']; c.value = '=J5/J7'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['K15']; c.value = '=K5/K7'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['A16']; c.value = 'EBITDA growth'
c.font = Font(size=10.0)
c = ws['G16']; c.value = '=G8/F8-1'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['H16']; c.value = '=H8/G8-1'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['I16']; c.value = '=I8/H8-1'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['J16']; c.value = '=J8/I8-1'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['K16']; c.value = '=K8/J8-1'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['A17']; c.value = 'Acquired EBITDA from BD (cumulative; spend/9.0x, contributes from year after spend)'
c.font = Font(size=10.0)
c = ws['F17']; c.value = 0
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['G17']; c.value = 0
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['H17']; c.value = "=G17+'03 Assumptions'!$F$13/'03 Assumptions'!$B$43"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['I17']; c.value = "=H17+'03 Assumptions'!$F$13/'03 Assumptions'!$B$43"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['J17']; c.value = "=I17+'03 Assumptions'!$F$13/'03 Assumptions'!$B$43"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['K17']; c.value = "=J17+'03 Assumptions'!$F$13/'03 Assumptions'!$B$43"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A18']; c.value = 'TOTAL EBITDA (organic + acquired)'
c.font = Font(bold=True, color='FF1F3864')
c = ws['F18']; c.value = '=F8+F17'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['G18']; c.value = '=G8+G17'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['H18']; c.value = '=H8+H17'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['I18']; c.value = '=I8+I17'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['J18']; c.value = '=J8+J17'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['K18']; c.value = '=K8+K17'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A20']; c.value = '=IF(AND(F8>995,F8<1031),"CHECK OK: 2026E EBITDA within guidance 995-1,030","CHECK: 2026E EBITDA outside guidance")'
c.font = Font(bold=True, color='FFFF0000', size=10.0)

ws = wb.create_sheet('08 Debt Schedule')
ws.column_dimensions['A'].width = 46.0
ws.column_dimensions['B'].width = 12.0
c = ws['A1']; c.value = 'DEBT SCHEDULE — mandatory amort + 100% cash sweep (TLB-EUR → TLB-USD; SSN bullet)'
c.font = Font(bold=True, color='FFFFFFFF', size=12.0)
c.fill = PatternFill("solid", fgColor='FF1F3864')
c = ws['A3']; c.value = '€m'
c.font = Font(bold=True, color='FF1F3864')
c = ws['G3']; c.value = '2027'
c.font = Font(bold=True, color='FF1F3864')
c = ws['H3']; c.value = '2028'
c.font = Font(bold=True, color='FF1F3864')
c = ws['I3']; c.value = '2029'
c.font = Font(bold=True, color='FF1F3864')
c = ws['J3']; c.value = '2030'
c.font = Font(bold=True, color='FF1F3864')
c = ws['K3']; c.value = '2031'
c.font = Font(bold=True, color='FF1F3864')
c = ws['A5']; c.value = 'TLB-EUR beginning'
c.font = Font(size=10.0)
c = ws['G5']; c.value = "='06 Sources & Uses'!B14"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF008000', size=10.0)
c = ws['H5']; c.value = '=G8'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['I5']; c.value = '=H8'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['J5']; c.value = '=I8'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['K5']; c.value = '=J8'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A6']; c.value = '  mandatory amort'
c.font = Font(size=10.0)
c = ws['G6']; c.value = "=-MIN(G5,('06 Sources & Uses'!$B$14)*'03 Assumptions'!$B$38)"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['H6']; c.value = "=-MIN(H5,('06 Sources & Uses'!$B$14)*'03 Assumptions'!$B$38)"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['I6']; c.value = "=-MIN(I5,('06 Sources & Uses'!$B$14)*'03 Assumptions'!$B$38)"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['J6']; c.value = "=-MIN(J5,('06 Sources & Uses'!$B$14)*'03 Assumptions'!$B$38)"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['K6']; c.value = "=-MIN(K5,('06 Sources & Uses'!$B$14)*'03 Assumptions'!$B$38)"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A7']; c.value = '  cash sweep'
c.font = Font(size=10.0)
c = ws['G7']; c.value = '=-MIN(G5+G6,G34)'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['H7']; c.value = '=-MIN(H5+H6,H34)'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['I7']; c.value = '=-MIN(I5+I6,I34)'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['J7']; c.value = '=-MIN(J5+J6,J34)'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['K7']; c.value = '=-MIN(K5+K6,K34)'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A8']; c.value = 'TLB-EUR ending'
c.font = Font(size=10.0)
c = ws['G8']; c.value = '=G5+G6+G7'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['H8']; c.value = '=H5+H6+H7'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['I8']; c.value = '=I5+I6+I7'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['J8']; c.value = '=J5+J6+J7'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['K8']; c.value = '=K5+K6+K7'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A10']; c.value = 'TLB-USD beginning'
c.font = Font(size=10.0)
c = ws['G10']; c.value = "='06 Sources & Uses'!B15"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF008000', size=10.0)
c = ws['H10']; c.value = '=G13'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['I10']; c.value = '=H13'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['J10']; c.value = '=I13'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['K10']; c.value = '=J13'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A11']; c.value = '  mandatory amort'
c.font = Font(size=10.0)
c = ws['G11']; c.value = "=-MIN(G10,('06 Sources & Uses'!$B$15)*'03 Assumptions'!$B$38)"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['H11']; c.value = "=-MIN(H10,('06 Sources & Uses'!$B$15)*'03 Assumptions'!$B$38)"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['I11']; c.value = "=-MIN(I10,('06 Sources & Uses'!$B$15)*'03 Assumptions'!$B$38)"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['J11']; c.value = "=-MIN(J10,('06 Sources & Uses'!$B$15)*'03 Assumptions'!$B$38)"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['K11']; c.value = "=-MIN(K10,('06 Sources & Uses'!$B$15)*'03 Assumptions'!$B$38)"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A12']; c.value = '  cash sweep (after EUR repaid)'
c.font = Font(size=10.0)
c = ws['G12']; c.value = '=-MIN(G10+G11,G34+G7)'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['H12']; c.value = '=-MIN(H10+H11,H34+H7)'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['I12']; c.value = '=-MIN(I10+I11,I34+I7)'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['J12']; c.value = '=-MIN(J10+J11,J34+J7)'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['K12']; c.value = '=-MIN(K10+K11,K34+K7)'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A13']; c.value = 'TLB-USD ending'
c.font = Font(size=10.0)
c = ws['G13']; c.value = '=G10+G11+G12'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['H13']; c.value = '=H10+H11+H12'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['I13']; c.value = '=I10+I11+I12'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['J13']; c.value = '=J10+J11+J12'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['K13']; c.value = '=K10+K11+K12'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A15']; c.value = 'SSN (bullet)'
c.font = Font(size=10.0)
c = ws['G15']; c.value = "='06 Sources & Uses'!B16"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF008000', size=10.0)
c = ws['H15']; c.value = "='06 Sources & Uses'!B16"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF008000', size=10.0)
c = ws['I15']; c.value = "='06 Sources & Uses'!B16"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF008000', size=10.0)
c = ws['J15']; c.value = "='06 Sources & Uses'!B16"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF008000', size=10.0)
c = ws['K15']; c.value = "='06 Sources & Uses'!B16"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF008000', size=10.0)
c = ws['A17']; c.value = 'Total debt ending'
c.font = Font(bold=True, color='FF1F3864')
c = ws['G17']; c.value = '=G8+G13+G15'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['H17']; c.value = '=H8+H13+H15'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['I17']; c.value = '=I8+I13+I15'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['J17']; c.value = '=J8+J13+J15'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['K17']; c.value = '=K8+K13+K15'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A18']; c.value = 'Excess cash (accumulates once TLBs repaid)'
c.font = Font(size=10.0)
c = ws['G18']; c.value = '=MAX(0,G34+G7+G12)'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['H18']; c.value = '=MAX(0,H34+H7+H12)+G18'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['I18']; c.value = '=MAX(0,I34+I7+I12)+H18'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['J18']; c.value = '=MAX(0,J34+J7+J12)+I18'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['K18']; c.value = '=MAX(0,K34+K7+K12)+J18'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A19']; c.value = 'NET DEBT ending'
c.font = Font(bold=True, color='FF1F3864')
c = ws['G19']; c.value = "=G17-G18-'03 Assumptions'!$B$44"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['H19']; c.value = "=H17-H18-'03 Assumptions'!$B$44"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['I19']; c.value = "=I17-I18-'03 Assumptions'!$B$44"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['J19']; c.value = "=J17-J18-'03 Assumptions'!$B$44"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['K19']; c.value = "=K17-K18-'03 Assumptions'!$B$44"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A21']; c.value = 'Interest: TLB-EUR (E+375 on beginning)'
c.font = Font(size=10.0)
c = ws['G21']; c.value = "=-G5*('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['H21']; c.value = "=-H5*('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['I21']; c.value = "=-I5*('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['J21']; c.value = "=-J5*('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['K21']; c.value = "=-K5*('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A22']; c.value = 'Interest: TLB-USD'
c.font = Font(size=10.0)
c = ws['G22']; c.value = "=-G10*'03 Assumptions'!$B$28"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['H22']; c.value = "=-H10*'03 Assumptions'!$B$28"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['I22']; c.value = "=-I10*'03 Assumptions'!$B$28"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['J22']; c.value = "=-J10*'03 Assumptions'!$B$28"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['K22']; c.value = "=-K10*'03 Assumptions'!$B$28"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A23']; c.value = 'Interest: SSN'
c.font = Font(size=10.0)
c = ws['G23']; c.value = "=-G15*'03 Assumptions'!$B$29"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['H23']; c.value = "=-H15*'03 Assumptions'!$B$29"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['I23']; c.value = "=-I15*'03 Assumptions'!$B$29"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['J23']; c.value = "=-J15*'03 Assumptions'!$B$29"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['K23']; c.value = "=-K15*'03 Assumptions'!$B$29"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A24']; c.value = 'RCF commitment fee'
c.font = Font(size=10.0)
c = ws['G24']; c.value = "=-'03 Assumptions'!$B$30*'03 Assumptions'!$B$31"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['H24']; c.value = "=-'03 Assumptions'!$B$30*'03 Assumptions'!$B$31"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['I24']; c.value = "=-'03 Assumptions'!$B$30*'03 Assumptions'!$B$31"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['J24']; c.value = "=-'03 Assumptions'!$B$30*'03 Assumptions'!$B$31"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['K24']; c.value = "=-'03 Assumptions'!$B$30*'03 Assumptions'!$B$31"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A25']; c.value = 'Total cash interest'
c.font = Font(bold=True, color='FF1F3864')
c = ws['G25']; c.value = '=SUM(G21:G24)'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['H25']; c.value = '=SUM(H21:H24)'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['I25']; c.value = '=SUM(I21:I24)'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['J25']; c.value = '=SUM(J21:J24)'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['K25']; c.value = '=SUM(K21:K24)'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A26']; c.value = '(-) BD / in-licensing spend (funded pre-sweep)'
c.font = Font(size=10.0)
c = ws['G26']; c.value = "=-'03 Assumptions'!$F$13"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF008000', size=10.0)
c = ws['H26']; c.value = "=-'03 Assumptions'!$F$13"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF008000', size=10.0)
c = ws['I26']; c.value = "=-'03 Assumptions'!$F$13"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF008000', size=10.0)
c = ws['J26']; c.value = "=-'03 Assumptions'!$F$13"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF008000', size=10.0)
c = ws['K26']; c.value = 0
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF008000', size=10.0)
c = ws['A27']; c.value = 'EBITDA'
c.font = Font(size=10.0)
c = ws['G27']; c.value = "='09 Operating Model'!G18"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF008000', size=10.0)
c = ws['H27']; c.value = "='09 Operating Model'!H18"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF008000', size=10.0)
c = ws['I27']; c.value = "='09 Operating Model'!I18"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF008000', size=10.0)
c = ws['J27']; c.value = "='09 Operating Model'!J18"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF008000', size=10.0)
c = ws['K27']; c.value = "='09 Operating Model'!K18"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF008000', size=10.0)
c = ws['A28']; c.value = '(-) Maintenance capex'
c.font = Font(size=10.0)
c = ws['G28']; c.value = "='09 Operating Model'!G12"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF008000', size=10.0)
c = ws['H28']; c.value = "='09 Operating Model'!H12"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF008000', size=10.0)
c = ws['I28']; c.value = "='09 Operating Model'!I12"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF008000', size=10.0)
c = ws['J28']; c.value = "='09 Operating Model'!J12"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF008000', size=10.0)
c = ws['K28']; c.value = "='09 Operating Model'!K12"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF008000', size=10.0)
c = ws['A29']; c.value = '(-) Change in NWC'
c.font = Font(size=10.0)
c = ws['G29']; c.value = "='09 Operating Model'!G13"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF008000', size=10.0)
c = ws['H29']; c.value = "='09 Operating Model'!H13"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF008000', size=10.0)
c = ws['I29']; c.value = "='09 Operating Model'!I13"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF008000', size=10.0)
c = ws['J29']; c.value = "='09 Operating Model'!J13"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF008000', size=10.0)
c = ws['K29']; c.value = "='09 Operating Model'!K13"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF008000', size=10.0)
c = ws['A30']; c.value = '(-) Cash taxes (24% of max(0, EBITDA - D&A - interest))'
c.font = Font(size=10.0)
c = ws['G30']; c.value = "=-'03 Assumptions'!$B$34*MAX(0,G27+'09 Operating Model'!G10-MIN(-G25,'03 Assumptions'!$B$45*G27))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['H30']; c.value = "=-'03 Assumptions'!$B$34*MAX(0,H27+'09 Operating Model'!H10-MIN(-H25,'03 Assumptions'!$B$45*H27))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['I30']; c.value = "=-'03 Assumptions'!$B$34*MAX(0,I27+'09 Operating Model'!I10-MIN(-I25,'03 Assumptions'!$B$45*I27))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['J30']; c.value = "=-'03 Assumptions'!$B$34*MAX(0,J27+'09 Operating Model'!J10-MIN(-J25,'03 Assumptions'!$B$45*J27))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['K30']; c.value = "=-'03 Assumptions'!$B$34*MAX(0,K27+'09 Operating Model'!K10-MIN(-K25,'03 Assumptions'!$B$45*K27))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A31']; c.value = '(-) Cash interest & fees'
c.font = Font(size=10.0)
c = ws['G31']; c.value = '=G25'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['H31']; c.value = '=H25'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['I31']; c.value = '=I25'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['J31']; c.value = '=J25'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['K31']; c.value = '=K25'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A32']; c.value = 'FCF before debt service'
c.font = Font(bold=True, color='FF1F3864')
c = ws['G32']; c.value = '=G27+G28+G29+G30+G31+G26'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['H32']; c.value = '=H27+H28+H29+H30+H31+H26'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['I32']; c.value = '=I27+I28+I29+I30+I31+I26'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['J32']; c.value = '=J27+J28+J29+J30+J31+J26'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['K32']; c.value = '=K27+K28+K29+K30+K31+K26'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A33']; c.value = '(-) Mandatory amortisation'
c.font = Font(size=10.0)
c = ws['G33']; c.value = '=G6+G11'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['H33']; c.value = '=H6+H11'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['I33']; c.value = '=I6+I11'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['J33']; c.value = '=J6+J11'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['K33']; c.value = '=K6+K11'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A34']; c.value = 'FCF available for sweep'
c.font = Font(size=10.0)
c = ws['G34']; c.value = "=MAX(0,G32+G33)*'03 Assumptions'!$F$12"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['H34']; c.value = "=MAX(0,H32+H33)*'03 Assumptions'!$F$12"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['I34']; c.value = "=MAX(0,I32+I33)*'03 Assumptions'!$F$12"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['J34']; c.value = "=MAX(0,J32+J33)*'03 Assumptions'!$F$12"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['K34']; c.value = "=MAX(0,K32+K33)*'03 Assumptions'!$F$12"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A36']; c.value = 'Leverage: net debt / EBITDA'
c.font = Font(bold=True, color='FF1F3864')
c = ws['G36']; c.value = '=G19/G27'
c.number_format = '0.00#'
c.font = Font(size=10.0)
c = ws['H36']; c.value = '=H19/H27'
c.number_format = '0.00#'
c.font = Font(size=10.0)
c = ws['I36']; c.value = '=I19/I27'
c.number_format = '0.00#'
c.font = Font(size=10.0)
c = ws['J36']; c.value = '=J19/J27'
c.number_format = '0.00#'
c.font = Font(size=10.0)
c = ws['K36']; c.value = '=K19/K27'
c.number_format = '0.00#'
c.font = Font(size=10.0)
c = ws['A37']; c.value = 'Interest coverage: EBITDA / cash interest'
c.font = Font(bold=True, color='FF1F3864')
c = ws['G37']; c.value = '=-G27/G25'
c.number_format = '0.00#'
c.font = Font(size=10.0)
c = ws['H37']; c.value = '=-H27/H25'
c.number_format = '0.00#'
c.font = Font(size=10.0)
c = ws['I37']; c.value = '=-I27/I25'
c.number_format = '0.00#'
c.font = Font(size=10.0)
c = ws['J37']; c.value = '=-J27/J25'
c.number_format = '0.00#'
c.font = Font(size=10.0)
c = ws['K37']; c.value = '=-K27/K25'
c.number_format = '0.00#'
c.font = Font(size=10.0)
c = ws['A38']; c.value = 'DSCR: (EBITDA-capex-tax)/(interest+amort)'
c.font = Font(size=10.0)
c = ws['G38']; c.value = '=(G27+G28+G30)/-(G25-G33+0.0001)'
c.number_format = '0.00#'
c.font = Font(size=10.0)
c = ws['H38']; c.value = '=(H27+H28+H30)/-(H25-H33+0.0001)'
c.number_format = '0.00#'
c.font = Font(size=10.0)
c = ws['I38']; c.value = '=(I27+I28+I30)/-(I25-I33+0.0001)'
c.number_format = '0.00#'
c.font = Font(size=10.0)
c = ws['J38']; c.value = '=(J27+J28+J30)/-(J25-J33+0.0001)'
c.number_format = '0.00#'
c.font = Font(size=10.0)
c = ws['K38']; c.value = '=(K27+K28+K30)/-(K25-K33+0.0001)'
c.number_format = '0.00#'
c.font = Font(size=10.0)
c = ws['A40']; c.value = '=IF(G37>2.5,"CHECK OK: year-1 interest coverage > 2.5x","CHECK: coverage tight — review leverage")'
c.font = Font(bold=True, color='FFFF0000', size=10.0)
c = ws['A42']; c.value = 'Sweep order: TLB-EUR first, then TLB-USD; SSN bullet. Net debt nets EUR 250m minimum cash + swept excess.'
c.font = Font(size=10.0)
c = ws['A43']; c.value = 'v2: BD spend deducted pre-sweep; taxes apply ATAD 30%-of-EBITDA interest deductibility cap (carryforward ignored - conservative).'
c.font = Font(size=10.0)
c = ws['A44']; c.value = 'v3: 2031 (exit-year) BD spend set to EUR 0 — it buys EBITDA only in 2032 (post-exit). Cum. BD 2027-2030 = EUR 1.6bn.'

ws = wb.create_sheet('10 Returns')
ws.column_dimensions['A'].width = 46.0
ws.column_dimensions['B'].width = 12.0
c = ws['A1']; c.value = 'RETURNS — IRR / MOIC + BRIDGES'
c.font = Font(bold=True, color='FFFFFFFF', size=12.0)
c.fill = PatternFill("solid", fgColor='FF1F3864')
c = ws['A3']; c.value = 'Entry equity (from S&U)'
c.font = Font(size=10.0)
c = ws['B3']; c.value = "='06 Sources & Uses'!B18"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF008000', size=10.0)
c = ws['A4']; c.value = 'Exit year EBITDA (2031, incl. BD)'
c.font = Font(size=10.0)
c = ws['B4']; c.value = "='09 Operating Model'!K18"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF008000', size=10.0)
c = ws['A5']; c.value = 'Exit multiple (scenario)'
c.font = Font(size=10.0)
c = ws['B5']; c.value = "='03 Assumptions'!F11"
c.number_format = '0.00#'
c.font = Font(color='FF008000', size=10.0)
c = ws['A6']; c.value = 'Exit enterprise value'
c.font = Font(size=10.0)
c = ws['B6']; c.value = '=B4*B5'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A7']; c.value = '(-) Net debt at exit'
c.font = Font(size=10.0)
c = ws['B7']; c.value = "='08 Debt Schedule'!K19"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF008000', size=10.0)
c = ws['A8']; c.value = 'Exit equity value'
c.font = Font(size=10.0)
c = ws['B8']; c.value = '=B6-B7'
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A9']; c.value = 'MOIC'
c.font = Font(size=10.0)
c = ws['B9']; c.value = '=B8/B3'
c.number_format = '0.00#'
c.font = Font(size=10.0)
c.fill = PatternFill("solid", fgColor='FFFFFF00')
c = ws['A10']; c.value = 'Hold (years)'
c.font = Font(size=10.0)
c = ws['B10']; c.value = "='03 Assumptions'!B42"
c.number_format = '#,##0;\\(#,##0\\);\\-'
c.font = Font(color='FF008000', size=10.0)
c = ws['A11']; c.value = 'IRR (single exit flow)'
c.font = Font(size=10.0)
c = ws['B11']; c.value = '=B9^(1/B10)-1'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c.fill = PatternFill("solid", fgColor='FFFFFF00')
c = ws['A14']; c.value = 'MOIC BRIDGE (multiples of entry equity)'
c.font = Font(bold=True, color='FF1F3864')
c = ws['A15']; c.value = 'Entry equity (1.00x)'
c.font = Font(size=10.0)
c = ws['B15']; c.value = 1
c.number_format = '0.00#'
c.font = Font(size=10.0)
c = ws['A16']; c.value = '(+) EBITDA growth at entry multiple'
c.font = Font(size=10.0)
c = ws['B16']; c.value = "=(B4-'03 Assumptions'!B16)*'07 Purchase Price'!B8/B3"
c.number_format = '0.00#'
c.font = Font(size=10.0)
c = ws['A17']; c.value = '(+) Multiple expansion/(compression)'
c.font = Font(size=10.0)
c = ws['B17']; c.value = "=(B5-'07 Purchase Price'!B8)*B4/B3"
c.number_format = '0.00#'
c.font = Font(size=10.0)
c = ws['A18']; c.value = '(+) Deleveraging (net debt entry - exit)'
c.font = Font(size=10.0)
c = ws['B18']; c.value = "=('06 Sources & Uses'!B13-'03 Assumptions'!B44-'08 Debt Schedule'!K19)/B3"
c.number_format = '0.00#'
c.font = Font(size=10.0)
c = ws['A19']; c.value = '(-) Fees & frictions (residual)'
c.font = Font(size=10.0)
c = ws['B19']; c.value = '=B9-B15-B16-B17-B18'
c.number_format = '0.00#'
c.font = Font(size=10.0)
c = ws['A20']; c.value = 'MOIC (check)'
c.font = Font(bold=True, color='FF1F3864')
c = ws['B20']; c.value = '=SUM(B15:B19)'
c.number_format = '0.00#'
c.font = Font(size=10.0)
c = ws['A21']; c.value = '=IF(ABS(B20-B9)<0.01,"CHECK OK: bridge ties to MOIC","ERROR: bridge broken")'
c.font = Font(bold=True, color='FFFF0000', size=10.0)
c = ws['A23']; c.value = 'IRR / MOIC BY EXIT YEAR (live)'
c.font = Font(bold=True, color='FF1F3864')
c = ws['A24']; c.value = 'Exit year'
c.font = Font(size=10.0)
c = ws['B24']; c.value = '2029'
c.font = Font(size=10.0)
c = ws['C24']; c.value = '2030'
c.font = Font(size=10.0)
c = ws['D24']; c.value = '2031'
c.font = Font(size=10.0)
c = ws['A25']; c.value = 'Exit equity (mult x EBITDA - net debt)'
c.font = Font(size=10.0)
c = ws['B25']; c.value = "='03 Assumptions'!F11*'09 Operating Model'!I18-'08 Debt Schedule'!I19"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['C25']; c.value = "='03 Assumptions'!F11*'09 Operating Model'!J18-'08 Debt Schedule'!J19"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['D25']; c.value = "='03 Assumptions'!F11*'09 Operating Model'!K18-'08 Debt Schedule'!K19"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A26']; c.value = 'MOIC'
c.font = Font(size=10.0)
c = ws['B26']; c.value = "=B25/'06 Sources & Uses'!B18"
c.number_format = '0.00#'
c.font = Font(size=10.0)
c = ws['C26']; c.value = "=C25/'06 Sources & Uses'!B18"
c.number_format = '0.00#'
c.font = Font(size=10.0)
c = ws['D26']; c.value = "=D25/'06 Sources & Uses'!B18"
c.number_format = '0.00#'
c.font = Font(size=10.0)
c = ws['A27']; c.value = 'IRR'
c.font = Font(size=10.0)
c = ws['B27']; c.value = "=(B25/'06 Sources & Uses'!B18)^(1/3)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['C27']; c.value = "=(C25/'06 Sources & Uses'!B18)^(1/4)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['D27']; c.value = "=(D25/'06 Sources & Uses'!B18)^(1/5)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['A29']; c.value = 'No interim dividends/recaps modeled (conservative). Dividend recap optionality would raise IRR, not MOIC.'
c.font = Font(size=10.0)
c = ws['A31']; c.value = 'RECAP OPTIONALITY: Sponsor-consistent case includes a dividend recap to ~4.0x leverage when headroom allows (~Yr4/5), adding ~1.0pt IRR.'
c = ws['A32']; c.value = 'Modeled here as conservative 100% sweep to ~1.0x exit leverage. Extending the hold to 7 years requires modeling this recap AND interest income on excess cash first.'

ws = wb.create_sheet('11 Sensitivities')
ws.column_dimensions['A'].width = 46.0
ws.column_dimensions['B'].width = 12.0
c = ws['A1']; c.value = 'SENSITIVITIES - EXACT full-recursion engines (no approximations)'
c.font = Font(bold=True, color='FFFFFFFF', size=12.0)
c.fill = PatternFill("solid", fgColor='FF1F3864')
c = ws['A2']; c.value = 'All grids run a complete 3-tranche debt recursion (interest, ATAD tax cap, BD spend, mandatory amort, sequential sweep,'
c.font = Font(size=10.0)
c = ws['A3']; c.value = 'excess cash) per case. Engine A scales operating items with EBITDA (axis definition); Engine B uses the live EBITDA path.'
c.font = Font(size=10.0)
c = ws['A5']; c.value = 'GRID 1 - Exit multiple (rows) x exit year (cols) - IRR, fully live'
c.font = Font(bold=True, color='FF1F3864')
c = ws['C6']; c.value = '2029'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['D6']; c.value = '2030'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['E6']; c.value = '2031'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['B7']; c.value = 9
c.number_format = '0.00#'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['C7']; c.value = "=(($B7*'09 Operating Model'!$I$18-'08 Debt Schedule'!$I$19)/'06 Sources & Uses'!$B$18)^(1/3)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['D7']; c.value = "=(($B7*'09 Operating Model'!$J$18-'08 Debt Schedule'!$J$19)/'06 Sources & Uses'!$B$18)^(1/4)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['E7']; c.value = "=(($B7*'09 Operating Model'!$K$18-'08 Debt Schedule'!$K$19)/'06 Sources & Uses'!$B$18)^(1/5)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['B8']; c.value = 10
c.number_format = '0.00#'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['C8']; c.value = "=(($B8*'09 Operating Model'!$I$18-'08 Debt Schedule'!$I$19)/'06 Sources & Uses'!$B$18)^(1/3)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['D8']; c.value = "=(($B8*'09 Operating Model'!$J$18-'08 Debt Schedule'!$J$19)/'06 Sources & Uses'!$B$18)^(1/4)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['E8']; c.value = "=(($B8*'09 Operating Model'!$K$18-'08 Debt Schedule'!$K$19)/'06 Sources & Uses'!$B$18)^(1/5)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['B9']; c.value = 11
c.number_format = '0.00#'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['C9']; c.value = "=(($B9*'09 Operating Model'!$I$18-'08 Debt Schedule'!$I$19)/'06 Sources & Uses'!$B$18)^(1/3)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['D9']; c.value = "=(($B9*'09 Operating Model'!$J$18-'08 Debt Schedule'!$J$19)/'06 Sources & Uses'!$B$18)^(1/4)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['E9']; c.value = "=(($B9*'09 Operating Model'!$K$18-'08 Debt Schedule'!$K$19)/'06 Sources & Uses'!$B$18)^(1/5)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['B10']; c.value = 11.5
c.number_format = '0.00#'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['C10']; c.value = "=(($B10*'09 Operating Model'!$I$18-'08 Debt Schedule'!$I$19)/'06 Sources & Uses'!$B$18)^(1/3)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['D10']; c.value = "=(($B10*'09 Operating Model'!$J$18-'08 Debt Schedule'!$J$19)/'06 Sources & Uses'!$B$18)^(1/4)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['E10']; c.value = "=(($B10*'09 Operating Model'!$K$18-'08 Debt Schedule'!$K$19)/'06 Sources & Uses'!$B$18)^(1/5)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['B11']; c.value = 12
c.number_format = '0.00#'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['C11']; c.value = "=(($B11*'09 Operating Model'!$I$18-'08 Debt Schedule'!$I$19)/'06 Sources & Uses'!$B$18)^(1/3)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['D11']; c.value = "=(($B11*'09 Operating Model'!$J$18-'08 Debt Schedule'!$J$19)/'06 Sources & Uses'!$B$18)^(1/4)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['E11']; c.value = "=(($B11*'09 Operating Model'!$K$18-'08 Debt Schedule'!$K$19)/'06 Sources & Uses'!$B$18)^(1/5)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['B12']; c.value = 12.9
c.number_format = '0.00#'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['C12']; c.value = "=(($B12*'09 Operating Model'!$I$18-'08 Debt Schedule'!$I$19)/'06 Sources & Uses'!$B$18)^(1/3)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['D12']; c.value = "=(($B12*'09 Operating Model'!$J$18-'08 Debt Schedule'!$J$19)/'06 Sources & Uses'!$B$18)^(1/4)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['E12']; c.value = "=(($B12*'09 Operating Model'!$K$18-'08 Debt Schedule'!$K$19)/'06 Sources & Uses'!$B$18)^(1/5)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['B13']; c.value = 14
c.number_format = '0.00#'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['C13']; c.value = "=(($B13*'09 Operating Model'!$I$18-'08 Debt Schedule'!$I$19)/'06 Sources & Uses'!$B$18)^(1/3)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['D13']; c.value = "=(($B13*'09 Operating Model'!$J$18-'08 Debt Schedule'!$J$19)/'06 Sources & Uses'!$B$18)^(1/4)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['E13']; c.value = "=(($B13*'09 Operating Model'!$K$18-'08 Debt Schedule'!$K$19)/'06 Sources & Uses'!$B$18)^(1/5)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['A16']; c.value = 'ENGINE A - total-EBITDA CAGR cases (cols C-G). Full recursion below; results in GRID 2.'
c.font = Font(bold=True, color='FF1F3864')
c = ws['A17']; c.value = 'EBITDA CAGR 2026-31'
c.font = Font(size=10.0)
c = ws['C17']; c.value = 0.02
c.number_format = '0.0%'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['D17']; c.value = 0.04
c.number_format = '0.0%'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['E17']; c.value = 0.06
c.number_format = '0.0%'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['F17']; c.value = 0.08
c.number_format = '0.0%'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['G17']; c.value = 0.1
c.number_format = '0.0%'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['A21']; c.value = 'EBITDA year 2027'
c.font = Font(size=10.0)
c = ws['C21']; c.value = "='09 Operating Model'!$F$18*(1+C$17)^1"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['D21']; c.value = "='09 Operating Model'!$F$18*(1+D$17)^1"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['E21']; c.value = "='09 Operating Model'!$F$18*(1+E$17)^1"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['F21']; c.value = "='09 Operating Model'!$F$18*(1+F$17)^1"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['G21']; c.value = "='09 Operating Model'!$F$18*(1+G$17)^1"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A22']; c.value = 'EBITDA year 2028'
c.font = Font(size=10.0)
c = ws['C22']; c.value = "='09 Operating Model'!$F$18*(1+C$17)^2"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['D22']; c.value = "='09 Operating Model'!$F$18*(1+D$17)^2"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['E22']; c.value = "='09 Operating Model'!$F$18*(1+E$17)^2"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['F22']; c.value = "='09 Operating Model'!$F$18*(1+F$17)^2"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['G22']; c.value = "='09 Operating Model'!$F$18*(1+G$17)^2"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A23']; c.value = 'EBITDA year 2029'
c.font = Font(size=10.0)
c = ws['C23']; c.value = "='09 Operating Model'!$F$18*(1+C$17)^3"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['D23']; c.value = "='09 Operating Model'!$F$18*(1+D$17)^3"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['E23']; c.value = "='09 Operating Model'!$F$18*(1+E$17)^3"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['F23']; c.value = "='09 Operating Model'!$F$18*(1+F$17)^3"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['G23']; c.value = "='09 Operating Model'!$F$18*(1+G$17)^3"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A24']; c.value = 'EBITDA year 2030'
c.font = Font(size=10.0)
c = ws['C24']; c.value = "='09 Operating Model'!$F$18*(1+C$17)^4"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['D24']; c.value = "='09 Operating Model'!$F$18*(1+D$17)^4"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['E24']; c.value = "='09 Operating Model'!$F$18*(1+E$17)^4"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['F24']; c.value = "='09 Operating Model'!$F$18*(1+F$17)^4"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['G24']; c.value = "='09 Operating Model'!$F$18*(1+G$17)^4"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A25']; c.value = 'EBITDA year 2031'
c.font = Font(size=10.0)
c = ws['C25']; c.value = "='09 Operating Model'!$F$18*(1+C$17)^5"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['D25']; c.value = "='09 Operating Model'!$F$18*(1+D$17)^5"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['E25']; c.value = "='09 Operating Model'!$F$18*(1+E$17)^5"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['F25']; c.value = "='09 Operating Model'!$F$18*(1+F$17)^5"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['G25']; c.value = "='09 Operating Model'!$F$18*(1+G$17)^5"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A28']; c.value = 'FCF pre-amort 2027'
c.font = Font(size=10.0)
c = ws['C28']; c.value = "=C21+(('08 Debt Schedule'!G28+'08 Debt Schedule'!G29)*C21/'09 Operating Model'!G18)+(-'03 Assumptions'!$B$34*MAX(0,C21+('09 Operating Model'!G10*C21/'09 Operating Model'!G18)-MIN((('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*'06 Sources & Uses'!$B$14+'03 Assumptions'!$B$28*'06 Sources & Uses'!$B$15+'03 Assumptions'!$B$29*'06 Sources & Uses'!$B$16+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31),'03 Assumptions'!$B$45*C21)))-(('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*'06 Sources & Uses'!$B$14+'03 Assumptions'!$B$28*'06 Sources & Uses'!$B$15+'03 Assumptions'!$B$29*'06 Sources & Uses'!$B$16+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31)-'03 Assumptions'!$F$13"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['D28']; c.value = "=D21+(('08 Debt Schedule'!G28+'08 Debt Schedule'!G29)*D21/'09 Operating Model'!G18)+(-'03 Assumptions'!$B$34*MAX(0,D21+('09 Operating Model'!G10*D21/'09 Operating Model'!G18)-MIN((('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*'06 Sources & Uses'!$B$14+'03 Assumptions'!$B$28*'06 Sources & Uses'!$B$15+'03 Assumptions'!$B$29*'06 Sources & Uses'!$B$16+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31),'03 Assumptions'!$B$45*D21)))-(('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*'06 Sources & Uses'!$B$14+'03 Assumptions'!$B$28*'06 Sources & Uses'!$B$15+'03 Assumptions'!$B$29*'06 Sources & Uses'!$B$16+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31)-'03 Assumptions'!$F$13"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['E28']; c.value = "=E21+(('08 Debt Schedule'!G28+'08 Debt Schedule'!G29)*E21/'09 Operating Model'!G18)+(-'03 Assumptions'!$B$34*MAX(0,E21+('09 Operating Model'!G10*E21/'09 Operating Model'!G18)-MIN((('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*'06 Sources & Uses'!$B$14+'03 Assumptions'!$B$28*'06 Sources & Uses'!$B$15+'03 Assumptions'!$B$29*'06 Sources & Uses'!$B$16+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31),'03 Assumptions'!$B$45*E21)))-(('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*'06 Sources & Uses'!$B$14+'03 Assumptions'!$B$28*'06 Sources & Uses'!$B$15+'03 Assumptions'!$B$29*'06 Sources & Uses'!$B$16+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31)-'03 Assumptions'!$F$13"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['F28']; c.value = "=F21+(('08 Debt Schedule'!G28+'08 Debt Schedule'!G29)*F21/'09 Operating Model'!G18)+(-'03 Assumptions'!$B$34*MAX(0,F21+('09 Operating Model'!G10*F21/'09 Operating Model'!G18)-MIN((('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*'06 Sources & Uses'!$B$14+'03 Assumptions'!$B$28*'06 Sources & Uses'!$B$15+'03 Assumptions'!$B$29*'06 Sources & Uses'!$B$16+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31),'03 Assumptions'!$B$45*F21)))-(('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*'06 Sources & Uses'!$B$14+'03 Assumptions'!$B$28*'06 Sources & Uses'!$B$15+'03 Assumptions'!$B$29*'06 Sources & Uses'!$B$16+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31)-'03 Assumptions'!$F$13"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['G28']; c.value = "=G21+(('08 Debt Schedule'!G28+'08 Debt Schedule'!G29)*G21/'09 Operating Model'!G18)+(-'03 Assumptions'!$B$34*MAX(0,G21+('09 Operating Model'!G10*G21/'09 Operating Model'!G18)-MIN((('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*'06 Sources & Uses'!$B$14+'03 Assumptions'!$B$28*'06 Sources & Uses'!$B$15+'03 Assumptions'!$B$29*'06 Sources & Uses'!$B$16+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31),'03 Assumptions'!$B$45*G21)))-(('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*'06 Sources & Uses'!$B$14+'03 Assumptions'!$B$28*'06 Sources & Uses'!$B$15+'03 Assumptions'!$B$29*'06 Sources & Uses'!$B$16+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31)-'03 Assumptions'!$F$13"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A29']; c.value = 'FCF pre-amort 2028'
c.font = Font(size=10.0)
c = ws['C29']; c.value = "=C22+(('08 Debt Schedule'!H28+'08 Debt Schedule'!H29)*C22/'09 Operating Model'!H18)+(-'03 Assumptions'!$B$34*MAX(0,C22+('09 Operating Model'!H10*C22/'09 Operating Model'!H18)-MIN((('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*C35+'03 Assumptions'!$B$28*C42+'03 Assumptions'!$B$29*'06 Sources & Uses'!$B$16+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31),'03 Assumptions'!$B$45*C22)))-(('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*C35+'03 Assumptions'!$B$28*C42+'03 Assumptions'!$B$29*'06 Sources & Uses'!$B$16+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31)-'03 Assumptions'!$F$13"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['D29']; c.value = "=D22+(('08 Debt Schedule'!H28+'08 Debt Schedule'!H29)*D22/'09 Operating Model'!H18)+(-'03 Assumptions'!$B$34*MAX(0,D22+('09 Operating Model'!H10*D22/'09 Operating Model'!H18)-MIN((('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*D35+'03 Assumptions'!$B$28*D42+'03 Assumptions'!$B$29*'06 Sources & Uses'!$B$16+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31),'03 Assumptions'!$B$45*D22)))-(('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*D35+'03 Assumptions'!$B$28*D42+'03 Assumptions'!$B$29*'06 Sources & Uses'!$B$16+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31)-'03 Assumptions'!$F$13"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['E29']; c.value = "=E22+(('08 Debt Schedule'!H28+'08 Debt Schedule'!H29)*E22/'09 Operating Model'!H18)+(-'03 Assumptions'!$B$34*MAX(0,E22+('09 Operating Model'!H10*E22/'09 Operating Model'!H18)-MIN((('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*E35+'03 Assumptions'!$B$28*E42+'03 Assumptions'!$B$29*'06 Sources & Uses'!$B$16+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31),'03 Assumptions'!$B$45*E22)))-(('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*E35+'03 Assumptions'!$B$28*E42+'03 Assumptions'!$B$29*'06 Sources & Uses'!$B$16+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31)-'03 Assumptions'!$F$13"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['F29']; c.value = "=F22+(('08 Debt Schedule'!H28+'08 Debt Schedule'!H29)*F22/'09 Operating Model'!H18)+(-'03 Assumptions'!$B$34*MAX(0,F22+('09 Operating Model'!H10*F22/'09 Operating Model'!H18)-MIN((('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*F35+'03 Assumptions'!$B$28*F42+'03 Assumptions'!$B$29*'06 Sources & Uses'!$B$16+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31),'03 Assumptions'!$B$45*F22)))-(('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*F35+'03 Assumptions'!$B$28*F42+'03 Assumptions'!$B$29*'06 Sources & Uses'!$B$16+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31)-'03 Assumptions'!$F$13"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['G29']; c.value = "=G22+(('08 Debt Schedule'!H28+'08 Debt Schedule'!H29)*G22/'09 Operating Model'!H18)+(-'03 Assumptions'!$B$34*MAX(0,G22+('09 Operating Model'!H10*G22/'09 Operating Model'!H18)-MIN((('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*G35+'03 Assumptions'!$B$28*G42+'03 Assumptions'!$B$29*'06 Sources & Uses'!$B$16+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31),'03 Assumptions'!$B$45*G22)))-(('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*G35+'03 Assumptions'!$B$28*G42+'03 Assumptions'!$B$29*'06 Sources & Uses'!$B$16+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31)-'03 Assumptions'!$F$13"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A30']; c.value = 'FCF pre-amort 2029'
c.font = Font(size=10.0)
c = ws['C30']; c.value = "=C23+(('08 Debt Schedule'!I28+'08 Debt Schedule'!I29)*C23/'09 Operating Model'!I18)+(-'03 Assumptions'!$B$34*MAX(0,C23+('09 Operating Model'!I10*C23/'09 Operating Model'!I18)-MIN((('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*C36+'03 Assumptions'!$B$28*C43+'03 Assumptions'!$B$29*'06 Sources & Uses'!$B$16+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31),'03 Assumptions'!$B$45*C23)))-(('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*C36+'03 Assumptions'!$B$28*C43+'03 Assumptions'!$B$29*'06 Sources & Uses'!$B$16+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31)-'03 Assumptions'!$F$13"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['D30']; c.value = "=D23+(('08 Debt Schedule'!I28+'08 Debt Schedule'!I29)*D23/'09 Operating Model'!I18)+(-'03 Assumptions'!$B$34*MAX(0,D23+('09 Operating Model'!I10*D23/'09 Operating Model'!I18)-MIN((('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*D36+'03 Assumptions'!$B$28*D43+'03 Assumptions'!$B$29*'06 Sources & Uses'!$B$16+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31),'03 Assumptions'!$B$45*D23)))-(('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*D36+'03 Assumptions'!$B$28*D43+'03 Assumptions'!$B$29*'06 Sources & Uses'!$B$16+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31)-'03 Assumptions'!$F$13"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['E30']; c.value = "=E23+(('08 Debt Schedule'!I28+'08 Debt Schedule'!I29)*E23/'09 Operating Model'!I18)+(-'03 Assumptions'!$B$34*MAX(0,E23+('09 Operating Model'!I10*E23/'09 Operating Model'!I18)-MIN((('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*E36+'03 Assumptions'!$B$28*E43+'03 Assumptions'!$B$29*'06 Sources & Uses'!$B$16+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31),'03 Assumptions'!$B$45*E23)))-(('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*E36+'03 Assumptions'!$B$28*E43+'03 Assumptions'!$B$29*'06 Sources & Uses'!$B$16+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31)-'03 Assumptions'!$F$13"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['F30']; c.value = "=F23+(('08 Debt Schedule'!I28+'08 Debt Schedule'!I29)*F23/'09 Operating Model'!I18)+(-'03 Assumptions'!$B$34*MAX(0,F23+('09 Operating Model'!I10*F23/'09 Operating Model'!I18)-MIN((('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*F36+'03 Assumptions'!$B$28*F43+'03 Assumptions'!$B$29*'06 Sources & Uses'!$B$16+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31),'03 Assumptions'!$B$45*F23)))-(('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*F36+'03 Assumptions'!$B$28*F43+'03 Assumptions'!$B$29*'06 Sources & Uses'!$B$16+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31)-'03 Assumptions'!$F$13"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['G30']; c.value = "=G23+(('08 Debt Schedule'!I28+'08 Debt Schedule'!I29)*G23/'09 Operating Model'!I18)+(-'03 Assumptions'!$B$34*MAX(0,G23+('09 Operating Model'!I10*G23/'09 Operating Model'!I18)-MIN((('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*G36+'03 Assumptions'!$B$28*G43+'03 Assumptions'!$B$29*'06 Sources & Uses'!$B$16+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31),'03 Assumptions'!$B$45*G23)))-(('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*G36+'03 Assumptions'!$B$28*G43+'03 Assumptions'!$B$29*'06 Sources & Uses'!$B$16+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31)-'03 Assumptions'!$F$13"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A31']; c.value = 'FCF pre-amort 2030'
c.font = Font(size=10.0)
c = ws['C31']; c.value = "=C24+(('08 Debt Schedule'!J28+'08 Debt Schedule'!J29)*C24/'09 Operating Model'!J18)+(-'03 Assumptions'!$B$34*MAX(0,C24+('09 Operating Model'!J10*C24/'09 Operating Model'!J18)-MIN((('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*C37+'03 Assumptions'!$B$28*C44+'03 Assumptions'!$B$29*'06 Sources & Uses'!$B$16+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31),'03 Assumptions'!$B$45*C24)))-(('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*C37+'03 Assumptions'!$B$28*C44+'03 Assumptions'!$B$29*'06 Sources & Uses'!$B$16+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31)-'03 Assumptions'!$F$13"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['D31']; c.value = "=D24+(('08 Debt Schedule'!J28+'08 Debt Schedule'!J29)*D24/'09 Operating Model'!J18)+(-'03 Assumptions'!$B$34*MAX(0,D24+('09 Operating Model'!J10*D24/'09 Operating Model'!J18)-MIN((('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*D37+'03 Assumptions'!$B$28*D44+'03 Assumptions'!$B$29*'06 Sources & Uses'!$B$16+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31),'03 Assumptions'!$B$45*D24)))-(('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*D37+'03 Assumptions'!$B$28*D44+'03 Assumptions'!$B$29*'06 Sources & Uses'!$B$16+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31)-'03 Assumptions'!$F$13"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['E31']; c.value = "=E24+(('08 Debt Schedule'!J28+'08 Debt Schedule'!J29)*E24/'09 Operating Model'!J18)+(-'03 Assumptions'!$B$34*MAX(0,E24+('09 Operating Model'!J10*E24/'09 Operating Model'!J18)-MIN((('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*E37+'03 Assumptions'!$B$28*E44+'03 Assumptions'!$B$29*'06 Sources & Uses'!$B$16+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31),'03 Assumptions'!$B$45*E24)))-(('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*E37+'03 Assumptions'!$B$28*E44+'03 Assumptions'!$B$29*'06 Sources & Uses'!$B$16+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31)-'03 Assumptions'!$F$13"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['F31']; c.value = "=F24+(('08 Debt Schedule'!J28+'08 Debt Schedule'!J29)*F24/'09 Operating Model'!J18)+(-'03 Assumptions'!$B$34*MAX(0,F24+('09 Operating Model'!J10*F24/'09 Operating Model'!J18)-MIN((('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*F37+'03 Assumptions'!$B$28*F44+'03 Assumptions'!$B$29*'06 Sources & Uses'!$B$16+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31),'03 Assumptions'!$B$45*F24)))-(('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*F37+'03 Assumptions'!$B$28*F44+'03 Assumptions'!$B$29*'06 Sources & Uses'!$B$16+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31)-'03 Assumptions'!$F$13"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['G31']; c.value = "=G24+(('08 Debt Schedule'!J28+'08 Debt Schedule'!J29)*G24/'09 Operating Model'!J18)+(-'03 Assumptions'!$B$34*MAX(0,G24+('09 Operating Model'!J10*G24/'09 Operating Model'!J18)-MIN((('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*G37+'03 Assumptions'!$B$28*G44+'03 Assumptions'!$B$29*'06 Sources & Uses'!$B$16+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31),'03 Assumptions'!$B$45*G24)))-(('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*G37+'03 Assumptions'!$B$28*G44+'03 Assumptions'!$B$29*'06 Sources & Uses'!$B$16+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31)-'03 Assumptions'!$F$13"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A32']; c.value = 'FCF pre-amort 2031'
c.font = Font(size=10.0)
c = ws['C32']; c.value = "=C25+(('08 Debt Schedule'!K28+'08 Debt Schedule'!K29)*C25/'09 Operating Model'!K18)+(-'03 Assumptions'!$B$34*MAX(0,C25+('09 Operating Model'!K10*C25/'09 Operating Model'!K18)-MIN((('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*C38+'03 Assumptions'!$B$28*C45+'03 Assumptions'!$B$29*'06 Sources & Uses'!$B$16+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31),'03 Assumptions'!$B$45*C25)))-(('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*C38+'03 Assumptions'!$B$28*C45+'03 Assumptions'!$B$29*'06 Sources & Uses'!$B$16+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31)-'03 Assumptions'!$F$13"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['D32']; c.value = "=D25+(('08 Debt Schedule'!K28+'08 Debt Schedule'!K29)*D25/'09 Operating Model'!K18)+(-'03 Assumptions'!$B$34*MAX(0,D25+('09 Operating Model'!K10*D25/'09 Operating Model'!K18)-MIN((('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*D38+'03 Assumptions'!$B$28*D45+'03 Assumptions'!$B$29*'06 Sources & Uses'!$B$16+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31),'03 Assumptions'!$B$45*D25)))-(('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*D38+'03 Assumptions'!$B$28*D45+'03 Assumptions'!$B$29*'06 Sources & Uses'!$B$16+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31)-'03 Assumptions'!$F$13"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['E32']; c.value = "=E25+(('08 Debt Schedule'!K28+'08 Debt Schedule'!K29)*E25/'09 Operating Model'!K18)+(-'03 Assumptions'!$B$34*MAX(0,E25+('09 Operating Model'!K10*E25/'09 Operating Model'!K18)-MIN((('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*E38+'03 Assumptions'!$B$28*E45+'03 Assumptions'!$B$29*'06 Sources & Uses'!$B$16+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31),'03 Assumptions'!$B$45*E25)))-(('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*E38+'03 Assumptions'!$B$28*E45+'03 Assumptions'!$B$29*'06 Sources & Uses'!$B$16+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31)-'03 Assumptions'!$F$13"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['F32']; c.value = "=F25+(('08 Debt Schedule'!K28+'08 Debt Schedule'!K29)*F25/'09 Operating Model'!K18)+(-'03 Assumptions'!$B$34*MAX(0,F25+('09 Operating Model'!K10*F25/'09 Operating Model'!K18)-MIN((('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*F38+'03 Assumptions'!$B$28*F45+'03 Assumptions'!$B$29*'06 Sources & Uses'!$B$16+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31),'03 Assumptions'!$B$45*F25)))-(('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*F38+'03 Assumptions'!$B$28*F45+'03 Assumptions'!$B$29*'06 Sources & Uses'!$B$16+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31)-'03 Assumptions'!$F$13"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['G32']; c.value = "=G25+(('08 Debt Schedule'!K28+'08 Debt Schedule'!K29)*G25/'09 Operating Model'!K18)+(-'03 Assumptions'!$B$34*MAX(0,G25+('09 Operating Model'!K10*G25/'09 Operating Model'!K18)-MIN((('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*G38+'03 Assumptions'!$B$28*G45+'03 Assumptions'!$B$29*'06 Sources & Uses'!$B$16+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31),'03 Assumptions'!$B$45*G25)))-(('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*G38+'03 Assumptions'!$B$28*G45+'03 Assumptions'!$B$29*'06 Sources & Uses'!$B$16+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31)-'03 Assumptions'!$F$13"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A35']; c.value = 'TLB-EUR end 2027'
c.font = Font(size=10.0)
c = ws['C35']; c.value = "='06 Sources & Uses'!$B$14-MIN('06 Sources & Uses'!$B$14,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN('06 Sources & Uses'!$B$14-MIN('06 Sources & Uses'!$B$14,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38),MAX(0,C28-MIN('06 Sources & Uses'!$B$14,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN('06 Sources & Uses'!$B$15,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['D35']; c.value = "='06 Sources & Uses'!$B$14-MIN('06 Sources & Uses'!$B$14,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN('06 Sources & Uses'!$B$14-MIN('06 Sources & Uses'!$B$14,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38),MAX(0,D28-MIN('06 Sources & Uses'!$B$14,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN('06 Sources & Uses'!$B$15,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['E35']; c.value = "='06 Sources & Uses'!$B$14-MIN('06 Sources & Uses'!$B$14,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN('06 Sources & Uses'!$B$14-MIN('06 Sources & Uses'!$B$14,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38),MAX(0,E28-MIN('06 Sources & Uses'!$B$14,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN('06 Sources & Uses'!$B$15,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['F35']; c.value = "='06 Sources & Uses'!$B$14-MIN('06 Sources & Uses'!$B$14,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN('06 Sources & Uses'!$B$14-MIN('06 Sources & Uses'!$B$14,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38),MAX(0,F28-MIN('06 Sources & Uses'!$B$14,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN('06 Sources & Uses'!$B$15,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['G35']; c.value = "='06 Sources & Uses'!$B$14-MIN('06 Sources & Uses'!$B$14,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN('06 Sources & Uses'!$B$14-MIN('06 Sources & Uses'!$B$14,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38),MAX(0,G28-MIN('06 Sources & Uses'!$B$14,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN('06 Sources & Uses'!$B$15,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A36']; c.value = 'TLB-EUR end 2028'
c.font = Font(size=10.0)
c = ws['C36']; c.value = "=C35-MIN(C35,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(C35-MIN(C35,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38),MAX(0,C29-MIN(C35,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(C42,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['D36']; c.value = "=D35-MIN(D35,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(D35-MIN(D35,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38),MAX(0,D29-MIN(D35,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(D42,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['E36']; c.value = "=E35-MIN(E35,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(E35-MIN(E35,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38),MAX(0,E29-MIN(E35,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(E42,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['F36']; c.value = "=F35-MIN(F35,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(F35-MIN(F35,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38),MAX(0,F29-MIN(F35,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(F42,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['G36']; c.value = "=G35-MIN(G35,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(G35-MIN(G35,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38),MAX(0,G29-MIN(G35,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(G42,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A37']; c.value = 'TLB-EUR end 2029'
c.font = Font(size=10.0)
c = ws['C37']; c.value = "=C36-MIN(C36,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(C36-MIN(C36,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38),MAX(0,C30-MIN(C36,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(C43,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['D37']; c.value = "=D36-MIN(D36,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(D36-MIN(D36,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38),MAX(0,D30-MIN(D36,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(D43,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['E37']; c.value = "=E36-MIN(E36,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(E36-MIN(E36,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38),MAX(0,E30-MIN(E36,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(E43,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['F37']; c.value = "=F36-MIN(F36,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(F36-MIN(F36,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38),MAX(0,F30-MIN(F36,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(F43,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['G37']; c.value = "=G36-MIN(G36,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(G36-MIN(G36,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38),MAX(0,G30-MIN(G36,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(G43,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A38']; c.value = 'TLB-EUR end 2030'
c.font = Font(size=10.0)
c = ws['C38']; c.value = "=C37-MIN(C37,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(C37-MIN(C37,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38),MAX(0,C31-MIN(C37,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(C44,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['D38']; c.value = "=D37-MIN(D37,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(D37-MIN(D37,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38),MAX(0,D31-MIN(D37,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(D44,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['E38']; c.value = "=E37-MIN(E37,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(E37-MIN(E37,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38),MAX(0,E31-MIN(E37,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(E44,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['F38']; c.value = "=F37-MIN(F37,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(F37-MIN(F37,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38),MAX(0,F31-MIN(F37,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(F44,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['G38']; c.value = "=G37-MIN(G37,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(G37-MIN(G37,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38),MAX(0,G31-MIN(G37,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(G44,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A39']; c.value = 'TLB-EUR end 2031'
c.font = Font(size=10.0)
c = ws['C39']; c.value = "=C38-MIN(C38,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(C38-MIN(C38,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38),MAX(0,C32-MIN(C38,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(C45,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['D39']; c.value = "=D38-MIN(D38,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(D38-MIN(D38,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38),MAX(0,D32-MIN(D38,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(D45,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['E39']; c.value = "=E38-MIN(E38,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(E38-MIN(E38,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38),MAX(0,E32-MIN(E38,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(E45,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['F39']; c.value = "=F38-MIN(F38,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(F38-MIN(F38,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38),MAX(0,F32-MIN(F38,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(F45,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['G39']; c.value = "=G38-MIN(G38,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(G38-MIN(G38,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38),MAX(0,G32-MIN(G38,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(G45,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A42']; c.value = 'TLB-USD end 2027'
c.font = Font(size=10.0)
c = ws['C42']; c.value = "='06 Sources & Uses'!$B$15-MIN('06 Sources & Uses'!$B$15,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)-MIN('06 Sources & Uses'!$B$15-MIN('06 Sources & Uses'!$B$15,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38),MAX(0,MAX(0,C28-MIN('06 Sources & Uses'!$B$14,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN('06 Sources & Uses'!$B$15,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38))-('06 Sources & Uses'!$B$14-MIN('06 Sources & Uses'!$B$14,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38))))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['D42']; c.value = "='06 Sources & Uses'!$B$15-MIN('06 Sources & Uses'!$B$15,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)-MIN('06 Sources & Uses'!$B$15-MIN('06 Sources & Uses'!$B$15,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38),MAX(0,MAX(0,D28-MIN('06 Sources & Uses'!$B$14,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN('06 Sources & Uses'!$B$15,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38))-('06 Sources & Uses'!$B$14-MIN('06 Sources & Uses'!$B$14,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38))))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['E42']; c.value = "='06 Sources & Uses'!$B$15-MIN('06 Sources & Uses'!$B$15,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)-MIN('06 Sources & Uses'!$B$15-MIN('06 Sources & Uses'!$B$15,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38),MAX(0,MAX(0,E28-MIN('06 Sources & Uses'!$B$14,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN('06 Sources & Uses'!$B$15,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38))-('06 Sources & Uses'!$B$14-MIN('06 Sources & Uses'!$B$14,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38))))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['F42']; c.value = "='06 Sources & Uses'!$B$15-MIN('06 Sources & Uses'!$B$15,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)-MIN('06 Sources & Uses'!$B$15-MIN('06 Sources & Uses'!$B$15,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38),MAX(0,MAX(0,F28-MIN('06 Sources & Uses'!$B$14,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN('06 Sources & Uses'!$B$15,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38))-('06 Sources & Uses'!$B$14-MIN('06 Sources & Uses'!$B$14,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38))))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['G42']; c.value = "='06 Sources & Uses'!$B$15-MIN('06 Sources & Uses'!$B$15,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)-MIN('06 Sources & Uses'!$B$15-MIN('06 Sources & Uses'!$B$15,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38),MAX(0,MAX(0,G28-MIN('06 Sources & Uses'!$B$14,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN('06 Sources & Uses'!$B$15,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38))-('06 Sources & Uses'!$B$14-MIN('06 Sources & Uses'!$B$14,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38))))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A43']; c.value = 'TLB-USD end 2028'
c.font = Font(size=10.0)
c = ws['C43']; c.value = "=C42-MIN(C42,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)-MIN(C42-MIN(C42,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38),MAX(0,MAX(0,C29-MIN(C35,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(C42,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38))-(C35-MIN(C35,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38))))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['D43']; c.value = "=D42-MIN(D42,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)-MIN(D42-MIN(D42,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38),MAX(0,MAX(0,D29-MIN(D35,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(D42,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38))-(D35-MIN(D35,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38))))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['E43']; c.value = "=E42-MIN(E42,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)-MIN(E42-MIN(E42,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38),MAX(0,MAX(0,E29-MIN(E35,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(E42,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38))-(E35-MIN(E35,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38))))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['F43']; c.value = "=F42-MIN(F42,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)-MIN(F42-MIN(F42,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38),MAX(0,MAX(0,F29-MIN(F35,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(F42,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38))-(F35-MIN(F35,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38))))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['G43']; c.value = "=G42-MIN(G42,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)-MIN(G42-MIN(G42,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38),MAX(0,MAX(0,G29-MIN(G35,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(G42,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38))-(G35-MIN(G35,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38))))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A44']; c.value = 'TLB-USD end 2029'
c.font = Font(size=10.0)
c = ws['C44']; c.value = "=C43-MIN(C43,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)-MIN(C43-MIN(C43,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38),MAX(0,MAX(0,C30-MIN(C36,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(C43,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38))-(C36-MIN(C36,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38))))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['D44']; c.value = "=D43-MIN(D43,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)-MIN(D43-MIN(D43,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38),MAX(0,MAX(0,D30-MIN(D36,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(D43,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38))-(D36-MIN(D36,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38))))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['E44']; c.value = "=E43-MIN(E43,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)-MIN(E43-MIN(E43,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38),MAX(0,MAX(0,E30-MIN(E36,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(E43,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38))-(E36-MIN(E36,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38))))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['F44']; c.value = "=F43-MIN(F43,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)-MIN(F43-MIN(F43,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38),MAX(0,MAX(0,F30-MIN(F36,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(F43,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38))-(F36-MIN(F36,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38))))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['G44']; c.value = "=G43-MIN(G43,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)-MIN(G43-MIN(G43,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38),MAX(0,MAX(0,G30-MIN(G36,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(G43,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38))-(G36-MIN(G36,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38))))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A45']; c.value = 'TLB-USD end 2030'
c.font = Font(size=10.0)
c = ws['C45']; c.value = "=C44-MIN(C44,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)-MIN(C44-MIN(C44,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38),MAX(0,MAX(0,C31-MIN(C37,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(C44,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38))-(C37-MIN(C37,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38))))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['D45']; c.value = "=D44-MIN(D44,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)-MIN(D44-MIN(D44,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38),MAX(0,MAX(0,D31-MIN(D37,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(D44,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38))-(D37-MIN(D37,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38))))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['E45']; c.value = "=E44-MIN(E44,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)-MIN(E44-MIN(E44,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38),MAX(0,MAX(0,E31-MIN(E37,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(E44,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38))-(E37-MIN(E37,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38))))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['F45']; c.value = "=F44-MIN(F44,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)-MIN(F44-MIN(F44,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38),MAX(0,MAX(0,F31-MIN(F37,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(F44,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38))-(F37-MIN(F37,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38))))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['G45']; c.value = "=G44-MIN(G44,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)-MIN(G44-MIN(G44,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38),MAX(0,MAX(0,G31-MIN(G37,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(G44,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38))-(G37-MIN(G37,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38))))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A46']; c.value = 'TLB-USD end 2031'
c.font = Font(size=10.0)
c = ws['C46']; c.value = "=C45-MIN(C45,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)-MIN(C45-MIN(C45,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38),MAX(0,MAX(0,C32-MIN(C38,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(C45,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38))-(C38-MIN(C38,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38))))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['D46']; c.value = "=D45-MIN(D45,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)-MIN(D45-MIN(D45,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38),MAX(0,MAX(0,D32-MIN(D38,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(D45,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38))-(D38-MIN(D38,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38))))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['E46']; c.value = "=E45-MIN(E45,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)-MIN(E45-MIN(E45,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38),MAX(0,MAX(0,E32-MIN(E38,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(E45,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38))-(E38-MIN(E38,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38))))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['F46']; c.value = "=F45-MIN(F45,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)-MIN(F45-MIN(F45,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38),MAX(0,MAX(0,F32-MIN(F38,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(F45,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38))-(F38-MIN(F38,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38))))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['G46']; c.value = "=G45-MIN(G45,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)-MIN(G45-MIN(G45,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38),MAX(0,MAX(0,G32-MIN(G38,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(G45,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38))-(G38-MIN(G38,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38))))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A49']; c.value = 'Excess cash end 2027'
c.font = Font(size=10.0)
c = ws['C49']; c.value = "=0+MAX(0,MAX(0,C28-MIN('06 Sources & Uses'!$B$14,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN('06 Sources & Uses'!$B$15,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38))-('06 Sources & Uses'!$B$14-MIN('06 Sources & Uses'!$B$14,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38))-('06 Sources & Uses'!$B$15-MIN('06 Sources & Uses'!$B$15,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['D49']; c.value = "=0+MAX(0,MAX(0,D28-MIN('06 Sources & Uses'!$B$14,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN('06 Sources & Uses'!$B$15,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38))-('06 Sources & Uses'!$B$14-MIN('06 Sources & Uses'!$B$14,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38))-('06 Sources & Uses'!$B$15-MIN('06 Sources & Uses'!$B$15,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['E49']; c.value = "=0+MAX(0,MAX(0,E28-MIN('06 Sources & Uses'!$B$14,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN('06 Sources & Uses'!$B$15,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38))-('06 Sources & Uses'!$B$14-MIN('06 Sources & Uses'!$B$14,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38))-('06 Sources & Uses'!$B$15-MIN('06 Sources & Uses'!$B$15,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['F49']; c.value = "=0+MAX(0,MAX(0,F28-MIN('06 Sources & Uses'!$B$14,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN('06 Sources & Uses'!$B$15,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38))-('06 Sources & Uses'!$B$14-MIN('06 Sources & Uses'!$B$14,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38))-('06 Sources & Uses'!$B$15-MIN('06 Sources & Uses'!$B$15,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['G49']; c.value = "=0+MAX(0,MAX(0,G28-MIN('06 Sources & Uses'!$B$14,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN('06 Sources & Uses'!$B$15,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38))-('06 Sources & Uses'!$B$14-MIN('06 Sources & Uses'!$B$14,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38))-('06 Sources & Uses'!$B$15-MIN('06 Sources & Uses'!$B$15,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A50']; c.value = 'Excess cash end 2028'
c.font = Font(size=10.0)
c = ws['C50']; c.value = "=C49+MAX(0,MAX(0,C29-MIN(C35,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(C42,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38))-(C35-MIN(C35,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38))-(C42-MIN(C42,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['D50']; c.value = "=D49+MAX(0,MAX(0,D29-MIN(D35,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(D42,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38))-(D35-MIN(D35,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38))-(D42-MIN(D42,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['E50']; c.value = "=E49+MAX(0,MAX(0,E29-MIN(E35,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(E42,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38))-(E35-MIN(E35,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38))-(E42-MIN(E42,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['F50']; c.value = "=F49+MAX(0,MAX(0,F29-MIN(F35,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(F42,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38))-(F35-MIN(F35,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38))-(F42-MIN(F42,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['G50']; c.value = "=G49+MAX(0,MAX(0,G29-MIN(G35,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(G42,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38))-(G35-MIN(G35,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38))-(G42-MIN(G42,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A51']; c.value = 'Excess cash end 2029'
c.font = Font(size=10.0)
c = ws['C51']; c.value = "=C50+MAX(0,MAX(0,C30-MIN(C36,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(C43,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38))-(C36-MIN(C36,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38))-(C43-MIN(C43,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['D51']; c.value = "=D50+MAX(0,MAX(0,D30-MIN(D36,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(D43,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38))-(D36-MIN(D36,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38))-(D43-MIN(D43,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['E51']; c.value = "=E50+MAX(0,MAX(0,E30-MIN(E36,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(E43,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38))-(E36-MIN(E36,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38))-(E43-MIN(E43,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['F51']; c.value = "=F50+MAX(0,MAX(0,F30-MIN(F36,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(F43,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38))-(F36-MIN(F36,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38))-(F43-MIN(F43,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['G51']; c.value = "=G50+MAX(0,MAX(0,G30-MIN(G36,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(G43,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38))-(G36-MIN(G36,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38))-(G43-MIN(G43,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A52']; c.value = 'Excess cash end 2030'
c.font = Font(size=10.0)
c = ws['C52']; c.value = "=C51+MAX(0,MAX(0,C31-MIN(C37,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(C44,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38))-(C37-MIN(C37,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38))-(C44-MIN(C44,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['D52']; c.value = "=D51+MAX(0,MAX(0,D31-MIN(D37,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(D44,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38))-(D37-MIN(D37,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38))-(D44-MIN(D44,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['E52']; c.value = "=E51+MAX(0,MAX(0,E31-MIN(E37,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(E44,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38))-(E37-MIN(E37,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38))-(E44-MIN(E44,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['F52']; c.value = "=F51+MAX(0,MAX(0,F31-MIN(F37,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(F44,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38))-(F37-MIN(F37,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38))-(F44-MIN(F44,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['G52']; c.value = "=G51+MAX(0,MAX(0,G31-MIN(G37,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(G44,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38))-(G37-MIN(G37,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38))-(G44-MIN(G44,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A53']; c.value = 'Excess cash end 2031'
c.font = Font(size=10.0)
c = ws['C53']; c.value = "=C52+MAX(0,MAX(0,C32-MIN(C38,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(C45,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38))-(C38-MIN(C38,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38))-(C45-MIN(C45,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['D53']; c.value = "=D52+MAX(0,MAX(0,D32-MIN(D38,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(D45,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38))-(D38-MIN(D38,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38))-(D45-MIN(D45,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['E53']; c.value = "=E52+MAX(0,MAX(0,E32-MIN(E38,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(E45,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38))-(E38-MIN(E38,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38))-(E45-MIN(E45,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['F53']; c.value = "=F52+MAX(0,MAX(0,F32-MIN(F38,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(F45,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38))-(F38-MIN(F38,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38))-(F45-MIN(F45,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['G53']; c.value = "=G52+MAX(0,MAX(0,G32-MIN(G38,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38)-MIN(G45,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38))-(G38-MIN(G38,'06 Sources & Uses'!$B$14*'03 Assumptions'!$B$38))-(G45-MIN(G45,'06 Sources & Uses'!$B$15*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A56']; c.value = 'GRID 2 - Exit multiple (rows) x EBITDA CAGR (cols) - IRR at 2031 exit, EXACT'
c.font = Font(bold=True, color='FF1F3864')
c = ws['C57']; c.value = '=C17'
c.number_format = '0.0%'
c.font = Font(color='FF008000', size=10.0)
c = ws['D57']; c.value = '=D17'
c.number_format = '0.0%'
c.font = Font(color='FF008000', size=10.0)
c = ws['E57']; c.value = '=E17'
c.number_format = '0.0%'
c.font = Font(color='FF008000', size=10.0)
c = ws['F57']; c.value = '=F17'
c.number_format = '0.0%'
c.font = Font(color='FF008000', size=10.0)
c = ws['G57']; c.value = '=G17'
c.number_format = '0.0%'
c.font = Font(color='FF008000', size=10.0)
c = ws['B58']; c.value = 9
c.number_format = '0.00#'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['C58']; c.value = "=(($B58*C25-(C39+C46+'06 Sources & Uses'!$B$16-C53-'03 Assumptions'!$B$44))/'06 Sources & Uses'!$B$18)^(1/5)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['D58']; c.value = "=(($B58*D25-(D39+D46+'06 Sources & Uses'!$B$16-D53-'03 Assumptions'!$B$44))/'06 Sources & Uses'!$B$18)^(1/5)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['E58']; c.value = "=(($B58*E25-(E39+E46+'06 Sources & Uses'!$B$16-E53-'03 Assumptions'!$B$44))/'06 Sources & Uses'!$B$18)^(1/5)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['F58']; c.value = "=(($B58*F25-(F39+F46+'06 Sources & Uses'!$B$16-F53-'03 Assumptions'!$B$44))/'06 Sources & Uses'!$B$18)^(1/5)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['G58']; c.value = "=(($B58*G25-(G39+G46+'06 Sources & Uses'!$B$16-G53-'03 Assumptions'!$B$44))/'06 Sources & Uses'!$B$18)^(1/5)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['B59']; c.value = 10
c.number_format = '0.00#'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['C59']; c.value = "=(($B59*C25-(C39+C46+'06 Sources & Uses'!$B$16-C53-'03 Assumptions'!$B$44))/'06 Sources & Uses'!$B$18)^(1/5)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['D59']; c.value = "=(($B59*D25-(D39+D46+'06 Sources & Uses'!$B$16-D53-'03 Assumptions'!$B$44))/'06 Sources & Uses'!$B$18)^(1/5)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['E59']; c.value = "=(($B59*E25-(E39+E46+'06 Sources & Uses'!$B$16-E53-'03 Assumptions'!$B$44))/'06 Sources & Uses'!$B$18)^(1/5)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['F59']; c.value = "=(($B59*F25-(F39+F46+'06 Sources & Uses'!$B$16-F53-'03 Assumptions'!$B$44))/'06 Sources & Uses'!$B$18)^(1/5)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['G59']; c.value = "=(($B59*G25-(G39+G46+'06 Sources & Uses'!$B$16-G53-'03 Assumptions'!$B$44))/'06 Sources & Uses'!$B$18)^(1/5)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['B60']; c.value = 11
c.number_format = '0.00#'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['C60']; c.value = "=(($B60*C25-(C39+C46+'06 Sources & Uses'!$B$16-C53-'03 Assumptions'!$B$44))/'06 Sources & Uses'!$B$18)^(1/5)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['D60']; c.value = "=(($B60*D25-(D39+D46+'06 Sources & Uses'!$B$16-D53-'03 Assumptions'!$B$44))/'06 Sources & Uses'!$B$18)^(1/5)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['E60']; c.value = "=(($B60*E25-(E39+E46+'06 Sources & Uses'!$B$16-E53-'03 Assumptions'!$B$44))/'06 Sources & Uses'!$B$18)^(1/5)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['F60']; c.value = "=(($B60*F25-(F39+F46+'06 Sources & Uses'!$B$16-F53-'03 Assumptions'!$B$44))/'06 Sources & Uses'!$B$18)^(1/5)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['G60']; c.value = "=(($B60*G25-(G39+G46+'06 Sources & Uses'!$B$16-G53-'03 Assumptions'!$B$44))/'06 Sources & Uses'!$B$18)^(1/5)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['B61']; c.value = 11.5
c.number_format = '0.00#'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['C61']; c.value = "=(($B61*C25-(C39+C46+'06 Sources & Uses'!$B$16-C53-'03 Assumptions'!$B$44))/'06 Sources & Uses'!$B$18)^(1/5)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['D61']; c.value = "=(($B61*D25-(D39+D46+'06 Sources & Uses'!$B$16-D53-'03 Assumptions'!$B$44))/'06 Sources & Uses'!$B$18)^(1/5)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['E61']; c.value = "=(($B61*E25-(E39+E46+'06 Sources & Uses'!$B$16-E53-'03 Assumptions'!$B$44))/'06 Sources & Uses'!$B$18)^(1/5)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['F61']; c.value = "=(($B61*F25-(F39+F46+'06 Sources & Uses'!$B$16-F53-'03 Assumptions'!$B$44))/'06 Sources & Uses'!$B$18)^(1/5)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['G61']; c.value = "=(($B61*G25-(G39+G46+'06 Sources & Uses'!$B$16-G53-'03 Assumptions'!$B$44))/'06 Sources & Uses'!$B$18)^(1/5)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['B62']; c.value = 12
c.number_format = '0.00#'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['C62']; c.value = "=(($B62*C25-(C39+C46+'06 Sources & Uses'!$B$16-C53-'03 Assumptions'!$B$44))/'06 Sources & Uses'!$B$18)^(1/5)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['D62']; c.value = "=(($B62*D25-(D39+D46+'06 Sources & Uses'!$B$16-D53-'03 Assumptions'!$B$44))/'06 Sources & Uses'!$B$18)^(1/5)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['E62']; c.value = "=(($B62*E25-(E39+E46+'06 Sources & Uses'!$B$16-E53-'03 Assumptions'!$B$44))/'06 Sources & Uses'!$B$18)^(1/5)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['F62']; c.value = "=(($B62*F25-(F39+F46+'06 Sources & Uses'!$B$16-F53-'03 Assumptions'!$B$44))/'06 Sources & Uses'!$B$18)^(1/5)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['G62']; c.value = "=(($B62*G25-(G39+G46+'06 Sources & Uses'!$B$16-G53-'03 Assumptions'!$B$44))/'06 Sources & Uses'!$B$18)^(1/5)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['B63']; c.value = 12.9
c.number_format = '0.00#'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['C63']; c.value = "=(($B63*C25-(C39+C46+'06 Sources & Uses'!$B$16-C53-'03 Assumptions'!$B$44))/'06 Sources & Uses'!$B$18)^(1/5)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['D63']; c.value = "=(($B63*D25-(D39+D46+'06 Sources & Uses'!$B$16-D53-'03 Assumptions'!$B$44))/'06 Sources & Uses'!$B$18)^(1/5)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['E63']; c.value = "=(($B63*E25-(E39+E46+'06 Sources & Uses'!$B$16-E53-'03 Assumptions'!$B$44))/'06 Sources & Uses'!$B$18)^(1/5)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['F63']; c.value = "=(($B63*F25-(F39+F46+'06 Sources & Uses'!$B$16-F53-'03 Assumptions'!$B$44))/'06 Sources & Uses'!$B$18)^(1/5)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['G63']; c.value = "=(($B63*G25-(G39+G46+'06 Sources & Uses'!$B$16-G53-'03 Assumptions'!$B$44))/'06 Sources & Uses'!$B$18)^(1/5)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['B64']; c.value = 14
c.number_format = '0.00#'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['C64']; c.value = "=(($B64*C25-(C39+C46+'06 Sources & Uses'!$B$16-C53-'03 Assumptions'!$B$44))/'06 Sources & Uses'!$B$18)^(1/5)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['D64']; c.value = "=(($B64*D25-(D39+D46+'06 Sources & Uses'!$B$16-D53-'03 Assumptions'!$B$44))/'06 Sources & Uses'!$B$18)^(1/5)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['E64']; c.value = "=(($B64*E25-(E39+E46+'06 Sources & Uses'!$B$16-E53-'03 Assumptions'!$B$44))/'06 Sources & Uses'!$B$18)^(1/5)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['F64']; c.value = "=(($B64*F25-(F39+F46+'06 Sources & Uses'!$B$16-F53-'03 Assumptions'!$B$44))/'06 Sources & Uses'!$B$18)^(1/5)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['G64']; c.value = "=(($B64*G25-(G39+G46+'06 Sources & Uses'!$B$16-G53-'03 Assumptions'!$B$44))/'06 Sources & Uses'!$B$18)^(1/5)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['A67']; c.value = 'ENGINE B - entry-leverage cases (cols C-E: 5.0x / 5.5x / 6.0x). Results in GRID 3.'
c.font = Font(bold=True, color='FF1F3864')
c = ws['A68']; c.value = 'Entry leverage'
c.font = Font(size=10.0)
c = ws['C68']; c.value = 5
c.number_format = '0.00#'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['D68']; c.value = 5.5
c.number_format = '0.00#'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['E68']; c.value = 6
c.number_format = '0.00#'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['A69']; c.value = 'Total debt'
c.font = Font(size=10.0)
c = ws['C69']; c.value = "=C68*'03 Assumptions'!$B$16"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['D69']; c.value = "=D68*'03 Assumptions'!$B$16"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['E69']; c.value = "=E68*'03 Assumptions'!$B$16"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A70']; c.value = 'Sponsor equity'
c.font = Font(size=10.0)
c = ws['C70']; c.value = "='06 Sources & Uses'!$B$4+'06 Sources & Uses'!$B$5+'06 Sources & Uses'!$B$6+'03 Assumptions'!$B$44+'03 Assumptions'!$B$25*C69-C69"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['D70']; c.value = "='06 Sources & Uses'!$B$4+'06 Sources & Uses'!$B$5+'06 Sources & Uses'!$B$6+'03 Assumptions'!$B$44+'03 Assumptions'!$B$25*D69-D69"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['E70']; c.value = "='06 Sources & Uses'!$B$4+'06 Sources & Uses'!$B$5+'06 Sources & Uses'!$B$6+'03 Assumptions'!$B$44+'03 Assumptions'!$B$25*E69-E69"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A72']; c.value = 'FCF pre-amort 2027'
c.font = Font(size=10.0)
c = ws['C72']; c.value = "='09 Operating Model'!G18+'08 Debt Schedule'!G28+'08 Debt Schedule'!G29+(-'03 Assumptions'!$B$34*MAX(0,'09 Operating Model'!G18+'09 Operating Model'!G10-MIN((('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*C$69*'03 Assumptions'!$B$36+'03 Assumptions'!$B$28*C$69*'03 Assumptions'!$B$37+'03 Assumptions'!$B$29*C$69*(1-'03 Assumptions'!$B$36-'03 Assumptions'!$B$37)+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31),'03 Assumptions'!$B$45*'09 Operating Model'!G18)))-(('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*C$69*'03 Assumptions'!$B$36+'03 Assumptions'!$B$28*C$69*'03 Assumptions'!$B$37+'03 Assumptions'!$B$29*C$69*(1-'03 Assumptions'!$B$36-'03 Assumptions'!$B$37)+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31)-'03 Assumptions'!$F$13"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['D72']; c.value = "='09 Operating Model'!G18+'08 Debt Schedule'!G28+'08 Debt Schedule'!G29+(-'03 Assumptions'!$B$34*MAX(0,'09 Operating Model'!G18+'09 Operating Model'!G10-MIN((('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*D$69*'03 Assumptions'!$B$36+'03 Assumptions'!$B$28*D$69*'03 Assumptions'!$B$37+'03 Assumptions'!$B$29*D$69*(1-'03 Assumptions'!$B$36-'03 Assumptions'!$B$37)+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31),'03 Assumptions'!$B$45*'09 Operating Model'!G18)))-(('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*D$69*'03 Assumptions'!$B$36+'03 Assumptions'!$B$28*D$69*'03 Assumptions'!$B$37+'03 Assumptions'!$B$29*D$69*(1-'03 Assumptions'!$B$36-'03 Assumptions'!$B$37)+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31)-'03 Assumptions'!$F$13"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['E72']; c.value = "='09 Operating Model'!G18+'08 Debt Schedule'!G28+'08 Debt Schedule'!G29+(-'03 Assumptions'!$B$34*MAX(0,'09 Operating Model'!G18+'09 Operating Model'!G10-MIN((('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*E$69*'03 Assumptions'!$B$36+'03 Assumptions'!$B$28*E$69*'03 Assumptions'!$B$37+'03 Assumptions'!$B$29*E$69*(1-'03 Assumptions'!$B$36-'03 Assumptions'!$B$37)+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31),'03 Assumptions'!$B$45*'09 Operating Model'!G18)))-(('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*E$69*'03 Assumptions'!$B$36+'03 Assumptions'!$B$28*E$69*'03 Assumptions'!$B$37+'03 Assumptions'!$B$29*E$69*(1-'03 Assumptions'!$B$36-'03 Assumptions'!$B$37)+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31)-'03 Assumptions'!$F$13"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A73']; c.value = 'FCF pre-amort 2028'
c.font = Font(size=10.0)
c = ws['C73']; c.value = "='09 Operating Model'!H18+'08 Debt Schedule'!H28+'08 Debt Schedule'!H29+(-'03 Assumptions'!$B$34*MAX(0,'09 Operating Model'!H18+'09 Operating Model'!H10-MIN((('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*C78+'03 Assumptions'!$B$28*C84+'03 Assumptions'!$B$29*C$69*(1-'03 Assumptions'!$B$36-'03 Assumptions'!$B$37)+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31),'03 Assumptions'!$B$45*'09 Operating Model'!H18)))-(('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*C78+'03 Assumptions'!$B$28*C84+'03 Assumptions'!$B$29*C$69*(1-'03 Assumptions'!$B$36-'03 Assumptions'!$B$37)+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31)-'03 Assumptions'!$F$13"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['D73']; c.value = "='09 Operating Model'!H18+'08 Debt Schedule'!H28+'08 Debt Schedule'!H29+(-'03 Assumptions'!$B$34*MAX(0,'09 Operating Model'!H18+'09 Operating Model'!H10-MIN((('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*D78+'03 Assumptions'!$B$28*D84+'03 Assumptions'!$B$29*D$69*(1-'03 Assumptions'!$B$36-'03 Assumptions'!$B$37)+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31),'03 Assumptions'!$B$45*'09 Operating Model'!H18)))-(('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*D78+'03 Assumptions'!$B$28*D84+'03 Assumptions'!$B$29*D$69*(1-'03 Assumptions'!$B$36-'03 Assumptions'!$B$37)+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31)-'03 Assumptions'!$F$13"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['E73']; c.value = "='09 Operating Model'!H18+'08 Debt Schedule'!H28+'08 Debt Schedule'!H29+(-'03 Assumptions'!$B$34*MAX(0,'09 Operating Model'!H18+'09 Operating Model'!H10-MIN((('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*E78+'03 Assumptions'!$B$28*E84+'03 Assumptions'!$B$29*E$69*(1-'03 Assumptions'!$B$36-'03 Assumptions'!$B$37)+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31),'03 Assumptions'!$B$45*'09 Operating Model'!H18)))-(('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*E78+'03 Assumptions'!$B$28*E84+'03 Assumptions'!$B$29*E$69*(1-'03 Assumptions'!$B$36-'03 Assumptions'!$B$37)+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31)-'03 Assumptions'!$F$13"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A74']; c.value = 'FCF pre-amort 2029'
c.font = Font(size=10.0)
c = ws['C74']; c.value = "='09 Operating Model'!I18+'08 Debt Schedule'!I28+'08 Debt Schedule'!I29+(-'03 Assumptions'!$B$34*MAX(0,'09 Operating Model'!I18+'09 Operating Model'!I10-MIN((('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*C79+'03 Assumptions'!$B$28*C85+'03 Assumptions'!$B$29*C$69*(1-'03 Assumptions'!$B$36-'03 Assumptions'!$B$37)+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31),'03 Assumptions'!$B$45*'09 Operating Model'!I18)))-(('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*C79+'03 Assumptions'!$B$28*C85+'03 Assumptions'!$B$29*C$69*(1-'03 Assumptions'!$B$36-'03 Assumptions'!$B$37)+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31)-'03 Assumptions'!$F$13"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['D74']; c.value = "='09 Operating Model'!I18+'08 Debt Schedule'!I28+'08 Debt Schedule'!I29+(-'03 Assumptions'!$B$34*MAX(0,'09 Operating Model'!I18+'09 Operating Model'!I10-MIN((('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*D79+'03 Assumptions'!$B$28*D85+'03 Assumptions'!$B$29*D$69*(1-'03 Assumptions'!$B$36-'03 Assumptions'!$B$37)+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31),'03 Assumptions'!$B$45*'09 Operating Model'!I18)))-(('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*D79+'03 Assumptions'!$B$28*D85+'03 Assumptions'!$B$29*D$69*(1-'03 Assumptions'!$B$36-'03 Assumptions'!$B$37)+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31)-'03 Assumptions'!$F$13"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['E74']; c.value = "='09 Operating Model'!I18+'08 Debt Schedule'!I28+'08 Debt Schedule'!I29+(-'03 Assumptions'!$B$34*MAX(0,'09 Operating Model'!I18+'09 Operating Model'!I10-MIN((('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*E79+'03 Assumptions'!$B$28*E85+'03 Assumptions'!$B$29*E$69*(1-'03 Assumptions'!$B$36-'03 Assumptions'!$B$37)+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31),'03 Assumptions'!$B$45*'09 Operating Model'!I18)))-(('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*E79+'03 Assumptions'!$B$28*E85+'03 Assumptions'!$B$29*E$69*(1-'03 Assumptions'!$B$36-'03 Assumptions'!$B$37)+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31)-'03 Assumptions'!$F$13"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A75']; c.value = 'FCF pre-amort 2030'
c.font = Font(size=10.0)
c = ws['C75']; c.value = "='09 Operating Model'!J18+'08 Debt Schedule'!J28+'08 Debt Schedule'!J29+(-'03 Assumptions'!$B$34*MAX(0,'09 Operating Model'!J18+'09 Operating Model'!J10-MIN((('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*C80+'03 Assumptions'!$B$28*C86+'03 Assumptions'!$B$29*C$69*(1-'03 Assumptions'!$B$36-'03 Assumptions'!$B$37)+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31),'03 Assumptions'!$B$45*'09 Operating Model'!J18)))-(('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*C80+'03 Assumptions'!$B$28*C86+'03 Assumptions'!$B$29*C$69*(1-'03 Assumptions'!$B$36-'03 Assumptions'!$B$37)+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31)-'03 Assumptions'!$F$13"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['D75']; c.value = "='09 Operating Model'!J18+'08 Debt Schedule'!J28+'08 Debt Schedule'!J29+(-'03 Assumptions'!$B$34*MAX(0,'09 Operating Model'!J18+'09 Operating Model'!J10-MIN((('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*D80+'03 Assumptions'!$B$28*D86+'03 Assumptions'!$B$29*D$69*(1-'03 Assumptions'!$B$36-'03 Assumptions'!$B$37)+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31),'03 Assumptions'!$B$45*'09 Operating Model'!J18)))-(('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*D80+'03 Assumptions'!$B$28*D86+'03 Assumptions'!$B$29*D$69*(1-'03 Assumptions'!$B$36-'03 Assumptions'!$B$37)+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31)-'03 Assumptions'!$F$13"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['E75']; c.value = "='09 Operating Model'!J18+'08 Debt Schedule'!J28+'08 Debt Schedule'!J29+(-'03 Assumptions'!$B$34*MAX(0,'09 Operating Model'!J18+'09 Operating Model'!J10-MIN((('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*E80+'03 Assumptions'!$B$28*E86+'03 Assumptions'!$B$29*E$69*(1-'03 Assumptions'!$B$36-'03 Assumptions'!$B$37)+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31),'03 Assumptions'!$B$45*'09 Operating Model'!J18)))-(('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*E80+'03 Assumptions'!$B$28*E86+'03 Assumptions'!$B$29*E$69*(1-'03 Assumptions'!$B$36-'03 Assumptions'!$B$37)+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31)-'03 Assumptions'!$F$13"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A76']; c.value = 'FCF pre-amort 2031'
c.font = Font(size=10.0)
c = ws['C76']; c.value = "='09 Operating Model'!K18+'08 Debt Schedule'!K28+'08 Debt Schedule'!K29+(-'03 Assumptions'!$B$34*MAX(0,'09 Operating Model'!K18+'09 Operating Model'!K10-MIN((('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*C81+'03 Assumptions'!$B$28*C87+'03 Assumptions'!$B$29*C$69*(1-'03 Assumptions'!$B$36-'03 Assumptions'!$B$37)+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31),'03 Assumptions'!$B$45*'09 Operating Model'!K18)))-(('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*C81+'03 Assumptions'!$B$28*C87+'03 Assumptions'!$B$29*C$69*(1-'03 Assumptions'!$B$36-'03 Assumptions'!$B$37)+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31)-'03 Assumptions'!$F$13"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['D76']; c.value = "='09 Operating Model'!K18+'08 Debt Schedule'!K28+'08 Debt Schedule'!K29+(-'03 Assumptions'!$B$34*MAX(0,'09 Operating Model'!K18+'09 Operating Model'!K10-MIN((('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*D81+'03 Assumptions'!$B$28*D87+'03 Assumptions'!$B$29*D$69*(1-'03 Assumptions'!$B$36-'03 Assumptions'!$B$37)+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31),'03 Assumptions'!$B$45*'09 Operating Model'!K18)))-(('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*D81+'03 Assumptions'!$B$28*D87+'03 Assumptions'!$B$29*D$69*(1-'03 Assumptions'!$B$36-'03 Assumptions'!$B$37)+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31)-'03 Assumptions'!$F$13"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['E76']; c.value = "='09 Operating Model'!K18+'08 Debt Schedule'!K28+'08 Debt Schedule'!K29+(-'03 Assumptions'!$B$34*MAX(0,'09 Operating Model'!K18+'09 Operating Model'!K10-MIN((('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*E81+'03 Assumptions'!$B$28*E87+'03 Assumptions'!$B$29*E$69*(1-'03 Assumptions'!$B$36-'03 Assumptions'!$B$37)+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31),'03 Assumptions'!$B$45*'09 Operating Model'!K18)))-(('03 Assumptions'!$B$26+'03 Assumptions'!$B$27)*E81+'03 Assumptions'!$B$28*E87+'03 Assumptions'!$B$29*E$69*(1-'03 Assumptions'!$B$36-'03 Assumptions'!$B$37)+'03 Assumptions'!$B$30*'03 Assumptions'!$B$31)-'03 Assumptions'!$F$13"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A78']; c.value = 'TLB-EUR end 2027'
c.font = Font(size=10.0)
c = ws['C78']; c.value = "=C$69*'03 Assumptions'!$B$36-MIN(C$69*'03 Assumptions'!$B$36,C$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38)-MIN(C$69*'03 Assumptions'!$B$36-MIN(C$69*'03 Assumptions'!$B$36,C$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38),MAX(0,C72-MIN(C$69*'03 Assumptions'!$B$36,C$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38)-MIN(C$69*'03 Assumptions'!$B$37,C$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['D78']; c.value = "=D$69*'03 Assumptions'!$B$36-MIN(D$69*'03 Assumptions'!$B$36,D$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38)-MIN(D$69*'03 Assumptions'!$B$36-MIN(D$69*'03 Assumptions'!$B$36,D$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38),MAX(0,D72-MIN(D$69*'03 Assumptions'!$B$36,D$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38)-MIN(D$69*'03 Assumptions'!$B$37,D$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['E78']; c.value = "=E$69*'03 Assumptions'!$B$36-MIN(E$69*'03 Assumptions'!$B$36,E$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38)-MIN(E$69*'03 Assumptions'!$B$36-MIN(E$69*'03 Assumptions'!$B$36,E$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38),MAX(0,E72-MIN(E$69*'03 Assumptions'!$B$36,E$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38)-MIN(E$69*'03 Assumptions'!$B$37,E$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A79']; c.value = 'TLB-EUR end 2028'
c.font = Font(size=10.0)
c = ws['C79']; c.value = "=C78-MIN(C78,C$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38)-MIN(C78-MIN(C78,C$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38),MAX(0,C73-MIN(C78,C$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38)-MIN(C84,C$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['D79']; c.value = "=D78-MIN(D78,D$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38)-MIN(D78-MIN(D78,D$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38),MAX(0,D73-MIN(D78,D$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38)-MIN(D84,D$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['E79']; c.value = "=E78-MIN(E78,E$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38)-MIN(E78-MIN(E78,E$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38),MAX(0,E73-MIN(E78,E$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38)-MIN(E84,E$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A80']; c.value = 'TLB-EUR end 2029'
c.font = Font(size=10.0)
c = ws['C80']; c.value = "=C79-MIN(C79,C$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38)-MIN(C79-MIN(C79,C$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38),MAX(0,C74-MIN(C79,C$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38)-MIN(C85,C$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['D80']; c.value = "=D79-MIN(D79,D$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38)-MIN(D79-MIN(D79,D$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38),MAX(0,D74-MIN(D79,D$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38)-MIN(D85,D$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['E80']; c.value = "=E79-MIN(E79,E$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38)-MIN(E79-MIN(E79,E$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38),MAX(0,E74-MIN(E79,E$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38)-MIN(E85,E$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A81']; c.value = 'TLB-EUR end 2030'
c.font = Font(size=10.0)
c = ws['C81']; c.value = "=C80-MIN(C80,C$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38)-MIN(C80-MIN(C80,C$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38),MAX(0,C75-MIN(C80,C$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38)-MIN(C86,C$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['D81']; c.value = "=D80-MIN(D80,D$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38)-MIN(D80-MIN(D80,D$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38),MAX(0,D75-MIN(D80,D$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38)-MIN(D86,D$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['E81']; c.value = "=E80-MIN(E80,E$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38)-MIN(E80-MIN(E80,E$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38),MAX(0,E75-MIN(E80,E$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38)-MIN(E86,E$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A82']; c.value = 'TLB-EUR end 2031'
c.font = Font(size=10.0)
c = ws['C82']; c.value = "=C81-MIN(C81,C$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38)-MIN(C81-MIN(C81,C$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38),MAX(0,C76-MIN(C81,C$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38)-MIN(C87,C$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['D82']; c.value = "=D81-MIN(D81,D$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38)-MIN(D81-MIN(D81,D$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38),MAX(0,D76-MIN(D81,D$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38)-MIN(D87,D$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['E82']; c.value = "=E81-MIN(E81,E$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38)-MIN(E81-MIN(E81,E$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38),MAX(0,E76-MIN(E81,E$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38)-MIN(E87,E$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A84']; c.value = 'TLB-USD end 2027'
c.font = Font(size=10.0)
c = ws['C84']; c.value = "=C$69*'03 Assumptions'!$B$37-MIN(C$69*'03 Assumptions'!$B$37,C$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38)-MIN(C$69*'03 Assumptions'!$B$37-MIN(C$69*'03 Assumptions'!$B$37,C$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38),MAX(0,MAX(0,C72-MIN(C$69*'03 Assumptions'!$B$36,C$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38)-MIN(C$69*'03 Assumptions'!$B$37,C$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38))-(C$69*'03 Assumptions'!$B$36-MIN(C$69*'03 Assumptions'!$B$36,C$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38))))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['D84']; c.value = "=D$69*'03 Assumptions'!$B$37-MIN(D$69*'03 Assumptions'!$B$37,D$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38)-MIN(D$69*'03 Assumptions'!$B$37-MIN(D$69*'03 Assumptions'!$B$37,D$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38),MAX(0,MAX(0,D72-MIN(D$69*'03 Assumptions'!$B$36,D$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38)-MIN(D$69*'03 Assumptions'!$B$37,D$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38))-(D$69*'03 Assumptions'!$B$36-MIN(D$69*'03 Assumptions'!$B$36,D$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38))))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['E84']; c.value = "=E$69*'03 Assumptions'!$B$37-MIN(E$69*'03 Assumptions'!$B$37,E$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38)-MIN(E$69*'03 Assumptions'!$B$37-MIN(E$69*'03 Assumptions'!$B$37,E$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38),MAX(0,MAX(0,E72-MIN(E$69*'03 Assumptions'!$B$36,E$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38)-MIN(E$69*'03 Assumptions'!$B$37,E$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38))-(E$69*'03 Assumptions'!$B$36-MIN(E$69*'03 Assumptions'!$B$36,E$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38))))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A85']; c.value = 'TLB-USD end 2028'
c.font = Font(size=10.0)
c = ws['C85']; c.value = "=C84-MIN(C84,C$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38)-MIN(C84-MIN(C84,C$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38),MAX(0,MAX(0,C73-MIN(C78,C$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38)-MIN(C84,C$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38))-(C78-MIN(C78,C$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38))))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['D85']; c.value = "=D84-MIN(D84,D$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38)-MIN(D84-MIN(D84,D$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38),MAX(0,MAX(0,D73-MIN(D78,D$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38)-MIN(D84,D$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38))-(D78-MIN(D78,D$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38))))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['E85']; c.value = "=E84-MIN(E84,E$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38)-MIN(E84-MIN(E84,E$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38),MAX(0,MAX(0,E73-MIN(E78,E$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38)-MIN(E84,E$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38))-(E78-MIN(E78,E$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38))))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A86']; c.value = 'TLB-USD end 2029'
c.font = Font(size=10.0)
c = ws['C86']; c.value = "=C85-MIN(C85,C$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38)-MIN(C85-MIN(C85,C$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38),MAX(0,MAX(0,C74-MIN(C79,C$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38)-MIN(C85,C$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38))-(C79-MIN(C79,C$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38))))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['D86']; c.value = "=D85-MIN(D85,D$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38)-MIN(D85-MIN(D85,D$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38),MAX(0,MAX(0,D74-MIN(D79,D$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38)-MIN(D85,D$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38))-(D79-MIN(D79,D$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38))))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['E86']; c.value = "=E85-MIN(E85,E$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38)-MIN(E85-MIN(E85,E$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38),MAX(0,MAX(0,E74-MIN(E79,E$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38)-MIN(E85,E$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38))-(E79-MIN(E79,E$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38))))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A87']; c.value = 'TLB-USD end 2030'
c.font = Font(size=10.0)
c = ws['C87']; c.value = "=C86-MIN(C86,C$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38)-MIN(C86-MIN(C86,C$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38),MAX(0,MAX(0,C75-MIN(C80,C$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38)-MIN(C86,C$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38))-(C80-MIN(C80,C$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38))))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['D87']; c.value = "=D86-MIN(D86,D$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38)-MIN(D86-MIN(D86,D$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38),MAX(0,MAX(0,D75-MIN(D80,D$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38)-MIN(D86,D$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38))-(D80-MIN(D80,D$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38))))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['E87']; c.value = "=E86-MIN(E86,E$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38)-MIN(E86-MIN(E86,E$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38),MAX(0,MAX(0,E75-MIN(E80,E$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38)-MIN(E86,E$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38))-(E80-MIN(E80,E$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38))))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A88']; c.value = 'TLB-USD end 2031'
c.font = Font(size=10.0)
c = ws['C88']; c.value = "=C87-MIN(C87,C$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38)-MIN(C87-MIN(C87,C$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38),MAX(0,MAX(0,C76-MIN(C81,C$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38)-MIN(C87,C$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38))-(C81-MIN(C81,C$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38))))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['D88']; c.value = "=D87-MIN(D87,D$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38)-MIN(D87-MIN(D87,D$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38),MAX(0,MAX(0,D76-MIN(D81,D$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38)-MIN(D87,D$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38))-(D81-MIN(D81,D$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38))))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['E88']; c.value = "=E87-MIN(E87,E$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38)-MIN(E87-MIN(E87,E$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38),MAX(0,MAX(0,E76-MIN(E81,E$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38)-MIN(E87,E$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38))-(E81-MIN(E81,E$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38))))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A90']; c.value = 'Excess cash end 2027'
c.font = Font(size=10.0)
c = ws['C90']; c.value = "=0+MAX(0,MAX(0,C72-MIN(C$69*'03 Assumptions'!$B$36,C$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38)-MIN(C$69*'03 Assumptions'!$B$37,C$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38))-(C$69*'03 Assumptions'!$B$36-MIN(C$69*'03 Assumptions'!$B$36,C$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38))-(C$69*'03 Assumptions'!$B$37-MIN(C$69*'03 Assumptions'!$B$37,C$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['D90']; c.value = "=0+MAX(0,MAX(0,D72-MIN(D$69*'03 Assumptions'!$B$36,D$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38)-MIN(D$69*'03 Assumptions'!$B$37,D$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38))-(D$69*'03 Assumptions'!$B$36-MIN(D$69*'03 Assumptions'!$B$36,D$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38))-(D$69*'03 Assumptions'!$B$37-MIN(D$69*'03 Assumptions'!$B$37,D$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['E90']; c.value = "=0+MAX(0,MAX(0,E72-MIN(E$69*'03 Assumptions'!$B$36,E$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38)-MIN(E$69*'03 Assumptions'!$B$37,E$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38))-(E$69*'03 Assumptions'!$B$36-MIN(E$69*'03 Assumptions'!$B$36,E$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38))-(E$69*'03 Assumptions'!$B$37-MIN(E$69*'03 Assumptions'!$B$37,E$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A91']; c.value = 'Excess cash end 2028'
c.font = Font(size=10.0)
c = ws['C91']; c.value = "=C90+MAX(0,MAX(0,C73-MIN(C78,C$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38)-MIN(C84,C$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38))-(C78-MIN(C78,C$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38))-(C84-MIN(C84,C$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['D91']; c.value = "=D90+MAX(0,MAX(0,D73-MIN(D78,D$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38)-MIN(D84,D$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38))-(D78-MIN(D78,D$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38))-(D84-MIN(D84,D$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['E91']; c.value = "=E90+MAX(0,MAX(0,E73-MIN(E78,E$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38)-MIN(E84,E$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38))-(E78-MIN(E78,E$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38))-(E84-MIN(E84,E$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A92']; c.value = 'Excess cash end 2029'
c.font = Font(size=10.0)
c = ws['C92']; c.value = "=C91+MAX(0,MAX(0,C74-MIN(C79,C$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38)-MIN(C85,C$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38))-(C79-MIN(C79,C$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38))-(C85-MIN(C85,C$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['D92']; c.value = "=D91+MAX(0,MAX(0,D74-MIN(D79,D$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38)-MIN(D85,D$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38))-(D79-MIN(D79,D$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38))-(D85-MIN(D85,D$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['E92']; c.value = "=E91+MAX(0,MAX(0,E74-MIN(E79,E$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38)-MIN(E85,E$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38))-(E79-MIN(E79,E$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38))-(E85-MIN(E85,E$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A93']; c.value = 'Excess cash end 2030'
c.font = Font(size=10.0)
c = ws['C93']; c.value = "=C92+MAX(0,MAX(0,C75-MIN(C80,C$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38)-MIN(C86,C$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38))-(C80-MIN(C80,C$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38))-(C86-MIN(C86,C$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['D93']; c.value = "=D92+MAX(0,MAX(0,D75-MIN(D80,D$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38)-MIN(D86,D$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38))-(D80-MIN(D80,D$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38))-(D86-MIN(D86,D$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['E93']; c.value = "=E92+MAX(0,MAX(0,E75-MIN(E80,E$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38)-MIN(E86,E$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38))-(E80-MIN(E80,E$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38))-(E86-MIN(E86,E$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A94']; c.value = 'Excess cash end 2031'
c.font = Font(size=10.0)
c = ws['C94']; c.value = "=C93+MAX(0,MAX(0,C76-MIN(C81,C$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38)-MIN(C87,C$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38))-(C81-MIN(C81,C$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38))-(C87-MIN(C87,C$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['D94']; c.value = "=D93+MAX(0,MAX(0,D76-MIN(D81,D$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38)-MIN(D87,D$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38))-(D81-MIN(D81,D$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38))-(D87-MIN(D87,D$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['E94']; c.value = "=E93+MAX(0,MAX(0,E76-MIN(E81,E$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38)-MIN(E87,E$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38))-(E81-MIN(E81,E$69*'03 Assumptions'!$B$36*'03 Assumptions'!$B$38))-(E87-MIN(E87,E$69*'03 Assumptions'!$B$37*'03 Assumptions'!$B$38)))"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(size=10.0)
c = ws['A97']; c.value = 'GRID 3 - Entry leverage (rows) x exit multiple (cols) - IRR at 2031 exit, EXACT'
c.font = Font(bold=True, color='FF1F3864')
c = ws['C98']; c.value = 10
c.number_format = '0.00#'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['D98']; c.value = 11.5
c.number_format = '0.00#'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['E98']; c.value = 12.9
c.number_format = '0.00#'
c.font = Font(color='FF0000FF', size=10.0)
c = ws['B99']; c.value = '=C68'
c.number_format = '0.00#'
c.font = Font(color='FF008000', size=10.0)
c = ws['C99']; c.value = "=((C$98*'09 Operating Model'!$K$18-(C82+C88+C$69*(1-'03 Assumptions'!$B$36-'03 Assumptions'!$B$37)-C94-'03 Assumptions'!$B$44))/C$70)^(1/5)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['D99']; c.value = "=((D$98*'09 Operating Model'!$K$18-(C82+C88+C$69*(1-'03 Assumptions'!$B$36-'03 Assumptions'!$B$37)-C94-'03 Assumptions'!$B$44))/C$70)^(1/5)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['E99']; c.value = "=((E$98*'09 Operating Model'!$K$18-(C82+C88+C$69*(1-'03 Assumptions'!$B$36-'03 Assumptions'!$B$37)-C94-'03 Assumptions'!$B$44))/C$70)^(1/5)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['B100']; c.value = '=D68'
c.number_format = '0.00#'
c.font = Font(color='FF008000', size=10.0)
c = ws['C100']; c.value = "=((C$98*'09 Operating Model'!$K$18-(D82+D88+D$69*(1-'03 Assumptions'!$B$36-'03 Assumptions'!$B$37)-D94-'03 Assumptions'!$B$44))/D$70)^(1/5)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['D100']; c.value = "=((D$98*'09 Operating Model'!$K$18-(D82+D88+D$69*(1-'03 Assumptions'!$B$36-'03 Assumptions'!$B$37)-D94-'03 Assumptions'!$B$44))/D$70)^(1/5)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['E100']; c.value = "=((E$98*'09 Operating Model'!$K$18-(D82+D88+D$69*(1-'03 Assumptions'!$B$36-'03 Assumptions'!$B$37)-D94-'03 Assumptions'!$B$44))/D$70)^(1/5)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['B101']; c.value = '=E68'
c.number_format = '0.00#'
c.font = Font(color='FF008000', size=10.0)
c = ws['C101']; c.value = "=((C$98*'09 Operating Model'!$K$18-(E82+E88+E$69*(1-'03 Assumptions'!$B$36-'03 Assumptions'!$B$37)-E94-'03 Assumptions'!$B$44))/E$70)^(1/5)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['D101']; c.value = "=((D$98*'09 Operating Model'!$K$18-(E82+E88+E$69*(1-'03 Assumptions'!$B$36-'03 Assumptions'!$B$37)-E94-'03 Assumptions'!$B$44))/E$70)^(1/5)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['E101']; c.value = "=((E$98*'09 Operating Model'!$K$18-(E82+E88+E$69*(1-'03 Assumptions'!$B$36-'03 Assumptions'!$B$37)-E94-'03 Assumptions'!$B$44))/E$70)^(1/5)-1"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['A103']; c.value = 'Engine B equity = base uses ex-OID + 2.5% OID on case debt - case debt. Advisory fee (1.25% EV) is debt-independent.'
c.font = Font(size=10.0)

ws = wb.create_sheet('12 Charts')
ws.column_dimensions['A'].width = 46.0
ws.column_dimensions['B'].width = 12.0
c = ws['A1']; c.value = 'CHARTS'
c.font = Font(bold=True, color='FFFFFFFF', size=12.0)
c.fill = PatternFill("solid", fgColor='FF1F3864')
c = ws['A3']; c.value = 'Chart data (live)'
c.font = Font(bold=True, color='FF1F3864')
c = ws['A4']; c.value = 'Year'
c.font = Font(size=10.0)
c = ws['B4']; c.value = 2027
c.number_format = '#,##0;\\(#,##0\\);\\-'
c.font = Font(size=10.0)
c = ws['C4']; c.value = 2028
c.number_format = '#,##0;\\(#,##0\\);\\-'
c.font = Font(size=10.0)
c = ws['D4']; c.value = 2029
c.number_format = '#,##0;\\(#,##0\\);\\-'
c.font = Font(size=10.0)
c = ws['E4']; c.value = 2030
c.number_format = '#,##0;\\(#,##0\\);\\-'
c.font = Font(size=10.0)
c = ws['F4']; c.value = 2031
c.number_format = '#,##0;\\(#,##0\\);\\-'
c.font = Font(size=10.0)
c = ws['A5']; c.value = 'EBITDA'
c.font = Font(size=10.0)
c = ws['B5']; c.value = "='09 Operating Model'!G18"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF008000', size=10.0)
c = ws['C5']; c.value = "='09 Operating Model'!H18"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF008000', size=10.0)
c = ws['D5']; c.value = "='09 Operating Model'!I18"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF008000', size=10.0)
c = ws['E5']; c.value = "='09 Operating Model'!J18"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF008000', size=10.0)
c = ws['F5']; c.value = "='09 Operating Model'!K18"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF008000', size=10.0)
c = ws['A6']; c.value = 'Net debt'
c.font = Font(size=10.0)
c = ws['B6']; c.value = "='08 Debt Schedule'!G19"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF008000', size=10.0)
c = ws['C6']; c.value = "='08 Debt Schedule'!H19"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF008000', size=10.0)
c = ws['D6']; c.value = "='08 Debt Schedule'!I19"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF008000', size=10.0)
c = ws['E6']; c.value = "='08 Debt Schedule'!J19"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF008000', size=10.0)
c = ws['F6']; c.value = "='08 Debt Schedule'!K19"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF008000', size=10.0)
c = ws['A7']; c.value = 'Leverage'
c.font = Font(size=10.0)
c = ws['B7']; c.value = "='08 Debt Schedule'!G36"
c.number_format = '0.00#'
c.font = Font(color='FF008000', size=10.0)
c = ws['C7']; c.value = "='08 Debt Schedule'!H36"
c.number_format = '0.00#'
c.font = Font(color='FF008000', size=10.0)
c = ws['D7']; c.value = "='08 Debt Schedule'!I36"
c.number_format = '0.00#'
c.font = Font(color='FF008000', size=10.0)
c = ws['E7']; c.value = "='08 Debt Schedule'!J36"
c.number_format = '0.00#'
c.font = Font(color='FF008000', size=10.0)
c = ws['F7']; c.value = "='08 Debt Schedule'!K36"
c.number_format = '0.00#'
c.font = Font(color='FF008000', size=10.0)
c = ws['A9']; c.value = 'MOIC bridge data'
c.font = Font(bold=True, color='FF1F3864')
c = ws['A10']; c.value = 'Entry'
c.font = Font(size=10.0)
c = ws['B10']; c.value = "='10 Returns'!B15"
c.number_format = '0.00#'
c.font = Font(color='FF008000', size=10.0)
c = ws['A11']; c.value = 'EBITDA growth'
c.font = Font(size=10.0)
c = ws['B11']; c.value = "='10 Returns'!B16"
c.number_format = '0.00#'
c.font = Font(color='FF008000', size=10.0)
c = ws['A12']; c.value = 'Multiple'
c.font = Font(size=10.0)
c = ws['B12']; c.value = "='10 Returns'!B17"
c.number_format = '0.00#'
c.font = Font(color='FF008000', size=10.0)
c = ws['A13']; c.value = 'Deleveraging'
c.font = Font(size=10.0)
c = ws['B13']; c.value = "='10 Returns'!B18"
c.number_format = '0.00#'
c.font = Font(color='FF008000', size=10.0)
c = ws['A14']; c.value = 'Fees/resid'
c.font = Font(size=10.0)
c = ws['B14']; c.value = "='10 Returns'!B19"
c.number_format = '0.00#'
c.font = Font(color='FF008000', size=10.0)

ws = wb.create_sheet('13 Dashboard')
ws.column_dimensions['A'].width = 46.0
ws.column_dimensions['B'].width = 12.0
c = ws['A1']; c.value = 'DASHBOARD + COMPARISON vs ACTUAL TRANSACTION'
c.font = Font(bold=True, color='FFFFFFFF', size=12.0)
c.fill = PatternFill("solid", fgColor='FF1F3864')
c = ws['A3']; c.value = 'KEY OUTPUTS (active scenario)'
c.font = Font(bold=True, color='FF1F3864')
c = ws['A4']; c.value = 'IRR (5-yr, 2031 exit)'
c.font = Font(size=10.0)
c = ws['B4']; c.value = "='10 Returns'!B11"
c.number_format = '0.0%'
c.font = Font(color='FF008000', size=10.0)
c = ws['A5']; c.value = 'MOIC'
c.font = Font(size=10.0)
c = ws['B5']; c.value = "='10 Returns'!B9"
c.number_format = '0.00#'
c.font = Font(color='FF008000', size=10.0)
c = ws['A6']; c.value = 'Entry EV / FY25A EBITDA'
c.font = Font(size=10.0)
c = ws['B6']; c.value = "='07 Purchase Price'!B8"
c.number_format = '0.00#'
c.font = Font(color='FF008000', size=10.0)
c = ws['A7']; c.value = 'Entry leverage'
c.font = Font(size=10.0)
c = ws['B7']; c.value = "='03 Assumptions'!F10"
c.number_format = '0.00#'
c.font = Font(color='FF008000', size=10.0)
c = ws['A8']; c.value = 'Sponsor equity (€m)'
c.font = Font(size=10.0)
c = ws['B8']; c.value = "='06 Sources & Uses'!B18"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF008000', size=10.0)
c = ws['A9']; c.value = 'Total debt (€m)'
c.font = Font(size=10.0)
c = ws['B9']; c.value = "='06 Sources & Uses'!B13"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF008000', size=10.0)
c = ws['A10']; c.value = 'Net debt at exit (€m)'
c.font = Font(size=10.0)
c = ws['B10']; c.value = "='08 Debt Schedule'!K19"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF008000', size=10.0)
c = ws['A11']; c.value = 'Debt paydown over hold (EUR m)'
c.font = Font(size=10.0)
c = ws['B11']; c.value = "='06 Sources & Uses'!B13-'08 Debt Schedule'!K17"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF008000', size=10.0)
c = ws['A12']; c.value = 'Exit EBITDA (EUR m, incl. BD)'
c.font = Font(size=10.0)
c = ws['B12']; c.value = "='09 Operating Model'!K18"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF008000', size=10.0)
c = ws['A13']; c.value = 'Year-1 interest coverage'
c.font = Font(size=10.0)
c = ws['B13']; c.value = "='08 Debt Schedule'!G37"
c.number_format = '0.00#'
c.font = Font(color='FF008000', size=10.0)
c = ws['A16']; c.value = 'INDEPENDENT MODEL vs ACTUAL ANNOUNCED DEAL'
c.font = Font(bold=True, color='FF1F3864')
c = ws['A17']; c.value = 'Metric'
c.font = Font(bold=True, color='FF1F3864')
c = ws['B17']; c.value = 'Our base'
c.font = Font(bold=True, color='FF1F3864')
c = ws['C17']; c.value = 'Actual deal'
c.font = Font(bold=True, color='FF1F3864')
c = ws['D17']; c.value = 'Comment'
c.font = Font(bold=True, color='FF1F3864')
c = ws['A18']; c.value = 'Entry EV/EBITDA'
c.font = Font(size=10.0)
c = ws['B18']; c.value = "='07 Purchase Price'!B8"
c.number_format = '0.00#'
c.font = Font(color='FF008000', size=10.0)
c = ws['C18']; c.value = '12.9x'
c.font = Font(size=10.0)
c = ws['D18']; c.value = 'Identical — price is a deal fact; we underwrite at it'
c.font = Font(size=10.0)
c = ws['A19']; c.value = 'Premium'
c.font = Font(size=10.0)
c = ws['B19']; c.value = '—'
c.font = Font(size=10.0)
c = ws['C19']; c.value = '12.89%'
c.font = Font(size=10.0)
c = ws['D19']; c.value = 'Low vs Italian norms (20-30%): Rossini 46.82% pre-committed'
c.font = Font(size=10.0)
c = ws['A20']; c.value = 'Debt quantum'
c.font = Font(size=10.0)
c = ws['B20']; c.value = "='06 Sources & Uses'!B13"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF008000', size=10.0)
c = ws['C20']; c.value = '€5.5-6.0bn sought'
c.font = Font(size=10.0)
c = ws['D20']; c.value = 'Our 5.5x = €5.45bn sits at bottom of actual range'
c.font = Font(size=10.0)
c = ws['A21']; c.value = 'Blended debt cost'
c.font = Font(size=10.0)
c = ws['B21']; c.value = '~6.3%'
c.font = Font(size=10.0)
c = ws['C21']; c.value = 'TBD (offer doc)'
c.font = Font(size=10.0)
c = ws['D21']; c.value = 'E+375 TLB / 6.75% SSN assumptions'
c.font = Font(size=10.0)
c = ws['A22']; c.value = 'Equity cheque'
c.font = Font(size=10.0)
c = ws['B22']; c.value = "='06 Sources & Uses'!B18"
c.number_format = '#,##0.0;\\(#,##0.0\\);\\-'
c.font = Font(color='FF008000', size=10.0)
c = ws['C22']; c.value = '~€7.5bn+ (club)'
c.font = Font(size=10.0)
c = ws['D22']; c.value = 'GBL ≤€1.3bn + CVC IX + ADIA/CPPIB/PSP et al.'
c.font = Font(size=10.0)
c = ws['A23']; c.value = 'Exit assumption'
c.font = Font(size=10.0)
c = ws['B23']; c.value = '11.5x (below entry)'
c.font = Font(size=10.0)
c = ws['C23']; c.value = 'n/a (private)'
c.font = Font(size=10.0)
c = ws['D23']; c.value = 'We refuse multiple expansion; Rossini precedent = 0 expansion'
c.font = Font(size=10.0)
c = ws['A24']; c.value = 'BD/M&A in base'
c.font = Font(size=10.0)
c = ws['B24']; c.value = 'EUR 400m/yr @ 9.0x (funded pre-sweep)'
c.font = Font(size=10.0)
c = ws['C24']; c.value = 'Core to thesis'
c.font = Font(size=10.0)
c = ws['D24']; c.value = 'Their 20%+ case needs the bolt-on machine — our bull case'
c.font = Font(size=10.0)
c = ws['A26']; c.value = "WHAT MUST CHANGE FOR OUR MODEL TO CONVERGE TO CVC/GBL's ACCEPTED ECONOMICS:"
c.font = Font(bold=True, color='FF1F3864')
c = ws['A27']; c.value = 'v3 BASE funds BD (EUR 400m/yr 2027-2030 at 9x; 2031 exit-year spend removed). No-BD conservative floor via F13=0.'
c.font = Font(size=10.0)
c = ws['A28']; c.value = '1. Exit at entry-flat 12.9x instead of 11.5x: ~+2.9pt IRR (the dominant lever). ~2.0-2.5pt per turn.'
c.font = Font(size=10.0)
c = ws['A29']; c.value = '2. Fund BD redeployment (€500m/yr at 8-10x): converts SPC cash into RD growth — our bull case, their base case.'
c.font = Font(size=10.0)
c = ws['A30']; c.value = '3. Accept 6.0x opening leverage (top of their range): ~+1pt IRR if coverage holds.'
c.font = Font(size=10.0)
c = ws['A31']; c.value = '4. Underwrite Isturisa to raised >€1.2bn peak (we hold RD fade): adds 2-3pts EBITDA CAGR.'
c.font = Font(size=10.0)
c = ws['A32']; c.value = 'BD VALUE-CREATION DECOMPOSITION (v3): no-BD floor ~15% IRR; + funding BD 400m/yr ~0 to -0.1pt (arbitrage offset by interest drag);'
c.font = Font(size=10.0)
c = ws['A33']; c.value = '+ exit at entry-flat 12.9x = +2.9pt. The exit multiple is ~100% of the sponsor-consistent case; BD is a near-wash on IRR (it grows MOIC dollars, not the rate).'
c.font = Font(size=10.0)

wb.save('Recordati_LBO_Model.xlsx')
print("wrote", 'Recordati_LBO_Model.xlsx')
"""
Deterministic reconstruction of the Technogym DCF model (15 sheets).
Auto-generated from the shipped workbook; running it reproduces the model byte-for-formula.
Requires: pip install openpyxl.   Run: python build.py   ->  writes the .xlsx next to this file.
Scenario switch lives at 03 Assumptions!B3  (1=Bear, 2=Base, 3=Bull).
"""
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
wb.remove(wb.active)
try: wb.calculation.fullCalcOnLoad = True
except Exception: pass

ws = wb.create_sheet('00 Executive Summary')
ws.column_dimensions['A'].width = 40.0
ws.column_dimensions['B'].width = 13.0
ws.column_dimensions['C'].width = 13.0
ws.column_dimensions['D'].width = 13.0
ws.column_dimensions['E'].width = 13.0
ws.column_dimensions['F'].width = 13.0
ws.column_dimensions['G'].width = 13.0
ws.column_dimensions['H'].width = 13.0
ws.column_dimensions['I'].width = 13.0
ws.column_dimensions['J'].width = 13.0
ws.column_dimensions['K'].width = 13.0
c = ws['A1']; c.value = 'TECHNOGYM S.p.A. (BIT: TGYM) - DCF EXECUTIVE SUMMARY | v2.0 | 17 JULY 2026'
c.font = Font(bold=True, color='00FFFFFF', size=12.0)
c.fill = PatternFill("solid", fgColor='001F3864')
c = ws['A3']; c.value = 'THE ANSWER'
c.font = Font(bold=True, color='001F3864', size=14.0)
c = ws['A4']; c.value = 'Base-case fair value (Gordon TV):'
c.font = Font(bold=True, color='001F3864')
c = ws['B4']; c.value = "='11 DCF Valuation'!B27"
c.number_format = '#,##0.00'
c.font = Font(color='00008000', size=10.0)
c.fill = PatternFill("solid", fgColor='00FFFF00')
c = ws['A5']; c.value = 'Cross-check (exit-multiple TV):'
c.font = Font(size=10.0)
c = ws['B5']; c.value = "='11 DCF Valuation'!B31"
c.number_format = '#,##0.00'
c.font = Font(color='00008000', size=10.0)
c = ws['A6']; c.value = 'Current price (08-Jul-26):'
c.font = Font(size=10.0)
c = ws['B6']; c.value = "='11 DCF Valuation'!B28"
c.number_format = '#,##0.00'
c.font = Font(color='00008000', size=10.0)
c = ws['A7']; c.value = 'Upside / (downside):'
c.font = Font(bold=True, color='001F3864')
c = ws['B7']; c.value = "='11 DCF Valuation'!B29"
c.number_format = '0.0%'
c.font = Font(color='00008000', size=10.0)
c.fill = PatternFill("solid", fgColor='00FFFF00')
c = ws['A8']; c.value = 'Scenario range (full model, recorded 16-Jul-26):'
c.font = Font(size=10.0)
c = ws['B8']; c.value = 'Bear EUR 8.82'
c.font = Font(color='000000FF', size=10.0)
c = ws['C8']; c.value = 'Bull EUR 24.05'
c.font = Font(color='000000FF', size=10.0)
c = ws['D8']; c.value = '(re-run switch on 03 Assumptions to refresh)'
c.font = Font(size=10.0)
c = ws['A10']; c.value = 'WHAT DRIVES THE VALUE'
c.font = Font(bold=True, color='001F3864', size=14.0)
c = ws['A11']; c.value = '1. BtoB growth path (81% of revenue): base fades 9% → 2.5% by 2035; every 1pp ≈ €2.5–2.9m EBITDA at incremental margins.'
c.font = Font(size=10.0)
c = ws['A12']; c.value = '2. EBITDA margin endpoint: 21.6% today → 23.5% base by 2032. Four consecutive years of expansion support it; bear case tests a stall.'
c.font = Font(size=10.0)
c = ws['A13']; c.value = '3. Digital is an OPTION, not the base: est. €45m (4.4% of revenue) today, ~7% by 2035 in base. Bull case (attach 20%→40%) is what consensus prices.'
c.font = Font(size=10.0)
c = ws['A15']; c.value = 'KEY ASSUMPTIONS (speed-read — full detail on 03)'
c.font = Font(bold=True, color='001F3864', size=14.0)
c = ws['A16']; c.value = 'Revenue CAGR 2026-35'
c.font = Font(size=10.0)
c = ws['B16']; c.value = "=('05 Revenue Build'!O27/'05 Revenue Build'!E27)^(1/10)-1"
c.number_format = '0.0%'
c.font = Font(color='00008000', size=10.0)
c = ws['A17']; c.value = 'EBITDA margin endpoint'
c.font = Font(size=10.0)
c = ws['B17']; c.value = "='03 Assumptions'!F13"
c.number_format = '0.0%'
c.font = Font(color='00008000', size=10.0)
c = ws['A18']; c.value = 'Terminal growth g'
c.font = Font(size=10.0)
c = ws['B18']; c.value = "='03 Assumptions'!F16"
c.number_format = '0.0%'
c.font = Font(color='00008000', size=10.0)
c = ws['A19']; c.value = 'WACC'
c.font = Font(size=10.0)
c = ws['B19']; c.value = "='10 WACC'!B15"
c.number_format = '0.0%'
c.font = Font(color='00008000', size=10.0)
c = ws['A20']; c.value = 'Exit multiple (cross-check)'
c.font = Font(size=10.0)
c = ws['B20']; c.value = "='03 Assumptions'!F17"
c.number_format = '0.0x'
c.font = Font(color='00008000', size=10.0)
c = ws['A21']; c.value = 'Tax rate'
c.font = Font(size=10.0)
c = ws['B21']; c.value = "='03 Assumptions'!B21"
c.number_format = '0.0%'
c.font = Font(color='00008000', size=10.0)
c = ws['A22']; c.value = 'TV % of EV'
c.font = Font(size=10.0)
c = ws['B22']; c.value = "='11 DCF Valuation'!B32"
c.number_format = '0.0%'
c.font = Font(color='00008000', size=10.0)
c = ws['A24']; c.value = "BRIDGE TO CONSENSUS (why we differ from the Street's €19.3 avg TP)"
c.font = Font(bold=True, color='001F3864', size=14.0)
c = ws['A25']; c.value = 'Consensus average TP'
c.font = Font(size=10.0)
c = ws['B25']; c.value = 19.3
c.number_format = '#,##0.00'
c.font = Font(color='000000FF', size=10.0)
c = ws['A26']; c.value = 'Our base-case DCF'
c.font = Font(size=10.0)
c = ws['B26']; c.value = "='11 DCF Valuation'!B27"
c.number_format = '#,##0.00'
c.font = Font(color='00008000', size=10.0)
c = ws['A27']; c.value = 'Gap'
c.font = Font(size=10.0)
c = ws['B27']; c.value = '=B26-B25'
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['A28']; c.value = 'Attribution (analyst estimates, blue):'
c.font = Font(size=10.0)
c = ws['A29']; c.value = '  Margin endpoint (Street ~24.5% vs our 23.5% total)'
c.font = Font(size=10.0)
c = ws['B29']; c.value = -1.6
c.number_format = '#,##0.00'
c.font = Font(color='000000FF', size=10.0)
c = ws['A30']; c.value = '  Digital ramp (Street implicitly faster / higher attach)'
c.font = Font(size=10.0)
c = ws['B30']; c.value = -1.2
c.number_format = '#,##0.00'
c.font = Font(color='000000FF', size=10.0)
c = ws['A31']; c.value = '  Growth path + WACC/terminal (residual)'
c.font = Font(size=10.0)
c = ws['B31']; c.value = -0.78
c.number_format = '#,##0.00'
c.font = Font(color='000000FF', size=10.0)
c = ws['A32']; c.value = '=IF(ABS(B29+B30+B31-B27)<0.5,"CHECK OK: attribution ties to gap","CHECK: attribution does not tie")'
c.font = Font(bold=True, color='00FF0000', size=10.0)
c = ws['A34']; c.value = 'FOOTBALL FIELD — value per share (€)'
c.font = Font(bold=True, color='001F3864', size=14.0)
c = ws['A35']; c.value = 'Method'
c.font = Font(bold=True, color='001F3864')
c = ws['B35']; c.value = 'Low'
c.font = Font(bold=True, color='001F3864')
c = ws['C35']; c.value = 'High'
c.font = Font(bold=True, color='001F3864')
c = ws['D35']; c.value = 'Range'
c.font = Font(bold=True, color='001F3864')
c = ws['A36']; c.value = '52-week trading range'
c.font = Font(size=10.0)
c = ws['B36']; c.value = 11
c.number_format = '#,##0.00'
c.font = Font(color='000000FF', size=10.0)
c = ws['C36']; c.value = 21.82
c.number_format = '#,##0.00'
c.font = Font(color='000000FF', size=10.0)
c = ws['D36']; c.value = '=C36-B36'
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['E36']; c.value = '<- Current EUR 15.19 and DCF base EUR 15.72 fall inside every range except the bear-scenario tail.'
c.font = Font(size=10.0)
c = ws['A37']; c.value = 'Comps 10.5x-14x 2026E EBITDA'
c.font = Font(size=10.0)
c = ws['B37']; c.value = "=(10.5*'06 Operating Model'!F6+'03 Assumptions'!B24)/'03 Assumptions'!B23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['C37']; c.value = "=(14*'06 Operating Model'!F6+'03 Assumptions'!B24)/'03 Assumptions'!B23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['D37']; c.value = '=C37-B37'
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['A38']; c.value = 'DCF grid (WACC 7-8.5% x g)'
c.font = Font(size=10.0)
c = ws['B38']; c.value = "=MIN('12 Sensitivity Analysis'!C6:G11)"
c.number_format = '#,##0.00'
c.font = Font(color='00008000', size=10.0)
c = ws['C38']; c.value = "=MAX('12 Sensitivity Analysis'!C6:G11)"
c.number_format = '#,##0.00'
c.font = Font(color='00008000', size=10.0)
c = ws['D38']; c.value = '=C38-B38'
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['A39']; c.value = 'DCF scenarios Bear-Bull'
c.font = Font(size=10.0)
c = ws['B39']; c.value = 8.82
c.number_format = '#,##0.00'
c.font = Font(color='000000FF', size=10.0)
c = ws['C39']; c.value = 22.42
c.number_format = '#,##0.00'
c.font = Font(color='000000FF', size=10.0)
c = ws['D39']; c.value = '=C39-B39'
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['A40']; c.value = 'Consensus TP range'
c.font = Font(size=10.0)
c = ws['B40']; c.value = 16
c.number_format = '#,##0.00'
c.font = Font(color='000000FF', size=10.0)
c = ws['C40']; c.value = 26.25
c.number_format = '#,##0.00'
c.font = Font(color='000000FF', size=10.0)
c = ws['D40']; c.value = '=C40-B40'
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['A41']; c.value = 'Current price'
c.font = Font(size=10.0)
c = ws['B41']; c.value = "='11 DCF Valuation'!B28"
c.number_format = '#,##0.00'
c.font = Font(color='00008000', size=10.0)
c = ws['A44']; c.value = 'HOW TO USE THIS MODEL'
c.font = Font(bold=True, color='001F3864', size=14.0)
c = ws['A45']; c.value = "1. Scenario switch: edit '03 Assumptions'!B3 (1=Bear, 2=Base, 3=Bull). Everything recalculates."
c.font = Font(size=10.0)
c = ws['A46']; c.value = '2. BLUE cells are editable inputs; BLACK are formulas; GREEN are cross-sheet links; RED are checks.'
c.font = Font(size=10.0)
c = ws['A47']; c.value = '3. If any RED check cell reads ERROR, stop and review before trusting outputs.'
c.font = Font(size=10.0)
c = ws['A48']; c.value = '4. Yellow-filled cells are the primary levers. Dashboard (14) and this page carry the headline outputs.'
c.font = Font(size=10.0)
c = ws['A49']; c.value = '5. Digital revenue is an explicit L-confidence ESTIMATE (company does not disclose) — see 05 rows 14-25.'
c.font = Font(size=10.0)
c = ws['A51']; c.value = 'GLOSSARY'
c.font = Font(bold=True, color='001F3864', size=14.0)
c = ws['A52']; c.value = 'FCFF'
c.font = Font(bold=True, color='001F3864')
c = ws['B52']; c.value = 'Free Cash Flow to Firm = NOPAT + D&A − Capex − ΔNWC; cash available to all capital providers'
c.font = Font(size=10.0)
c = ws['A53']; c.value = 'NOPAT'
c.font = Font(bold=True, color='001F3864')
c = ws['B53']; c.value = 'Net Operating Profit After Tax = EBIT × (1 − tax rate)'
c.font = Font(size=10.0)
c = ws['A54']; c.value = 'NWC'
c.font = Font(bold=True, color='001F3864')
c = ws['B54']; c.value = 'Net Working Capital = Inventory + Trade Receivables − Trade Payables'
c.font = Font(size=10.0)
c = ws['A55']; c.value = 'WACC'
c.font = Font(bold=True, color='001F3864')
c = ws['B55']; c.value = 'Weighted Average Cost of Capital — discount rate for FCFF; here ≈ cost of equity (net-cash company)'
c.font = Font(size=10.0)
c = ws['A56']; c.value = 'ERP / CRP'
c.font = Font(bold=True, color='001F3864')
c = ws['B56']; c.value = 'Equity Risk Premium (mature market) / Country Risk Premium (sovereign-risk add-on, exposure-weighted)'
c.font = Font(size=10.0)
c = ws['A57']; c.value = 'Gordon TV'
c.font = Font(bold=True, color='001F3864')
c = ws['B57']; c.value = 'Terminal value = FCFF₁₀ × (1+g) / (WACC − g); values all cash flows beyond 2035'
c.font = Font(size=10.0)
c = ws['A58']; c.value = 'Exit multiple TV'
c.font = Font(bold=True, color='001F3864')
c = ws['B58']; c.value = 'Alternative terminal value = 2035 EBITDA × multiple; cross-checks Gordon (two methods should broadly agree)'
c.font = Font(size=10.0)
c = ws['A59']; c.value = 'TV % of EV'
c.font = Font(bold=True, color='001F3864')
c = ws['B59']; c.value = 'Share of value from the terminal period; >80% means the answer rests on terminal assumptions (ours ≈ 63%)'
c.font = Font(size=10.0)
c = ws['A61']; c.value = 'RECOMMENDED ACTION'
c.font = Font(bold=True, color='001F3864', size=14.0)
c = ws['A62']; c.value = 'HOLD (v3). Fair value ~EUR 15.72 Gordon / ~EUR 18.3 exit-mult vs EUR 15.19 (+3.5%): below the >20% BUY hurdle. (Live: 00!B4.)'
c.font = Font(bold=True, size=10.0)
c.fill = PatternFill("solid", fgColor='00FFF2CC')
c = ws['A63']; c.value = 'Buy trigger < EUR 14.50; 5-yr horizon cross-check ~EUR 17.7 (live 11!B39; TV ~83% of EV - shown for transparency).'
c.font = Font(size=10.0)
c = ws['A64']; c.value = 'v2 changes: IFRS16 lease principal charged in FCFF (+NFP excl leases in bridge, -~EUR 1.6/sh); NWC on DSO/DIO/DPO days;'
c.font = Font(size=10.0)
c = ws['A65']; c.value = 'bear = explicit 2027 cyclical downturn (EBITDA -18% trough) not a linear fade. Rating rule: >20% upside required for BUY.'
c.font = Font(size=10.0)
c = ws['A66']; c.value = 'v3: digital P&L ISOLATED (hardware margin expands on operating leverage to ~20.75%; digital carries its own ~60% margin). Digital option worth ~EUR 3.3/sh (base->bull-digital), vs ~EUR 0.6 when it was blended. Base total margin held at ~23.5%.'
c = ws['A67']; c.value = 'KEY RISKS (ranked by expected value impact)'
c.font = Font(bold=True, color='001F3864', size=14.0)
c = ws['A68']; c.value = '1. BtoB capex cycle turn 2026-27: 5pp growth deceleration ≈ −€55-60m EBITDA by yr 5 ≈ −€3.5-4.0/share.'
c.font = Font(size=10.0)
c = ws['A69']; c.value = '2. Margin stall at ~21.5% (bear): mean reversion after 4 years of expansion → value falls to ~€11.'
c.font = Font(size=10.0)
c = ws['A70']; c.value = '3. Digital thesis failure / continued non-disclosure: removes the bull option; base case largely unaffected.'
c.font = Font(size=10.0)
c = ws['A71']; c.value = '4. Americas/tariffs (17% of revenue, Q1-26 +2.5%): structural headwind if the tariff-deferral hypothesis is wrong.'
c.font = Font(size=10.0)
c = ws['A72']; c.value = '5. Key-man (founder-CEO-chairman, family control) and FX (USD/EM).'
c.font = Font(size=10.0)
c = ws['A74']; c.value = 'Monitoring triggers: H1-26 print (30-Jul-26), first digital KPI disclosure, Americas re-acceleration.'
c.font = Font(size=10.0)

ws = wb.create_sheet('01 Cover')
ws.column_dimensions['A'].width = 46.0
ws.column_dimensions['B'].width = 12.0
ws.column_dimensions['C'].width = 12.0
ws.column_dimensions['D'].width = 12.0
ws.column_dimensions['E'].width = 12.0
ws.column_dimensions['F'].width = 12.0
ws.column_dimensions['G'].width = 12.0
ws.column_dimensions['H'].width = 12.0
ws.column_dimensions['I'].width = 12.0
ws.column_dimensions['J'].width = 12.0
ws.column_dimensions['K'].width = 12.0
ws.column_dimensions['L'].width = 12.0
ws.column_dimensions['M'].width = 12.0
ws.column_dimensions['N'].width = 12.0
ws.column_dimensions['O'].width = 12.0
ws.column_dimensions['P'].width = 12.0
ws.column_dimensions['Q'].width = 12.0
c = ws['A1']; c.value = 'TECHNOGYM S.p.A. — DCF MODEL'
c.font = Font(bold=True, color='00FFFFFF', size=12.0)
c.fill = PatternFill("solid", fgColor='001F3864')
c = ws['A2']; c.value = 'TECHNOGYM S.p.A. (BIT: TGYM)'
c.font = Font(bold=True, size=14.0)
c = ws['A3']; c.value = 'Discounted Cash Flow Valuation — Institutional Model'
c.font = Font(bold=True)
c = ws['A5']; c.value = 'Investment question: Can Technogym transition from premium fitness equipment'
c.font = Font(size=10.0)
c = ws['A6']; c.value = 'manufacturer to a recurring software & connected wellness platform deserving software multiples?'
c.font = Font(size=10.0)
c = ws['A8']; c.value = 'Prepared: 17 July 2026 | v2.0 (partner revision round) | Analyst: Tanmay Gambhir | EUR millions'
c.font = Font(size=10.0)
c = ws['A10']; c.value = 'Sheet map:'
c.font = Font(bold=True)
c = ws['A11']; c.value = '02 Sources'
c.font = Font(color='00008000', size=10.0)
c = ws['D11']; c.value = 'Every figure traced: source, date, confidence'
c.font = Font(size=10.0)
c = ws['A12']; c.value = '03 Assumptions'
c.font = Font(color='00008000', size=10.0)
c = ws['D12']; c.value = 'ALL editable inputs (blue). Scenario switch drives the model'
c.font = Font(size=10.0)
c = ws['A13']; c.value = '04 Historical Financials'
c.font = Font(color='00008000', size=10.0)
c = ws['D13']; c.value = 'FY2022–25 reported (company PRs)'
c.font = Font(size=10.0)
c = ws['A14']; c.value = '05 Revenue Build'
c.font = Font(color='00008000', size=10.0)
c = ws['D14']; c.value = 'Driver-based: BtoB / BtoC / digital estimation block'
c.font = Font(size=10.0)
c = ws['A15']; c.value = '06 Operating Model'
c.font = Font(color='00008000', size=10.0)
c = ws['D15']; c.value = 'Margins → EBITDA → EBIT → NOPAT'
c.font = Font(size=10.0)
c = ws['A16']; c.value = '07 Working Capital'
c.font = Font(color='00008000', size=10.0)
c = ws['D16']; c.value = 'NWC % revenue method'
c.font = Font(size=10.0)
c = ws['A17']; c.value = '08 Capex & Depreciation'
c.font = Font(color='00008000', size=10.0)
c = ws['D17']; c.value = '% revenue with normalisation'
c.font = Font(size=10.0)
c = ws['A18']; c.value = '09 Free Cash Flow'
c.font = Font(color='00008000', size=10.0)
c = ws['D18']; c.value = 'FCFF bridge'
c.font = Font(size=10.0)
c = ws['A19']; c.value = '10 WACC'
c.font = Font(color='00008000', size=10.0)
c = ws['D19']; c.value = 'Bund-based build (no CRP double-count)'
c.font = Font(size=10.0)
c = ws['A20']; c.value = '11 DCF Valuation'
c.font = Font(color='00008000', size=10.0)
c = ws['D20']; c.value = 'Gordon + exit-multiple cross-check'
c.font = Font(size=10.0)
c = ws['A21']; c.value = '12 Sensitivity Analysis'
c.font = Font(color='00008000', size=10.0)
c = ws['D21']; c.value = 'WACC×g, WACC×exit, margin×growth'
c.font = Font(size=10.0)
c = ws['A22']; c.value = '13 Scenarios'
c.font = Font(color='00008000', size=10.0)
c = ws['D22']; c.value = 'Bear/Base/Bull side-by-side (simplified engine)'
c.font = Font(size=10.0)
c = ws['A23']; c.value = '14 Charts & Dashboard'
c.font = Font(color='00008000', size=10.0)
c = ws['D23']; c.value = 'Outputs + charts'
c.font = Font(size=10.0)
c = ws['A26']; c.value = 'Color code: BLUE = input (edit me) | BLACK = formula | GREEN = cross-sheet link | RED = check'
c.font = Font(size=10.0)
c = ws['A27']; c.value = 'Legend: yellow-filled cells are the primary levers to edit (scenario switch, key assumptions).'
c.font = Font(size=10.0)

ws = wb.create_sheet('02 Sources')
ws.column_dimensions['A'].width = 46.0
ws.column_dimensions['B'].width = 18.0
ws.column_dimensions['C'].width = 12.0
ws.column_dimensions['D'].width = 90.0
ws.column_dimensions['E'].width = 12.0
ws.column_dimensions['F'].width = 12.0
ws.column_dimensions['G'].width = 12.0
ws.column_dimensions['H'].width = 12.0
ws.column_dimensions['I'].width = 12.0
ws.column_dimensions['J'].width = 12.0
ws.column_dimensions['K'].width = 12.0
ws.column_dimensions['L'].width = 12.0
ws.column_dimensions['M'].width = 12.0
ws.column_dimensions['N'].width = 12.0
ws.column_dimensions['O'].width = 12.0
ws.column_dimensions['P'].width = 12.0
ws.column_dimensions['Q'].width = 12.0
c = ws['A1']; c.value = 'SOURCE REGISTER'
c.font = Font(bold=True, color='00FFFFFF', size=12.0)
c.fill = PatternFill("solid", fgColor='001F3864')
c = ws['A3']; c.value = 'Item'
c.font = Font(bold=True)
c = ws['B3']; c.value = 'Value'
c.font = Font(bold=True)
c = ws['C3']; c.value = 'Confidence'
c.font = Font(bold=True)
c = ws['D3']; c.value = 'Source / date'
c.font = Font(bold=True)
c = ws['A4']; c.value = 'FY2025 revenue / EBITDA adj / NP adj'
c.font = Font(size=10.0)
c = ws['B4']; c.value = '1,019.3 / 220.1 / 119.9'
c.font = Font(size=10.0)
c = ws['C4']; c.value = 'H'
c.font = Font(size=10.0)
c = ws['D4']; c.value = 'Technogym FY25 results PR, 19-Mar-2026 (corporate.technogym.com)'
c.font = Font(size=10.0)
c = ws['A5']; c.value = 'Segment splits, geography, channels'
c.font = Font(size=10.0)
c = ws['B5']; c.value = 'see 04'
c.font = Font(size=10.0)
c = ws['C5']; c.value = 'H'
c.font = Font(size=10.0)
c = ws['D5']; c.value = 'FY25 PR tables'
c.font = Font(size=10.0)
c = ws['A6']; c.value = 'FY2022-24 revenue & EBITDA'
c.font = Font(size=10.0)
c = ws['B6']; c.value = '721.5/808.1/901.3; 132/152/178.4'
c.font = Font(size=10.0)
c = ws['C6']; c.value = 'H'
c.font = Font(size=10.0)
c = ws['D6']; c.value = 'FY23 & FY25 PRs'
c.font = Font(size=10.0)
c = ws['A7']; c.value = 'Net financial position (incl IFRS16)'
c.font = Font(size=10.0)
c = ws['B7']; c.value = '+156.0'
c.font = Font(size=10.0)
c = ws['C7']; c.value = 'H'
c.font = Font(size=10.0)
c = ws['D7']; c.value = 'FY25 PR'
c.font = Font(size=10.0)
c = ws['A8']; c.value = 'Shares out (ex-treasury)'
c.font = Font(size=10.0)
c = ws['B8']; c.value = '199.3m'
c.font = Font(size=10.0)
c = ws['C8']; c.value = 'H'
c.font = Font(size=10.0)
c = ws['D8']; c.value = 'FY25 PR: 201.3m total less 2.036m treasury'
c.font = Font(size=10.0)
c = ws['A9']; c.value = 'Share price'
c.font = Font(size=10.0)
c = ws['B9']; c.value = '15.19 (08-Jul-26)'
c.font = Font(size=10.0)
c = ws['C9']; c.value = 'H'
c.font = Font(size=10.0)
c = ws['D9']; c.value = 'Investing.com'
c.font = Font(size=10.0)
c = ws['A10']; c.value = 'Q1 2026 revenue'
c.font = Font(size=10.0)
c = ws['B10']; c.value = '236.8 (+10.1%)'
c.font = Font(size=10.0)
c = ws['C10']; c.value = 'H'
c.font = Font(size=10.0)
c = ws['D10']; c.value = 'Q1-26 IR presentation, 6-May-26'
c.font = Font(size=10.0)
c = ws['A11']; c.value = 'Bund 10Y'
c.font = Font(size=10.0)
c = ws['B11']; c.value = '3.13%'
c.font = Font(size=10.0)
c = ws['C11']; c.value = 'H'
c.font = Font(size=10.0)
c = ws['D11']; c.value = 'TradingEconomics 16-Jul-26'
c.font = Font(size=10.0)
c = ws['A12']; c.value = 'Mature ERP / Italy CRP'
c.font = Font(size=10.0)
c = ws['B12']; c.value = '4.23% / 2.46%'
c.font = Font(size=10.0)
c = ws['C12']; c.value = 'H'
c.font = Font(size=10.0)
c = ws['D12']; c.value = 'Damodaran, Jan-2026 update'
c.font = Font(size=10.0)
c = ws['A13']; c.value = 'Beta (levered)'
c.font = Font(size=10.0)
c = ws['B13']; c.value = '0.95 triangulated'
c.font = Font(size=10.0)
c = ws['C13']; c.value = 'M'
c.font = Font(size=10.0)
c = ws['D13']; c.value = 'Yahoo 0.86 5Y-M; TradingView 0.94; range 0.80-1.16'
c.font = Font(size=10.0)
c = ws['A14']; c.value = 'Digital revenue 2025'
c.font = Font(size=10.0)
c = ws['B14']; c.value = '~45 (ESTIMATE)'
c.font = Font(size=10.0)
c = ws['C14']; c.value = 'L'
c.font = Font(size=10.0)
c = ws['D14']; c.value = 'Estimation block: 100k sites/500k homes x attach x ARPU. NOT DISCLOSED by company'
c.font = Font(size=10.0)
c = ws['A15']; c.value = 'Digital split of 2025 segments'
c.font = Font(size=10.0)
c = ws['B15']; c.value = 'B2B 36 / B2C 9'
c.font = Font(size=10.0)
c = ws['C15']; c.value = 'L'
c.font = Font(size=10.0)
c = ws['D15']; c.value = 'Derived from estimation block; carved out of reported segments'
c.font = Font(size=10.0)
c = ws['A16']; c.value = 'Effective tax rate'
c.font = Font(size=10.0)
c = ws['B16']; c.value = '27.4%'
c.font = Font(size=10.0)
c = ws['C16']; c.value = 'H'
c.font = Font(size=10.0)
c = ws['D16']; c.value = 'FY25 PR: tax 43.7 on PBT 159.7'
c.font = Font(size=10.0)
c = ws['A17']; c.value = 'Comps: JHT/Amer/Garmin/Ipsen'
c.font = Font(size=10.0)
c = ws['B17']; c.value = '10.5x/20.1x/19.8x/8.4x'
c.font = Font(size=10.0)
c = ws['C17']; c.value = 'H/M'
c.font = Font(size=10.0)
c = ws['D17']; c.value = 'stockanalysis.com, multiples.vc, GuruFocus Jun-Jul 26'
c.font = Font(size=10.0)
c = ws['A18']; c.value = 'Consensus TP'
c.font = Font(size=10.0)
c = ws['B18']; c.value = '19.3-20.6'
c.font = Font(size=10.0)
c = ws['C18']; c.value = 'M'
c.font = Font(size=10.0)
c = ws['D18']; c.value = 'MarketScreener/TipRanks Jul-26'
c.font = Font(size=10.0)
c = ws['A19']; c.value = 'All scenario growth/margin paths'
c.font = Font(size=10.0)
c = ws['B19']; c.value = '03 Assumptions'
c.font = Font(size=10.0)
c = ws['C19']; c.value = 'L/M'
c.font = Font(size=10.0)
c = ws['D19']; c.value = 'Analyst judgment anchored to history - see Assumptions Register (project 01_Research)'
c.font = Font(size=10.0)
c = ws['A22']; c.value = 'Confidence: H = reported/primary | M = derived/consensus | L = analyst estimate (clearly labelled)'
c.font = Font(size=10.0)
c = ws['A24']; c.value = 'ANNOTATED-SCREENSHOT NOTE: source-page screenshots are not embedded (build environment cannot capture'
c.font = Font(size=10.0)
c = ws['A25']; c.value = 'annotated images from the source PDFs). In lieu: precise table-level citations below. All primary PDFs are linked'
c.font = Font(size=10.0)
c = ws['A26']; c.value = 'in the project Bibliography for page-level verification.'
c.font = Font(size=10.0)
c = ws['A28']; c.value = 'Precise citations:'
c.font = Font(bold=True, color='001F3864')
c = ws['A29']; c.value = 'Revenue by customer type'
c.font = Font(size=10.0)
c = ws['D29']; c.value = "FY25 PR p.3, table 'Revenue by customer type'"
c.font = Font(size=10.0)
c = ws['A30']; c.value = 'Revenue by geography'
c.font = Font(size=10.0)
c = ws['D30']; c.value = "FY25 PR p.3-4, table 'Revenue by geographies'"
c.font = Font(size=10.0)
c = ws['A31']; c.value = 'Revenue by channel'
c.font = Font(size=10.0)
c = ws['D31']; c.value = "FY25 PR p.4, table 'Revenue by distribution channels'"
c.font = Font(size=10.0)
c = ws['A32']; c.value = 'EBITDA / EBIT / NP adj.'
c.font = Font(size=10.0)
c = ws['D32']; c.value = 'FY25 PR p.4-5, section 2'
c.font = Font(size=10.0)
c = ws['A33']; c.value = 'NFP & FCF'
c.font = Font(size=10.0)
c = ws['D33']; c.value = "FY25 PR p.5, 'Net Financial Position and Free Cash Flow'"
c.font = Font(size=10.0)
c = ws['A34']; c.value = 'Full P&L / BS / CF'
c.font = Font(size=10.0)
c = ws['D34']; c.value = 'FY25 PR p.10-12, consolidated statements'
c.font = Font(size=10.0)
c = ws['A35']; c.value = 'Margin history 2021-24'
c.font = Font(size=10.0)
c = ws['D35']; c.value = 'FY25 PR p.2 (17.5/18.3/18.8/19.8%)'
c.font = Font(size=10.0)

ws = wb.create_sheet('03 Assumptions')
ws.column_dimensions['A'].width = 46.0
ws.column_dimensions['B'].width = 10.0
ws.column_dimensions['C'].width = 10.0
ws.column_dimensions['D'].width = 10.0
ws.column_dimensions['E'].width = 10.0
ws.column_dimensions['F'].width = 12.0
ws.column_dimensions['G'].width = 12.0
ws.column_dimensions['H'].width = 12.0
ws.column_dimensions['I'].width = 12.0
ws.column_dimensions['J'].width = 12.0
ws.column_dimensions['K'].width = 12.0
ws.column_dimensions['L'].width = 12.0
ws.column_dimensions['M'].width = 12.0
ws.column_dimensions['N'].width = 12.0
ws.column_dimensions['O'].width = 12.0
ws.column_dimensions['P'].width = 12.0
ws.column_dimensions['Q'].width = 12.0
c = ws['A1']; c.value = 'ASSUMPTIONS & SCENARIO SWITCH'
c.font = Font(bold=True, color='00FFFFFF', size=12.0)
c.fill = PatternFill("solid", fgColor='001F3864')
c = ws['A3']; c.value = 'SCENARIO SWITCH (1=Bear, 2=Base, 3=Bull)'
c.font = Font(bold=True)
c = ws['B3']; c.value = 2
c.number_format = '#,##0;(#,##0);-'
c.font = Font(color='000000FF', size=10.0)
c.fill = PatternFill("solid", fgColor='00FFFF00')
c = ws['C3']; c.value = '=IF(OR($B$3=1,$B$3=2,$B$3=3),"OK","ERROR: switch must be 1, 2 or 3")'
c.font = Font(bold=True, color='00FF0000', size=10.0)
c = ws['A5']; c.value = 'Scenario-driven inputs'
c.font = Font(bold=True)
c = ws['C5']; c.value = 'Bear'
c.font = Font(bold=True)
c = ws['D5']; c.value = 'Base'
c.font = Font(bold=True)
c = ws['E5']; c.value = 'Bull'
c.font = Font(bold=True)
c = ws['F5']; c.value = 'ACTIVE'
c.font = Font(bold=True)
c = ws['A6']; c.value = 'BtoB equipment growth, 2026'
c.font = Font(size=10.0)
c = ws['C6']; c.value = 0.04
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['D6']; c.value = 0.09
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['E6']; c.value = 0.13
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['F6']; c.value = '=CHOOSE($B$3,C6,D6,E6)'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c.fill = PatternFill("solid", fgColor='00FFFF00')
c = ws['A7']; c.value = 'BtoB growth by 2035 (fade to)'
c.font = Font(size=10.0)
c = ws['C7']; c.value = 0.02
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['D7']; c.value = 0.025
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['E7']; c.value = 0.03
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['F7']; c.value = '=CHOOSE($B$3,C7,D7,E7)'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c.fill = PatternFill("solid", fgColor='00FFFF00')
c = ws['A8']; c.value = 'BtoC growth, 2026'
c.font = Font(size=10.0)
c = ws['C8']; c.value = 0
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['D8']; c.value = 0.05
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['E8']; c.value = 0.09
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['F8']; c.value = '=CHOOSE($B$3,C8,D8,E8)'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c.fill = PatternFill("solid", fgColor='00FFFF00')
c = ws['A9']; c.value = 'BtoC growth by 2035 (fade to)'
c.font = Font(size=10.0)
c = ws['C9']; c.value = 0.01
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['D9']; c.value = 0.02
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['E9']; c.value = 0.025
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['F9']; c.value = '=CHOOSE($B$3,C9,D9,E9)'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c.fill = PatternFill("solid", fgColor='00FFFF00')
c = ws['A10']; c.value = 'B2B digital attach endpoint (by 2030)'
c.font = Font(size=10.0)
c = ws['C10']; c.value = 0.2
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['D10']; c.value = 0.3
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['E10']; c.value = 0.4
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['F10']; c.value = '=CHOOSE($B$3,C10,D10,E10)'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c.fill = PatternFill("solid", fgColor='00FFFF00')
c = ws['A11']; c.value = 'Consumer digital attach endpoint (by 2030)'
c.font = Font(size=10.0)
c = ws['C11']; c.value = 0.15
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['D11']; c.value = 0.2
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['E11']; c.value = 0.3
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['F11']; c.value = '=CHOOSE($B$3,C11,D11,E11)'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c.fill = PatternFill("solid", fgColor='00FFFF00')
c = ws['A12']; c.value = 'Digital ARPU growth p.a.'
c.font = Font(size=10.0)
c = ws['C12']; c.value = 0
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['D12']; c.value = 0.03
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['E12']; c.value = 0.05
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['F12']; c.value = '=CHOOSE($B$3,C12,D12,E12)'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c.fill = PatternFill("solid", fgColor='00FFFF00')
c = ws['A13']; c.value = 'Hardware EBITDA margin endpoint by 2032 (v3 isolated; ex-digital)'
c.font = Font(size=10.0)
c = ws['C13']; c.value = 0.195
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['D13']; c.value = 0.2075
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['E13']; c.value = 0.222
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['F13']; c.value = '=CHOOSE($B$3,C13,D13,E13)'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c.fill = PatternFill("solid", fgColor='00FFFF00')
c = ws['A14']; c.value = 'Capex % revenue'
c.font = Font(size=10.0)
c = ws['C14']; c.value = 0.055
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['D14']; c.value = 0.05
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['E14']; c.value = 0.05
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['F14']; c.value = '=CHOOSE($B$3,C14,D14,E14)'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c.fill = PatternFill("solid", fgColor='00FFFF00')
c = ws['A15']; c.value = 'WC stress: DSO shift (+days, bear case)'
c.font = Font(size=10.0)
c = ws['C15']; c.value = 10
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='000000FF', size=10.0)
c = ws['D15']; c.value = 0
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='000000FF', size=10.0)
c = ws['E15']; c.value = 0
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='000000FF', size=10.0)
c = ws['F15']; c.value = '=CHOOSE($B$3,C15,D15,E15)'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c.fill = PatternFill("solid", fgColor='00FFFF00')
c = ws['A16']; c.value = 'Terminal growth g'
c.font = Font(size=10.0)
c = ws['C16']; c.value = 0.015
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['D16']; c.value = 0.0225
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['E16']; c.value = 0.0275
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['F16']; c.value = '=CHOOSE($B$3,C16,D16,E16)'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c.fill = PatternFill("solid", fgColor='00FFFF00')
c = ws['A17']; c.value = 'Exit EV/EBITDA multiple (cross-check)'
c.font = Font(size=10.0)
c = ws['C17']; c.value = 10
c.number_format = '0.0x'
c.font = Font(color='000000FF', size=10.0)
c = ws['D17']; c.value = 12
c.number_format = '0.0x'
c.font = Font(color='000000FF', size=10.0)
c = ws['E17']; c.value = 14
c.number_format = '0.0x'
c.font = Font(color='000000FF', size=10.0)
c = ws['F17']; c.value = '=CHOOSE($B$3,C17,D17,E17)'
c.number_format = '0.0x'
c.font = Font(size=10.0)
c.fill = PatternFill("solid", fgColor='00FFFF00')
c = ws['A18']; c.value = 'Digital EBITDA margin (v3; editable, sensitize 50-75%)'
c = ws['C18']; c.value = 0.5
c = ws['D18']; c.value = 0.6
c = ws['E18']; c.value = 0.75
c = ws['F18']; c.value = '=CHOOSE($B$3,C18,D18,E18)'
c = ws['A20']; c.value = 'Fixed inputs (edit if evidence changes)'
c.font = Font(bold=True)
c = ws['A21']; c.value = 'Effective tax rate'
c.font = Font(size=10.0)
c = ws['B21']; c.value = 0.274
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['A22']; c.value = 'D&A % revenue'
c.font = Font(size=10.0)
c = ws['B22']; c.value = 0.051
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['A23']; c.value = 'Shares outstanding ex-treasury (m)'
c.font = Font(size=10.0)
c = ws['B23']; c.value = 199.3
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='000000FF', size=10.0)
c = ws['A24']; c.value = 'Net financial position incl IFRS16 (€m, + = net cash)'
c.font = Font(size=10.0)
c = ws['B24']; c.value = 156
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='000000FF', size=10.0)
c = ws['A25']; c.value = 'Share price (08-Jul-26, €)'
c.font = Font(size=10.0)
c = ws['B25']; c.value = 15.19
c.number_format = '#,##0.00'
c.font = Font(color='000000FF', size=10.0)
c = ws['A26']; c.value = 'B2B installed sites (2025)'
c.font = Font(size=10.0)
c = ws['B26']; c.value = 100000
c.number_format = '#,##0;(#,##0);-'
c.font = Font(color='000000FF', size=10.0)
c = ws['A27']; c.value = 'B2B sites growth p.a.'
c.font = Font(size=10.0)
c = ws['B27']; c.value = 0.03
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['A28']; c.value = 'B2B digital ARPU €/site/yr (2025)'
c.font = Font(size=10.0)
c = ws['B28']; c.value = 1800
c.number_format = '#,##0;(#,##0);-'
c.font = Font(color='000000FF', size=10.0)
c = ws['A29']; c.value = 'Consumer connected homes (2025)'
c.font = Font(size=10.0)
c = ws['B29']; c.value = 500000
c.number_format = '#,##0;(#,##0);-'
c.font = Font(color='000000FF', size=10.0)
c = ws['A30']; c.value = 'Homes growth p.a.'
c.font = Font(size=10.0)
c = ws['B30']; c.value = 0.05
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['A31']; c.value = 'Consumer ARPU €/user/yr (2025)'
c.font = Font(size=10.0)
c = ws['B31']; c.value = 120
c.number_format = '#,##0;(#,##0);-'
c.font = Font(color='000000FF', size=10.0)
c = ws['A32']; c.value = 'B2B attach rate 2025 (est.)'
c.font = Font(size=10.0)
c = ws['B32']; c.value = 0.2
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['A33']; c.value = 'Consumer attach rate 2025 (est.)'
c.font = Font(size=10.0)
c = ws['B33']; c.value = 0.15
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['A34']; c.value = '2025 BtoB revenue ex-digital (€m)'
c.font = Font(size=10.0)
c = ws['B34']; c.value = 787.6
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='000000FF', size=10.0)
c = ws['A35']; c.value = '2025 BtoC revenue ex-digital (€m)'
c.font = Font(size=10.0)
c = ws['B35']; c.value = 186.7
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='000000FF', size=10.0)
c = ws['A36']; c.value = '2025 digital revenue est. (€m)'
c.font = Font(size=10.0)
c = ws['B36']; c.value = 45
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='000000FF', size=10.0)
c = ws['A38']; c.value = 'Note: digital revenue is NOT disclosed. Rows 26-33 + 36 are an explicit L-confidence estimation'
c.font = Font(size=10.0)
c = ws['A39']; c.value = 'block; 2025 segment bases (rows 34-35) are reported segments minus the digital estimate.'
c.font = Font(size=10.0)
c = ws['A40']; c.value = '=IF(ABS(B34+B35+B36-1019.3)<0.5,"CHECK OK: 2025 splits sum to reported 1,019.3","ERROR: 2025 revenue splits do not sum")'
c.font = Font(bold=True, color='00FF0000', size=10.0)
c = ws['A42']; c.value = 'IFRS 16 & working-capital inputs (v2)'
c.font = Font(bold=True)
c = ws['A43']; c.value = 'Lease principal repayment % revenue (2025: 15.1/1019.3)'
c.font = Font(size=10.0)
c = ws['B43']; c.value = 0.0148
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['A44']; c.value = 'NFP EXCLUDING IFRS16 (€m) — used in DCF bridge'
c.font = Font(size=10.0)
c = ws['B44']; c.value = 209.6
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='000000FF', size=10.0)
c = ws['A45']; c.value = 'DSO days (2025 actual)'
c.font = Font(size=10.0)
c = ws['B45']; c.value = 47.3
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='000000FF', size=10.0)
c = ws['A46']; c.value = 'DIO days on materials (2025 actual)'
c.font = Font(size=10.0)
c = ws['B46']; c.value = 127
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='000000FF', size=10.0)
c = ws['A47']; c.value = 'DPO days on materials (2025 actual)'
c.font = Font(size=10.0)
c = ws['B47']; c.value = 221
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='000000FF', size=10.0)
c = ws['A48']; c.value = 'Materials cost % revenue'
c.font = Font(size=10.0)
c = ws['B48']; c.value = 0.315
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['A49']; c.value = 'Bear-case DPO compression (−days)'
c.font = Font(size=10.0)
c = ws['B49']; c.value = 30
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='000000FF', size=10.0)
c = ws['A51']; c.value = 'v2 methodology: FCFF charges lease principal (IFRS16 consistency) and the equity bridge uses NFP EXCL leases (B44).'
c.font = Font(size=10.0)
c = ws['A52']; c.value = 'NWC built on DSO/DIO/DPO days (07); bear case applies DSO +10d / DPO −30d stress. Bear growth/margin use explicit'
c.font = Font(size=10.0)
c = ws['A53']; c.value = 'cyclical paths on 05 (2027 downturn, partial recovery) instead of a linear fade.'
c.font = Font(size=10.0)
c = ws['A54']; c.value = 'v3: digital P&L isolated. Hardware margin expands from a derived 2025 level (~19.8%, = (220.1 - digital EBITDA)/hardware rev) to F13 endpoint;'
c = ws['A55']; c.value = 'digital revenue carries its own margin (F18). Base hardware endpoint 20.75% holds TOTAL base margin at the defensible ~23.5% (margin-neutral restructure);'
c = ws['A56']; c.value = 'the 22.75% hardware-endpoint variant (total ~25.4%, base ~EUR 17.2) is an UPSIDE sensitivity, not the base. Bear keeps the explicit cyclical total-margin path (05!row33).'

ws = wb.create_sheet('04 Historical Financials')
ws.column_dimensions['A'].width = 46.0
ws.column_dimensions['B'].width = 12.0
ws.column_dimensions['C'].width = 12.0
ws.column_dimensions['D'].width = 12.0
ws.column_dimensions['E'].width = 12.0
ws.column_dimensions['F'].width = 12.0
ws.column_dimensions['G'].width = 12.0
ws.column_dimensions['H'].width = 12.0
ws.column_dimensions['I'].width = 12.0
ws.column_dimensions['J'].width = 12.0
ws.column_dimensions['K'].width = 12.0
ws.column_dimensions['L'].width = 12.0
ws.column_dimensions['M'].width = 12.0
ws.column_dimensions['N'].width = 12.0
ws.column_dimensions['O'].width = 12.0
ws.column_dimensions['P'].width = 12.0
ws.column_dimensions['Q'].width = 12.0
c = ws['A1']; c.value = 'HISTORICAL FINANCIALS & TRENDS'
c.font = Font(bold=True, color='00FFFFFF', size=12.0)
c.fill = PatternFill("solid", fgColor='001F3864')
c = ws['A3']; c.value = '€m'
c.font = Font(bold=True)
c = ws['B3']; c.value = '2022'
c.font = Font(bold=True)
c = ws['C3']; c.value = '2023'
c.font = Font(bold=True)
c = ws['D3']; c.value = '2024'
c.font = Font(bold=True)
c = ws['E3']; c.value = '2025'
c.font = Font(bold=True)
c = ws['A4']; c.value = 'Revenue'
c.font = Font(size=10.0)
c = ws['B4']; c.value = 721.5
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='000000FF', size=10.0)
c = ws['C4']; c.value = 808.1
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='000000FF', size=10.0)
c = ws['D4']; c.value = 901.3
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='000000FF', size=10.0)
c = ws['E4']; c.value = 1019.3
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='000000FF', size=10.0)
c = ws['A5']; c.value = 'Adj. EBITDA'
c.font = Font(size=10.0)
c = ws['B5']; c.value = 132
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='000000FF', size=10.0)
c = ws['C5']; c.value = 152
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='000000FF', size=10.0)
c = ws['D5']; c.value = 178.4
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='000000FF', size=10.0)
c = ws['E5']; c.value = 220.1
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='000000FF', size=10.0)
c = ws['A6']; c.value = 'Adj. EBIT'
c.font = Font(size=10.0)
c = ws['D6']; c.value = 122.1
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='000000FF', size=10.0)
c = ws['E6']; c.value = 163.9
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='000000FF', size=10.0)
c = ws['A7']; c.value = 'Adj. Net profit'
c.font = Font(size=10.0)
c = ws['D7']; c.value = 90.2
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='000000FF', size=10.0)
c = ws['E7']; c.value = 119.9
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='000000FF', size=10.0)
c = ws['A8']; c.value = 'D&A'
c.font = Font(size=10.0)
c = ws['D8']; c.value = 51.8
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='000000FF', size=10.0)
c = ws['E8']; c.value = 52.3
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='000000FF', size=10.0)
c = ws['A9']; c.value = 'Recurring FCF'
c.font = Font(size=10.0)
c = ws['D9']; c.value = 116.4
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='000000FF', size=10.0)
c = ws['E9']; c.value = 129.9
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='000000FF', size=10.0)
c = ws['A10']; c.value = 'Net financial position (+=cash)'
c.font = Font(size=10.0)
c = ws['D10']; c.value = 160.1
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='000000FF', size=10.0)
c = ws['E10']; c.value = 156
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='000000FF', size=10.0)
c = ws['A11']; c.value = 'Inventory'
c.font = Font(size=10.0)
c = ws['D11']; c.value = 110.9
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='000000FF', size=10.0)
c = ws['E11']; c.value = 112
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='000000FF', size=10.0)
c = ws['A12']; c.value = 'Trade receivables'
c.font = Font(size=10.0)
c = ws['D12']; c.value = 132.8
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='000000FF', size=10.0)
c = ws['E12']; c.value = 131.8
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='000000FF', size=10.0)
c = ws['A13']; c.value = 'Trade payables'
c.font = Font(size=10.0)
c = ws['D13']; c.value = 179.1
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='000000FF', size=10.0)
c = ws['E13']; c.value = 194.4
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='000000FF', size=10.0)
c = ws['A15']; c.value = 'Derived ratios'
c.font = Font(bold=True)
c = ws['A16']; c.value = 'EBITDA margin'
c.font = Font(size=10.0)
c = ws['B16']; c.value = '=B5/B4'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['C16']; c.value = '=C5/C4'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['D16']; c.value = '=D5/D4'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['E16']; c.value = '=E5/E4'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['A17']; c.value = 'Revenue growth'
c.font = Font(size=10.0)
c = ws['C17']; c.value = '=C4/B4-1'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['D17']; c.value = '=D4/C4-1'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['E17']; c.value = '=E4/D4-1'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['A18']; c.value = 'Net trade WC (inv+AR-AP)'
c.font = Font(size=10.0)
c = ws['D18']; c.value = '=D11+D12-D13'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['E18']; c.value = '=E11+E12-E13'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['A19']; c.value = 'NWC % revenue'
c.font = Font(size=10.0)
c = ws['D19']; c.value = '=D18/D4'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['E19']; c.value = '=E18/E4'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['A21']; c.value = 'Source: FY2023 & FY2025 results press releases (see 02 Sources). Blue = reported hardcodes.'
c.font = Font(size=10.0)
c = ws['A23']; c.value = 'TREND EXHIBITS'
c.font = Font(bold=True, color='001F3864')
c = ws['A24']; c.value = 'EBITDA margin 2021-2025 (reported)'
c.font = Font(size=10.0)
c = ws['A25']; c.value = 'Year'
c.font = Font(size=10.0)
c = ws['B25']; c.value = '2021'
c.font = Font(size=10.0)
c = ws['C25']; c.value = '2022'
c.font = Font(size=10.0)
c = ws['D25']; c.value = '2023'
c.font = Font(size=10.0)
c = ws['E25']; c.value = '2024'
c.font = Font(size=10.0)
c = ws['F25']; c.value = '2025'
c.font = Font(size=10.0)
c = ws['A26']; c.value = 'Margin'
c.font = Font(size=10.0)
c = ws['B26']; c.value = 0.175
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['C26']; c.value = 0.183
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['D26']; c.value = 0.188
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['E26']; c.value = 0.198
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['F26']; c.value = 0.216
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['A28']; c.value = 'Segment & geography mix (reported, €m)'
c.font = Font(size=10.0)
c = ws['B29']; c.value = '2024'
c.font = Font(size=10.0)
c = ws['C29']; c.value = '2025'
c.font = Font(size=10.0)
c = ws['A30']; c.value = 'BtoB'
c.font = Font(size=10.0)
c = ws['B30']; c.value = 716.9
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='000000FF', size=10.0)
c = ws['C30']; c.value = 823.6
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='000000FF', size=10.0)
c = ws['A31']; c.value = 'BtoC'
c.font = Font(size=10.0)
c = ws['B31']; c.value = 184.4
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='000000FF', size=10.0)
c = ws['C31']; c.value = 195.7
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='000000FF', size=10.0)
c = ws['A32']; c.value = 'Europe ex-Italy'
c.font = Font(size=10.0)
c = ws['B32']; c.value = 416.1
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='000000FF', size=10.0)
c = ws['C32']; c.value = 479.7
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='000000FF', size=10.0)
c = ws['A33']; c.value = 'Americas'
c.font = Font(size=10.0)
c = ws['B33']; c.value = 146.4
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='000000FF', size=10.0)
c = ws['C33']; c.value = 168.2
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='000000FF', size=10.0)
c = ws['A34']; c.value = 'MEIA'
c.font = Font(size=10.0)
c = ws['B34']; c.value = 126.4
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='000000FF', size=10.0)
c = ws['C34']; c.value = 134.7
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='000000FF', size=10.0)
c = ws['A35']; c.value = 'APAC'
c.font = Font(size=10.0)
c = ws['B35']; c.value = 122.2
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='000000FF', size=10.0)
c = ws['C35']; c.value = 128.7
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='000000FF', size=10.0)
c = ws['A36']; c.value = 'Italy'
c.font = Font(size=10.0)
c = ws['B36']; c.value = 90.1
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='000000FF', size=10.0)
c = ws['C36']; c.value = 108
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='000000FF', size=10.0)
c = ws['A38']; c.value = 'Cash conversion (recurring pre-tax FCF / EBITDA): 86% (2024) → 82% (2025). ROCE ≈ 24% (2025, approx).'
c.font = Font(size=10.0)

ws = wb.create_sheet('05 Revenue Build')
ws.column_dimensions['A'].width = 46.0
ws.column_dimensions['B'].width = 12.0
ws.column_dimensions['C'].width = 12.0
ws.column_dimensions['D'].width = 12.0
ws.column_dimensions['E'].width = 12.0
ws.column_dimensions['F'].width = 12.0
ws.column_dimensions['G'].width = 12.0
ws.column_dimensions['H'].width = 12.0
ws.column_dimensions['I'].width = 12.0
ws.column_dimensions['J'].width = 12.0
ws.column_dimensions['K'].width = 12.0
ws.column_dimensions['L'].width = 12.0
ws.column_dimensions['M'].width = 12.0
ws.column_dimensions['N'].width = 12.0
ws.column_dimensions['O'].width = 12.0
ws.column_dimensions['P'].width = 12.0
ws.column_dimensions['Q'].width = 12.0
c = ws['A1']; c.value = 'REVENUE BUILD'
c.font = Font(bold=True, color='00FFFFFF', size=12.0)
c.fill = PatternFill("solid", fgColor='001F3864')
c = ws['A3']; c.value = '€m unless noted'
c.font = Font(bold=True)
c = ws['B3']; c.value = '2022A'
c.font = Font(bold=True)
c = ws['C3']; c.value = '2023A'
c.font = Font(bold=True)
c = ws['D3']; c.value = '2024A'
c.font = Font(bold=True)
c = ws['E3']; c.value = '2025A'
c.font = Font(bold=True)
c = ws['F3']; c.value = '2026E'
c.font = Font(bold=True)
c = ws['G3']; c.value = '2027E'
c.font = Font(bold=True)
c = ws['H3']; c.value = '2028E'
c.font = Font(bold=True)
c = ws['I3']; c.value = '2029E'
c.font = Font(bold=True)
c = ws['J3']; c.value = '2030E'
c.font = Font(bold=True)
c = ws['K3']; c.value = '2031E'
c.font = Font(bold=True)
c = ws['L3']; c.value = '2032E'
c.font = Font(bold=True)
c = ws['M3']; c.value = '2033E'
c.font = Font(bold=True)
c = ws['N3']; c.value = '2034E'
c.font = Font(bold=True)
c = ws['O3']; c.value = '2035E'
c.font = Font(bold=True)
c = ws['A4']; c.value = 'Year index (t)'
c.font = Font(size=10.0)
c = ws['F4']; c.value = 1
c.number_format = '#,##0;(#,##0);-'
c.font = Font(size=10.0)
c = ws['G4']; c.value = 2
c.number_format = '#,##0;(#,##0);-'
c.font = Font(size=10.0)
c = ws['H4']; c.value = 3
c.number_format = '#,##0;(#,##0);-'
c.font = Font(size=10.0)
c = ws['I4']; c.value = 4
c.number_format = '#,##0;(#,##0);-'
c.font = Font(size=10.0)
c = ws['J4']; c.value = 5
c.number_format = '#,##0;(#,##0);-'
c.font = Font(size=10.0)
c = ws['K4']; c.value = 6
c.number_format = '#,##0;(#,##0);-'
c.font = Font(size=10.0)
c = ws['L4']; c.value = 7
c.number_format = '#,##0;(#,##0);-'
c.font = Font(size=10.0)
c = ws['M4']; c.value = 8
c.number_format = '#,##0;(#,##0);-'
c.font = Font(size=10.0)
c = ws['N4']; c.value = 9
c.number_format = '#,##0;(#,##0);-'
c.font = Font(size=10.0)
c = ws['O4']; c.value = 10
c.number_format = '#,##0;(#,##0);-'
c.font = Font(size=10.0)
c = ws['A6']; c.value = 'BtoB equipment & services (ex-digital)'
c.font = Font(bold=True)
c = ws['A7']; c.value = 'Growth (base/bull: linear fade; bear: explicit cyclical path row 32)'
c.font = Font(size=10.0)
c = ws['F7']; c.value = "=IF('03 Assumptions'!$B$3=1,F32,'03 Assumptions'!$F$6+('03 Assumptions'!$F$7-'03 Assumptions'!$F$6)*(F4-1)/9)"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['G7']; c.value = "=IF('03 Assumptions'!$B$3=1,G32,'03 Assumptions'!$F$6+('03 Assumptions'!$F$7-'03 Assumptions'!$F$6)*(G4-1)/9)"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['H7']; c.value = "=IF('03 Assumptions'!$B$3=1,H32,'03 Assumptions'!$F$6+('03 Assumptions'!$F$7-'03 Assumptions'!$F$6)*(H4-1)/9)"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['I7']; c.value = "=IF('03 Assumptions'!$B$3=1,I32,'03 Assumptions'!$F$6+('03 Assumptions'!$F$7-'03 Assumptions'!$F$6)*(I4-1)/9)"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['J7']; c.value = "=IF('03 Assumptions'!$B$3=1,J32,'03 Assumptions'!$F$6+('03 Assumptions'!$F$7-'03 Assumptions'!$F$6)*(J4-1)/9)"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['K7']; c.value = "=IF('03 Assumptions'!$B$3=1,K32,'03 Assumptions'!$F$6+('03 Assumptions'!$F$7-'03 Assumptions'!$F$6)*(K4-1)/9)"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['L7']; c.value = "=IF('03 Assumptions'!$B$3=1,L32,'03 Assumptions'!$F$6+('03 Assumptions'!$F$7-'03 Assumptions'!$F$6)*(L4-1)/9)"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['M7']; c.value = "=IF('03 Assumptions'!$B$3=1,M32,'03 Assumptions'!$F$6+('03 Assumptions'!$F$7-'03 Assumptions'!$F$6)*(M4-1)/9)"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['N7']; c.value = "=IF('03 Assumptions'!$B$3=1,N32,'03 Assumptions'!$F$6+('03 Assumptions'!$F$7-'03 Assumptions'!$F$6)*(N4-1)/9)"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['O7']; c.value = "=IF('03 Assumptions'!$B$3=1,O32,'03 Assumptions'!$F$6+('03 Assumptions'!$F$7-'03 Assumptions'!$F$6)*(O4-1)/9)"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['A8']; c.value = 'Revenue'
c.font = Font(size=10.0)
c = ws['E8']; c.value = "='03 Assumptions'!$B$34"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['F8']; c.value = '=E8*(1+F7)'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['G8']; c.value = '=F8*(1+G7)'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['H8']; c.value = '=G8*(1+H7)'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['I8']; c.value = '=H8*(1+I7)'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['J8']; c.value = '=I8*(1+J7)'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['K8']; c.value = '=J8*(1+K7)'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['L8']; c.value = '=K8*(1+L7)'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['M8']; c.value = '=L8*(1+M7)'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['N8']; c.value = '=M8*(1+N7)'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['O8']; c.value = '=N8*(1+O7)'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['A10']; c.value = 'BtoC (ex-digital)'
c.font = Font(bold=True)
c = ws['A11']; c.value = 'Growth (linear fade)'
c.font = Font(size=10.0)
c = ws['F11']; c.value = "='03 Assumptions'!$F$8+('03 Assumptions'!$F$9-'03 Assumptions'!$F$8)*(F4-1)/9"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['G11']; c.value = "='03 Assumptions'!$F$8+('03 Assumptions'!$F$9-'03 Assumptions'!$F$8)*(G4-1)/9"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['H11']; c.value = "='03 Assumptions'!$F$8+('03 Assumptions'!$F$9-'03 Assumptions'!$F$8)*(H4-1)/9"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['I11']; c.value = "='03 Assumptions'!$F$8+('03 Assumptions'!$F$9-'03 Assumptions'!$F$8)*(I4-1)/9"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['J11']; c.value = "='03 Assumptions'!$F$8+('03 Assumptions'!$F$9-'03 Assumptions'!$F$8)*(J4-1)/9"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['K11']; c.value = "='03 Assumptions'!$F$8+('03 Assumptions'!$F$9-'03 Assumptions'!$F$8)*(K4-1)/9"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['L11']; c.value = "='03 Assumptions'!$F$8+('03 Assumptions'!$F$9-'03 Assumptions'!$F$8)*(L4-1)/9"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['M11']; c.value = "='03 Assumptions'!$F$8+('03 Assumptions'!$F$9-'03 Assumptions'!$F$8)*(M4-1)/9"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['N11']; c.value = "='03 Assumptions'!$F$8+('03 Assumptions'!$F$9-'03 Assumptions'!$F$8)*(N4-1)/9"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['O11']; c.value = "='03 Assumptions'!$F$8+('03 Assumptions'!$F$9-'03 Assumptions'!$F$8)*(O4-1)/9"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['A12']; c.value = 'Revenue'
c.font = Font(size=10.0)
c = ws['E12']; c.value = "='03 Assumptions'!$B$35"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['F12']; c.value = '=E12*(1+F11)'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['G12']; c.value = '=F12*(1+G11)'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['H12']; c.value = '=G12*(1+H11)'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['I12']; c.value = '=H12*(1+I11)'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['J12']; c.value = '=I12*(1+J11)'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['K12']; c.value = '=J12*(1+K11)'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['L12']; c.value = '=K12*(1+L11)'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['M12']; c.value = '=L12*(1+M11)'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['N12']; c.value = '=M12*(1+N11)'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['O12']; c.value = '=N12*(1+O11)'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['A14']; c.value = 'DIGITAL ESTIMATION BLOCK (L confidence — every input editable on 03)'
c.font = Font(bold=True)
c = ws['A15']; c.value = 'B2B installed sites (#)'
c.font = Font(size=10.0)
c = ws['E15']; c.value = "='03 Assumptions'!$B$26"
c.number_format = '#,##0;(#,##0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['F15']; c.value = "=E15*(1+'03 Assumptions'!$B$27)"
c.number_format = '#,##0;(#,##0);-'
c.font = Font(size=10.0)
c = ws['G15']; c.value = "=F15*(1+'03 Assumptions'!$B$27)"
c.number_format = '#,##0;(#,##0);-'
c.font = Font(size=10.0)
c = ws['H15']; c.value = "=G15*(1+'03 Assumptions'!$B$27)"
c.number_format = '#,##0;(#,##0);-'
c.font = Font(size=10.0)
c = ws['I15']; c.value = "=H15*(1+'03 Assumptions'!$B$27)"
c.number_format = '#,##0;(#,##0);-'
c.font = Font(size=10.0)
c = ws['J15']; c.value = "=I15*(1+'03 Assumptions'!$B$27)"
c.number_format = '#,##0;(#,##0);-'
c.font = Font(size=10.0)
c = ws['K15']; c.value = "=J15*(1+'03 Assumptions'!$B$27)"
c.number_format = '#,##0;(#,##0);-'
c.font = Font(size=10.0)
c = ws['L15']; c.value = "=K15*(1+'03 Assumptions'!$B$27)"
c.number_format = '#,##0;(#,##0);-'
c.font = Font(size=10.0)
c = ws['M15']; c.value = "=L15*(1+'03 Assumptions'!$B$27)"
c.number_format = '#,##0;(#,##0);-'
c.font = Font(size=10.0)
c = ws['N15']; c.value = "=M15*(1+'03 Assumptions'!$B$27)"
c.number_format = '#,##0;(#,##0);-'
c.font = Font(size=10.0)
c = ws['O15']; c.value = "=N15*(1+'03 Assumptions'!$B$27)"
c.number_format = '#,##0;(#,##0);-'
c.font = Font(size=10.0)
c = ws['A16']; c.value = 'B2B attach rate (ramp to endpoint by 2030)'
c.font = Font(size=10.0)
c = ws['E16']; c.value = "='03 Assumptions'!$B$32"
c.number_format = '0.0%'
c.font = Font(color='00008000', size=10.0)
c = ws['F16']; c.value = "='03 Assumptions'!$B$32+('03 Assumptions'!$F$10-'03 Assumptions'!$B$32)*MIN(F4,5)/5"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['G16']; c.value = "='03 Assumptions'!$B$32+('03 Assumptions'!$F$10-'03 Assumptions'!$B$32)*MIN(G4,5)/5"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['H16']; c.value = "='03 Assumptions'!$B$32+('03 Assumptions'!$F$10-'03 Assumptions'!$B$32)*MIN(H4,5)/5"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['I16']; c.value = "='03 Assumptions'!$B$32+('03 Assumptions'!$F$10-'03 Assumptions'!$B$32)*MIN(I4,5)/5"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['J16']; c.value = "='03 Assumptions'!$B$32+('03 Assumptions'!$F$10-'03 Assumptions'!$B$32)*MIN(J4,5)/5"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['K16']; c.value = "='03 Assumptions'!$B$32+('03 Assumptions'!$F$10-'03 Assumptions'!$B$32)*MIN(K4,5)/5"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['L16']; c.value = "='03 Assumptions'!$B$32+('03 Assumptions'!$F$10-'03 Assumptions'!$B$32)*MIN(L4,5)/5"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['M16']; c.value = "='03 Assumptions'!$B$32+('03 Assumptions'!$F$10-'03 Assumptions'!$B$32)*MIN(M4,5)/5"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['N16']; c.value = "='03 Assumptions'!$B$32+('03 Assumptions'!$F$10-'03 Assumptions'!$B$32)*MIN(N4,5)/5"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['O16']; c.value = "='03 Assumptions'!$B$32+('03 Assumptions'!$F$10-'03 Assumptions'!$B$32)*MIN(O4,5)/5"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['A17']; c.value = 'B2B ARPU (€/site/yr)'
c.font = Font(size=10.0)
c = ws['E17']; c.value = "='03 Assumptions'!$B$28"
c.number_format = '#,##0;(#,##0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['F17']; c.value = "=E17*(1+'03 Assumptions'!$F$12)"
c.number_format = '#,##0;(#,##0);-'
c.font = Font(size=10.0)
c = ws['G17']; c.value = "=F17*(1+'03 Assumptions'!$F$12)"
c.number_format = '#,##0;(#,##0);-'
c.font = Font(size=10.0)
c = ws['H17']; c.value = "=G17*(1+'03 Assumptions'!$F$12)"
c.number_format = '#,##0;(#,##0);-'
c.font = Font(size=10.0)
c = ws['I17']; c.value = "=H17*(1+'03 Assumptions'!$F$12)"
c.number_format = '#,##0;(#,##0);-'
c.font = Font(size=10.0)
c = ws['J17']; c.value = "=I17*(1+'03 Assumptions'!$F$12)"
c.number_format = '#,##0;(#,##0);-'
c.font = Font(size=10.0)
c = ws['K17']; c.value = "=J17*(1+'03 Assumptions'!$F$12)"
c.number_format = '#,##0;(#,##0);-'
c.font = Font(size=10.0)
c = ws['L17']; c.value = "=K17*(1+'03 Assumptions'!$F$12)"
c.number_format = '#,##0;(#,##0);-'
c.font = Font(size=10.0)
c = ws['M17']; c.value = "=L17*(1+'03 Assumptions'!$F$12)"
c.number_format = '#,##0;(#,##0);-'
c.font = Font(size=10.0)
c = ws['N17']; c.value = "=M17*(1+'03 Assumptions'!$F$12)"
c.number_format = '#,##0;(#,##0);-'
c.font = Font(size=10.0)
c = ws['O17']; c.value = "=N17*(1+'03 Assumptions'!$F$12)"
c.number_format = '#,##0;(#,##0);-'
c.font = Font(size=10.0)
c = ws['A18']; c.value = 'B2B digital revenue (€m)'
c.font = Font(size=10.0)
c = ws['E18']; c.value = '=E15*E16*E17/1000000'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['F18']; c.value = '=F15*F16*F17/1000000'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['G18']; c.value = '=G15*G16*G17/1000000'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['H18']; c.value = '=H15*H16*H17/1000000'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['I18']; c.value = '=I15*I16*I17/1000000'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['J18']; c.value = '=J15*J16*J17/1000000'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['K18']; c.value = '=K15*K16*K17/1000000'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['L18']; c.value = '=L15*L16*L17/1000000'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['M18']; c.value = '=M15*M16*M17/1000000'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['N18']; c.value = '=N15*N16*N17/1000000'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['O18']; c.value = '=O15*O16*O17/1000000'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['A20']; c.value = 'Consumer connected homes (#)'
c.font = Font(size=10.0)
c = ws['E20']; c.value = "='03 Assumptions'!$B$29"
c.number_format = '#,##0;(#,##0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['F20']; c.value = "=E20*(1+'03 Assumptions'!$B$30)"
c.number_format = '#,##0;(#,##0);-'
c.font = Font(size=10.0)
c = ws['G20']; c.value = "=F20*(1+'03 Assumptions'!$B$30)"
c.number_format = '#,##0;(#,##0);-'
c.font = Font(size=10.0)
c = ws['H20']; c.value = "=G20*(1+'03 Assumptions'!$B$30)"
c.number_format = '#,##0;(#,##0);-'
c.font = Font(size=10.0)
c = ws['I20']; c.value = "=H20*(1+'03 Assumptions'!$B$30)"
c.number_format = '#,##0;(#,##0);-'
c.font = Font(size=10.0)
c = ws['J20']; c.value = "=I20*(1+'03 Assumptions'!$B$30)"
c.number_format = '#,##0;(#,##0);-'
c.font = Font(size=10.0)
c = ws['K20']; c.value = "=J20*(1+'03 Assumptions'!$B$30)"
c.number_format = '#,##0;(#,##0);-'
c.font = Font(size=10.0)
c = ws['L20']; c.value = "=K20*(1+'03 Assumptions'!$B$30)"
c.number_format = '#,##0;(#,##0);-'
c.font = Font(size=10.0)
c = ws['M20']; c.value = "=L20*(1+'03 Assumptions'!$B$30)"
c.number_format = '#,##0;(#,##0);-'
c.font = Font(size=10.0)
c = ws['N20']; c.value = "=M20*(1+'03 Assumptions'!$B$30)"
c.number_format = '#,##0;(#,##0);-'
c.font = Font(size=10.0)
c = ws['O20']; c.value = "=N20*(1+'03 Assumptions'!$B$30)"
c.number_format = '#,##0;(#,##0);-'
c.font = Font(size=10.0)
c = ws['A21']; c.value = 'Consumer attach rate'
c.font = Font(size=10.0)
c = ws['E21']; c.value = "='03 Assumptions'!$B$33"
c.number_format = '0.0%'
c.font = Font(color='00008000', size=10.0)
c = ws['F21']; c.value = "='03 Assumptions'!$B$33+('03 Assumptions'!$F$11-'03 Assumptions'!$B$33)*MIN(F4,5)/5"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['G21']; c.value = "='03 Assumptions'!$B$33+('03 Assumptions'!$F$11-'03 Assumptions'!$B$33)*MIN(G4,5)/5"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['H21']; c.value = "='03 Assumptions'!$B$33+('03 Assumptions'!$F$11-'03 Assumptions'!$B$33)*MIN(H4,5)/5"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['I21']; c.value = "='03 Assumptions'!$B$33+('03 Assumptions'!$F$11-'03 Assumptions'!$B$33)*MIN(I4,5)/5"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['J21']; c.value = "='03 Assumptions'!$B$33+('03 Assumptions'!$F$11-'03 Assumptions'!$B$33)*MIN(J4,5)/5"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['K21']; c.value = "='03 Assumptions'!$B$33+('03 Assumptions'!$F$11-'03 Assumptions'!$B$33)*MIN(K4,5)/5"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['L21']; c.value = "='03 Assumptions'!$B$33+('03 Assumptions'!$F$11-'03 Assumptions'!$B$33)*MIN(L4,5)/5"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['M21']; c.value = "='03 Assumptions'!$B$33+('03 Assumptions'!$F$11-'03 Assumptions'!$B$33)*MIN(M4,5)/5"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['N21']; c.value = "='03 Assumptions'!$B$33+('03 Assumptions'!$F$11-'03 Assumptions'!$B$33)*MIN(N4,5)/5"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['O21']; c.value = "='03 Assumptions'!$B$33+('03 Assumptions'!$F$11-'03 Assumptions'!$B$33)*MIN(O4,5)/5"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['A22']; c.value = 'Consumer ARPU (€/user/yr)'
c.font = Font(size=10.0)
c = ws['E22']; c.value = "='03 Assumptions'!$B$31"
c.number_format = '#,##0;(#,##0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['F22']; c.value = "=E22*(1+'03 Assumptions'!$F$12)"
c.number_format = '#,##0;(#,##0);-'
c.font = Font(size=10.0)
c = ws['G22']; c.value = "=F22*(1+'03 Assumptions'!$F$12)"
c.number_format = '#,##0;(#,##0);-'
c.font = Font(size=10.0)
c = ws['H22']; c.value = "=G22*(1+'03 Assumptions'!$F$12)"
c.number_format = '#,##0;(#,##0);-'
c.font = Font(size=10.0)
c = ws['I22']; c.value = "=H22*(1+'03 Assumptions'!$F$12)"
c.number_format = '#,##0;(#,##0);-'
c.font = Font(size=10.0)
c = ws['J22']; c.value = "=I22*(1+'03 Assumptions'!$F$12)"
c.number_format = '#,##0;(#,##0);-'
c.font = Font(size=10.0)
c = ws['K22']; c.value = "=J22*(1+'03 Assumptions'!$F$12)"
c.number_format = '#,##0;(#,##0);-'
c.font = Font(size=10.0)
c = ws['L22']; c.value = "=K22*(1+'03 Assumptions'!$F$12)"
c.number_format = '#,##0;(#,##0);-'
c.font = Font(size=10.0)
c = ws['M22']; c.value = "=L22*(1+'03 Assumptions'!$F$12)"
c.number_format = '#,##0;(#,##0);-'
c.font = Font(size=10.0)
c = ws['N22']; c.value = "=M22*(1+'03 Assumptions'!$F$12)"
c.number_format = '#,##0;(#,##0);-'
c.font = Font(size=10.0)
c = ws['O22']; c.value = "=N22*(1+'03 Assumptions'!$F$12)"
c.number_format = '#,##0;(#,##0);-'
c.font = Font(size=10.0)
c = ws['A23']; c.value = 'Consumer digital revenue (€m)'
c.font = Font(size=10.0)
c = ws['E23']; c.value = '=E20*E21*E22/1000000'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['F23']; c.value = '=F20*F21*F22/1000000'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['G23']; c.value = '=G20*G21*G22/1000000'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['H23']; c.value = '=H20*H21*H22/1000000'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['I23']; c.value = '=I20*I21*I22/1000000'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['J23']; c.value = '=J20*J21*J22/1000000'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['K23']; c.value = '=K20*K21*K22/1000000'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['L23']; c.value = '=L20*L21*L22/1000000'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['M23']; c.value = '=M20*M21*M22/1000000'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['N23']; c.value = '=N20*N21*N22/1000000'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['O23']; c.value = '=O20*O21*O22/1000000'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['A25']; c.value = 'Total digital revenue'
c.font = Font(bold=True)
c = ws['E25']; c.value = '=E18+E23'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['F25']; c.value = '=F18+F23'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['G25']; c.value = '=G18+G23'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['H25']; c.value = '=H18+H23'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['I25']; c.value = '=I18+I23'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['J25']; c.value = '=J18+J23'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['K25']; c.value = '=K18+K23'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['L25']; c.value = '=L18+L23'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['M25']; c.value = '=M18+M23'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['N25']; c.value = '=N18+N23'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['O25']; c.value = '=O18+O23'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['A27']; c.value = 'TOTAL REVENUE'
c.font = Font(bold=True)
c = ws['E27']; c.value = '=E8+E12+E25'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['F27']; c.value = '=F8+F12+F25'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['G27']; c.value = '=G8+G12+G25'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['H27']; c.value = '=H8+H12+H25'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['I27']; c.value = '=I8+I12+I25'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['J27']; c.value = '=J8+J12+J25'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['K27']; c.value = '=K8+K12+K25'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['L27']; c.value = '=L8+L12+L25'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['M27']; c.value = '=M8+M12+M25'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['N27']; c.value = '=N8+N12+N25'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['O27']; c.value = '=O8+O12+O25'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['A28']; c.value = 'Total growth'
c.font = Font(size=10.0)
c = ws['F28']; c.value = '=F27/E27-1'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['G28']; c.value = '=G27/F27-1'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['H28']; c.value = '=H27/G27-1'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['I28']; c.value = '=I27/H27-1'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['J28']; c.value = '=J27/I27-1'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['K28']; c.value = '=K27/J27-1'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['L28']; c.value = '=L27/K27-1'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['M28']; c.value = '=M27/L27-1'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['N28']; c.value = '=N27/M27-1'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['O28']; c.value = '=O27/N27-1'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['A29']; c.value = 'Digital % of revenue'
c.font = Font(size=10.0)
c = ws['E29']; c.value = '=E25/E27'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['F29']; c.value = '=F25/F27'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['G29']; c.value = '=G25/G27'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['H29']; c.value = '=H25/H27'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['I29']; c.value = '=I25/I27'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['J29']; c.value = '=J25/J27'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['K29']; c.value = '=K25/K27'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['L29']; c.value = '=L25/L27'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['M29']; c.value = '=M25/M27'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['N29']; c.value = '=N25/N27'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['O29']; c.value = '=O25/O27'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['A31']; c.value = '=IF(ABS(E27-1019.3)<1,"CHECK OK: 2025 build ties to reported revenue","ERROR: 2025 revenue build broken")'
c.font = Font(bold=True, color='00FF0000', size=10.0)
c = ws['A32']; c.value = 'BEAR BtoB growth path (explicit cyclical: 2027 downturn −10%, recovery from 2029)'
c.font = Font(size=10.0)
c = ws['F32']; c.value = 0.04
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['G32']; c.value = -0.1
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['H32']; c.value = -0.03
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['I32']; c.value = 0.06
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['J32']; c.value = 0.05
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['K32']; c.value = 0.045
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['L32']; c.value = 0.04
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['M32']; c.value = 0.035
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['N32']; c.value = 0.03
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['O32']; c.value = 0.025
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['A33']; c.value = 'BEAR EBITDA margin path (2027 trough 19.0%, recovers to 21.5%)'
c.font = Font(size=10.0)
c = ws['F33']; c.value = 0.205
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['G33']; c.value = 0.19
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['H33']; c.value = 0.198
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['I33']; c.value = 0.205
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['J33']; c.value = 0.208
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['K33']; c.value = 0.21
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['L33']; c.value = 0.212
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['M33']; c.value = 0.213
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['N33']; c.value = 0.214
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['O33']; c.value = 0.215
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)

ws = wb.create_sheet('06 Operating Model')
ws.column_dimensions['A'].width = 46.0
ws.column_dimensions['B'].width = 12.0
ws.column_dimensions['C'].width = 12.0
ws.column_dimensions['D'].width = 12.0
ws.column_dimensions['E'].width = 12.0
ws.column_dimensions['F'].width = 12.0
ws.column_dimensions['G'].width = 12.0
ws.column_dimensions['H'].width = 12.0
ws.column_dimensions['I'].width = 12.0
ws.column_dimensions['J'].width = 12.0
ws.column_dimensions['K'].width = 12.0
ws.column_dimensions['L'].width = 12.0
ws.column_dimensions['M'].width = 12.0
ws.column_dimensions['N'].width = 12.0
ws.column_dimensions['O'].width = 12.0
ws.column_dimensions['P'].width = 12.0
ws.column_dimensions['Q'].width = 12.0
c = ws['A1']; c.value = 'OPERATING MODEL'
c.font = Font(bold=True, color='00FFFFFF', size=12.0)
c.fill = PatternFill("solid", fgColor='001F3864')
c = ws['A3']; c.value = '€m unless noted'
c.font = Font(bold=True)
c = ws['F3']; c.value = '2026E'
c.font = Font(bold=True)
c = ws['G3']; c.value = '2027E'
c.font = Font(bold=True)
c = ws['H3']; c.value = '2028E'
c.font = Font(bold=True)
c = ws['I3']; c.value = '2029E'
c.font = Font(bold=True)
c = ws['J3']; c.value = '2030E'
c.font = Font(bold=True)
c = ws['K3']; c.value = '2031E'
c.font = Font(bold=True)
c = ws['L3']; c.value = '2032E'
c.font = Font(bold=True)
c = ws['M3']; c.value = '2033E'
c.font = Font(bold=True)
c = ws['N3']; c.value = '2034E'
c.font = Font(bold=True)
c = ws['O3']; c.value = '2035E'
c.font = Font(bold=True)
c = ws['A4']; c.value = 'Revenue'
c.font = Font(size=10.0)
c = ws['F4']; c.value = "='05 Revenue Build'!F27"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['G4']; c.value = "='05 Revenue Build'!G27"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['H4']; c.value = "='05 Revenue Build'!H27"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['I4']; c.value = "='05 Revenue Build'!I27"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['J4']; c.value = "='05 Revenue Build'!J27"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['K4']; c.value = "='05 Revenue Build'!K27"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['L4']; c.value = "='05 Revenue Build'!L27"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['M4']; c.value = "='05 Revenue Build'!M27"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['N4']; c.value = "='05 Revenue Build'!N27"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['O4']; c.value = "='05 Revenue Build'!O27"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['A5']; c.value = 'Total EBITDA margin (OUTPUT = hardware + digital blocks, rows 20-26)'
c.font = Font(size=10.0)
c = ws['F5']; c.value = '=F6/F4'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['G5']; c.value = '=G6/G4'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['H5']; c.value = '=H6/H4'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['I5']; c.value = '=I6/I4'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['J5']; c.value = '=J6/J4'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['K5']; c.value = '=K6/K4'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['L5']; c.value = '=L6/L4'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['M5']; c.value = '=M6/M4'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['N5']; c.value = '=N6/N4'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['O5']; c.value = '=O6/O4'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['A6']; c.value = 'EBITDA'
c.font = Font(bold=True)
c = ws['F6']; c.value = "=IF('03 Assumptions'!$B$3=1,F4*'05 Revenue Build'!F33,F26)"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['G6']; c.value = "=IF('03 Assumptions'!$B$3=1,G4*'05 Revenue Build'!G33,G26)"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['H6']; c.value = "=IF('03 Assumptions'!$B$3=1,H4*'05 Revenue Build'!H33,H26)"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['I6']; c.value = "=IF('03 Assumptions'!$B$3=1,I4*'05 Revenue Build'!I33,I26)"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['J6']; c.value = "=IF('03 Assumptions'!$B$3=1,J4*'05 Revenue Build'!J33,J26)"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['K6']; c.value = "=IF('03 Assumptions'!$B$3=1,K4*'05 Revenue Build'!K33,K26)"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['L6']; c.value = "=IF('03 Assumptions'!$B$3=1,L4*'05 Revenue Build'!L33,L26)"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['M6']; c.value = "=IF('03 Assumptions'!$B$3=1,M4*'05 Revenue Build'!M33,M26)"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['N6']; c.value = "=IF('03 Assumptions'!$B$3=1,N4*'05 Revenue Build'!N33,N26)"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['O6']; c.value = "=IF('03 Assumptions'!$B$3=1,O4*'05 Revenue Build'!O33,O26)"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['A7']; c.value = 'D&A (% revenue)'
c.font = Font(size=10.0)
c = ws['F7']; c.value = "=-F4*'03 Assumptions'!$B$22"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['G7']; c.value = "=-G4*'03 Assumptions'!$B$22"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['H7']; c.value = "=-H4*'03 Assumptions'!$B$22"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['I7']; c.value = "=-I4*'03 Assumptions'!$B$22"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['J7']; c.value = "=-J4*'03 Assumptions'!$B$22"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['K7']; c.value = "=-K4*'03 Assumptions'!$B$22"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['L7']; c.value = "=-L4*'03 Assumptions'!$B$22"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['M7']; c.value = "=-M4*'03 Assumptions'!$B$22"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['N7']; c.value = "=-N4*'03 Assumptions'!$B$22"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['O7']; c.value = "=-O4*'03 Assumptions'!$B$22"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['A8']; c.value = 'EBIT'
c.font = Font(bold=True)
c = ws['F8']; c.value = '=F6+F7'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['G8']; c.value = '=G6+G7'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['H8']; c.value = '=H6+H7'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['I8']; c.value = '=I6+I7'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['J8']; c.value = '=J6+J7'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['K8']; c.value = '=K6+K7'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['L8']; c.value = '=L6+L7'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['M8']; c.value = '=M6+M7'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['N8']; c.value = '=N6+N7'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['O8']; c.value = '=O6+O7'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['A9']; c.value = 'Tax on EBIT'
c.font = Font(size=10.0)
c = ws['F9']; c.value = "=-F8*'03 Assumptions'!$B$21"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['G9']; c.value = "=-G8*'03 Assumptions'!$B$21"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['H9']; c.value = "=-H8*'03 Assumptions'!$B$21"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['I9']; c.value = "=-I8*'03 Assumptions'!$B$21"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['J9']; c.value = "=-J8*'03 Assumptions'!$B$21"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['K9']; c.value = "=-K8*'03 Assumptions'!$B$21"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['L9']; c.value = "=-L8*'03 Assumptions'!$B$21"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['M9']; c.value = "=-M8*'03 Assumptions'!$B$21"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['N9']; c.value = "=-N8*'03 Assumptions'!$B$21"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['O9']; c.value = "=-O8*'03 Assumptions'!$B$21"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['A10']; c.value = 'NOPAT'
c.font = Font(bold=True)
c = ws['F10']; c.value = '=F8+F9'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['G10']; c.value = '=G8+G9'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['H10']; c.value = '=H8+H9'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['I10']; c.value = '=I8+I9'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['J10']; c.value = '=J8+J9'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['K10']; c.value = '=K8+K9'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['L10']; c.value = '=L8+L9'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['M10']; c.value = '=M8+M9'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['N10']; c.value = '=N8+N9'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['O10']; c.value = '=O8+O9'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['A12']; c.value = 'EBIT margin'
c.font = Font(size=10.0)
c = ws['F12']; c.value = '=F8/F4'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['G12']; c.value = '=G8/G4'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['H12']; c.value = '=H8/H4'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['I12']; c.value = '=I8/I4'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['J12']; c.value = '=J8/J4'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['K12']; c.value = '=K8/K4'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['L12']; c.value = '=L8/L4'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['M12']; c.value = '=M8/M4'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['N12']; c.value = '=N8/N4'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['O12']; c.value = '=O8/O4'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['A13']; c.value = '=IF(AND(F5>0.1,O5<0.35),"CHECK OK: margins in sane range","CHECK: margin path out of range")'
c.font = Font(bold=True, color='00FF0000', size=10.0)
c = ws['A15']; c.value = 'Note: SBC immaterial (<0.35% dilution) — not modeled. Non-recurring items excluded (adjusted basis).'
c.font = Font(size=10.0)
c = ws['A17']; c.value = 'Margin note: 17.5% (2021) -> 21.6% (2025) benefited from post-COVID pent-up demand, supply-chain normalisation and one-time procurement gains. Sustaining a 23.5% TOTAL by 2032 needs the digital/services mix shift + operating leverage, not more procurement.'
c = ws['A19']; c.value = 'DIGITAL-ISOLATED EBITDA BUILD (v3) — base/bull isolate; bear uses explicit cyclical total-margin path (05!row33)'
c = ws['A20']; c.value = 'Hardware revenue (ex-digital)'
c = ws['F20']; c.value = "='05 Revenue Build'!F8+'05 Revenue Build'!F12"
c = ws['G20']; c.value = "='05 Revenue Build'!G8+'05 Revenue Build'!G12"
c = ws['H20']; c.value = "='05 Revenue Build'!H8+'05 Revenue Build'!H12"
c = ws['I20']; c.value = "='05 Revenue Build'!I8+'05 Revenue Build'!I12"
c = ws['J20']; c.value = "='05 Revenue Build'!J8+'05 Revenue Build'!J12"
c = ws['K20']; c.value = "='05 Revenue Build'!K8+'05 Revenue Build'!K12"
c = ws['L20']; c.value = "='05 Revenue Build'!L8+'05 Revenue Build'!L12"
c = ws['M20']; c.value = "='05 Revenue Build'!M8+'05 Revenue Build'!M12"
c = ws['N20']; c.value = "='05 Revenue Build'!N8+'05 Revenue Build'!N12"
c = ws['O20']; c.value = "='05 Revenue Build'!O8+'05 Revenue Build'!O12"
c = ws['A21']; c.value = 'Digital revenue'
c = ws['F21']; c.value = "='05 Revenue Build'!F25"
c = ws['G21']; c.value = "='05 Revenue Build'!G25"
c = ws['H21']; c.value = "='05 Revenue Build'!H25"
c = ws['I21']; c.value = "='05 Revenue Build'!I25"
c = ws['J21']; c.value = "='05 Revenue Build'!J25"
c = ws['K21']; c.value = "='05 Revenue Build'!K25"
c = ws['L21']; c.value = "='05 Revenue Build'!L25"
c = ws['M21']; c.value = "='05 Revenue Build'!M25"
c = ws['N21']; c.value = "='05 Revenue Build'!N25"
c = ws['O21']; c.value = "='05 Revenue Build'!O25"
c = ws['A22']; c.value = 'Hardware 2025 EBITDA margin (derived to tie reported EUR 220.1m)'
c = ws['B22']; c.value = "=(220.1-'03 Assumptions'!B36*'03 Assumptions'!F18)/('03 Assumptions'!B34+'03 Assumptions'!B35)"
c = ws['A23']; c.value = 'Hardware EBITDA margin (path B22 -> 03!F13 endpoint by 2032)'
c = ws['F23']; c.value = "=$B$22+('03 Assumptions'!$F$13-$B$22)*MIN('05 Revenue Build'!F$4,7)/7"
c = ws['G23']; c.value = "=$B$22+('03 Assumptions'!$F$13-$B$22)*MIN('05 Revenue Build'!G$4,7)/7"
c = ws['H23']; c.value = "=$B$22+('03 Assumptions'!$F$13-$B$22)*MIN('05 Revenue Build'!H$4,7)/7"
c = ws['I23']; c.value = "=$B$22+('03 Assumptions'!$F$13-$B$22)*MIN('05 Revenue Build'!I$4,7)/7"
c = ws['J23']; c.value = "=$B$22+('03 Assumptions'!$F$13-$B$22)*MIN('05 Revenue Build'!J$4,7)/7"
c = ws['K23']; c.value = "=$B$22+('03 Assumptions'!$F$13-$B$22)*MIN('05 Revenue Build'!K$4,7)/7"
c = ws['L23']; c.value = "=$B$22+('03 Assumptions'!$F$13-$B$22)*MIN('05 Revenue Build'!L$4,7)/7"
c = ws['M23']; c.value = "=$B$22+('03 Assumptions'!$F$13-$B$22)*MIN('05 Revenue Build'!M$4,7)/7"
c = ws['N23']; c.value = "=$B$22+('03 Assumptions'!$F$13-$B$22)*MIN('05 Revenue Build'!N$4,7)/7"
c = ws['O23']; c.value = "=$B$22+('03 Assumptions'!$F$13-$B$22)*MIN('05 Revenue Build'!O$4,7)/7"
c = ws['A24']; c.value = 'Hardware EBITDA'
c = ws['F24']; c.value = '=F20*F23'
c = ws['G24']; c.value = '=G20*G23'
c = ws['H24']; c.value = '=H20*H23'
c = ws['I24']; c.value = '=I20*I23'
c = ws['J24']; c.value = '=J20*J23'
c = ws['K24']; c.value = '=K20*K23'
c = ws['L24']; c.value = '=L20*L23'
c = ws['M24']; c.value = '=M20*M23'
c = ws['N24']; c.value = '=N20*N23'
c = ws['O24']; c.value = '=O20*O23'
c = ws['A25']; c.value = 'Digital EBITDA (margin = 03!F18)'
c = ws['F25']; c.value = "=F21*'03 Assumptions'!$F$18"
c = ws['G25']; c.value = "=G21*'03 Assumptions'!$F$18"
c = ws['H25']; c.value = "=H21*'03 Assumptions'!$F$18"
c = ws['I25']; c.value = "=I21*'03 Assumptions'!$F$18"
c = ws['J25']; c.value = "=J21*'03 Assumptions'!$F$18"
c = ws['K25']; c.value = "=K21*'03 Assumptions'!$F$18"
c = ws['L25']; c.value = "=L21*'03 Assumptions'!$F$18"
c = ws['M25']; c.value = "=M21*'03 Assumptions'!$F$18"
c = ws['N25']; c.value = "=N21*'03 Assumptions'!$F$18"
c = ws['O25']; c.value = "=O21*'03 Assumptions'!$F$18"
c = ws['A26']; c.value = 'TOTAL EBITDA (hardware + digital) [feeds row 6 in base/bull]'
c = ws['F26']; c.value = '=F24+F25'
c = ws['G26']; c.value = '=G24+G25'
c = ws['H26']; c.value = '=H24+H25'
c = ws['I26']; c.value = '=I24+I25'
c = ws['J26']; c.value = '=J24+J25'
c = ws['K26']; c.value = '=K24+K25'
c = ws['L26']; c.value = '=L24+L25'
c = ws['M26']; c.value = '=M24+M25'
c = ws['N26']; c.value = '=N24+N25'
c = ws['O26']; c.value = '=O24+O25'

ws = wb.create_sheet('07 Working Capital')
ws.column_dimensions['A'].width = 46.0
ws.column_dimensions['B'].width = 12.0
ws.column_dimensions['C'].width = 12.0
ws.column_dimensions['D'].width = 12.0
ws.column_dimensions['E'].width = 12.0
ws.column_dimensions['F'].width = 12.0
ws.column_dimensions['G'].width = 12.0
ws.column_dimensions['H'].width = 12.0
ws.column_dimensions['I'].width = 12.0
ws.column_dimensions['J'].width = 12.0
ws.column_dimensions['K'].width = 12.0
ws.column_dimensions['L'].width = 12.0
ws.column_dimensions['M'].width = 12.0
ws.column_dimensions['N'].width = 12.0
ws.column_dimensions['O'].width = 12.0
ws.column_dimensions['P'].width = 12.0
ws.column_dimensions['Q'].width = 12.0
c = ws['A1']; c.value = 'WORKING CAPITAL'
c.font = Font(bold=True, color='00FFFFFF', size=12.0)
c.fill = PatternFill("solid", fgColor='001F3864')
c = ws['A3']; c.value = '€m unless noted'
c.font = Font(bold=True)
c = ws['F3']; c.value = '2026E'
c.font = Font(bold=True)
c = ws['G3']; c.value = '2027E'
c.font = Font(bold=True)
c = ws['H3']; c.value = '2028E'
c.font = Font(bold=True)
c = ws['I3']; c.value = '2029E'
c.font = Font(bold=True)
c = ws['J3']; c.value = '2030E'
c.font = Font(bold=True)
c = ws['K3']; c.value = '2031E'
c.font = Font(bold=True)
c = ws['L3']; c.value = '2032E'
c.font = Font(bold=True)
c = ws['M3']; c.value = '2033E'
c.font = Font(bold=True)
c = ws['N3']; c.value = '2034E'
c.font = Font(bold=True)
c = ws['O3']; c.value = '2035E'
c.font = Font(bold=True)
c = ws['A4']; c.value = 'Revenue'
c.font = Font(size=10.0)
c = ws['F4']; c.value = "='05 Revenue Build'!F27"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['G4']; c.value = "='05 Revenue Build'!G27"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['H4']; c.value = "='05 Revenue Build'!H27"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['I4']; c.value = "='05 Revenue Build'!I27"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['J4']; c.value = "='05 Revenue Build'!J27"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['K4']; c.value = "='05 Revenue Build'!K27"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['L4']; c.value = "='05 Revenue Build'!L27"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['M4']; c.value = "='05 Revenue Build'!M27"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['N4']; c.value = "='05 Revenue Build'!N27"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['O4']; c.value = "='05 Revenue Build'!O27"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['A5']; c.value = 'DSO (days, incl. scenario stress)'
c.font = Font(size=10.0)
c = ws['F5']; c.value = "='03 Assumptions'!$B$45+'03 Assumptions'!$F$15"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['G5']; c.value = "='03 Assumptions'!$B$45+'03 Assumptions'!$F$15"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['H5']; c.value = "='03 Assumptions'!$B$45+'03 Assumptions'!$F$15"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['I5']; c.value = "='03 Assumptions'!$B$45+'03 Assumptions'!$F$15"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['J5']; c.value = "='03 Assumptions'!$B$45+'03 Assumptions'!$F$15"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['K5']; c.value = "='03 Assumptions'!$B$45+'03 Assumptions'!$F$15"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['L5']; c.value = "='03 Assumptions'!$B$45+'03 Assumptions'!$F$15"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['M5']; c.value = "='03 Assumptions'!$B$45+'03 Assumptions'!$F$15"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['N5']; c.value = "='03 Assumptions'!$B$45+'03 Assumptions'!$F$15"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['O5']; c.value = "='03 Assumptions'!$B$45+'03 Assumptions'!$F$15"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['A6']; c.value = 'DIO (days on materials)'
c.font = Font(size=10.0)
c = ws['F6']; c.value = "='03 Assumptions'!$B$46"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['G6']; c.value = "='03 Assumptions'!$B$46"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['H6']; c.value = "='03 Assumptions'!$B$46"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['I6']; c.value = "='03 Assumptions'!$B$46"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['J6']; c.value = "='03 Assumptions'!$B$46"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['K6']; c.value = "='03 Assumptions'!$B$46"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['L6']; c.value = "='03 Assumptions'!$B$46"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['M6']; c.value = "='03 Assumptions'!$B$46"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['N6']; c.value = "='03 Assumptions'!$B$46"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['O6']; c.value = "='03 Assumptions'!$B$46"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['A7']; c.value = 'DPO (days on materials; bear compresses)'
c.font = Font(size=10.0)
c = ws['F7']; c.value = "='03 Assumptions'!$B$47-IF('03 Assumptions'!$B$3=1,'03 Assumptions'!$B$49,0)"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['G7']; c.value = "='03 Assumptions'!$B$47-IF('03 Assumptions'!$B$3=1,'03 Assumptions'!$B$49,0)"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['H7']; c.value = "='03 Assumptions'!$B$47-IF('03 Assumptions'!$B$3=1,'03 Assumptions'!$B$49,0)"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['I7']; c.value = "='03 Assumptions'!$B$47-IF('03 Assumptions'!$B$3=1,'03 Assumptions'!$B$49,0)"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['J7']; c.value = "='03 Assumptions'!$B$47-IF('03 Assumptions'!$B$3=1,'03 Assumptions'!$B$49,0)"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['K7']; c.value = "='03 Assumptions'!$B$47-IF('03 Assumptions'!$B$3=1,'03 Assumptions'!$B$49,0)"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['L7']; c.value = "='03 Assumptions'!$B$47-IF('03 Assumptions'!$B$3=1,'03 Assumptions'!$B$49,0)"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['M7']; c.value = "='03 Assumptions'!$B$47-IF('03 Assumptions'!$B$3=1,'03 Assumptions'!$B$49,0)"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['N7']; c.value = "='03 Assumptions'!$B$47-IF('03 Assumptions'!$B$3=1,'03 Assumptions'!$B$49,0)"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['O7']; c.value = "='03 Assumptions'!$B$47-IF('03 Assumptions'!$B$3=1,'03 Assumptions'!$B$49,0)"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['A8']; c.value = 'Net working capital = Rev x DSO/365 + Mat x (DIO-DPO)/365'
c.font = Font(size=10.0)
c = ws['F8']; c.value = "=F4*(F5/365+'03 Assumptions'!$B$48*(F6-F7)/365)"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['G8']; c.value = "=G4*(G5/365+'03 Assumptions'!$B$48*(G6-G7)/365)"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['H8']; c.value = "=H4*(H5/365+'03 Assumptions'!$B$48*(H6-H7)/365)"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['I8']; c.value = "=I4*(I5/365+'03 Assumptions'!$B$48*(I6-I7)/365)"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['J8']; c.value = "=J4*(J5/365+'03 Assumptions'!$B$48*(J6-J7)/365)"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['K8']; c.value = "=K4*(K5/365+'03 Assumptions'!$B$48*(K6-K7)/365)"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['L8']; c.value = "=L4*(L5/365+'03 Assumptions'!$B$48*(L6-L7)/365)"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['M8']; c.value = "=M4*(M5/365+'03 Assumptions'!$B$48*(M6-M7)/365)"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['N8']; c.value = "=N4*(N5/365+'03 Assumptions'!$B$48*(N6-N7)/365)"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['O8']; c.value = "=O4*(O5/365+'03 Assumptions'!$B$48*(O6-O7)/365)"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['A9']; c.value = 'Change in NWC (- = cash out)'
c.font = Font(bold=True)
c = ws['F9']; c.value = "=-(F8-(1019.3*(F5/365+'03 Assumptions'!$B$48*(F6-'03 Assumptions'!$B$47)/365)))"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['G9']; c.value = '=-(G8-F8)'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['H9']; c.value = '=-(H8-G8)'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['I9']; c.value = '=-(I8-H8)'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['J9']; c.value = '=-(J8-I8)'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['K9']; c.value = '=-(K8-J8)'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['L9']; c.value = '=-(L8-K8)'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['M9']; c.value = '=-(M8-L8)'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['N9']; c.value = '=-(N8-M8)'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['O9']; c.value = '=-(O8-N8)'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['A12']; c.value = 'Effective NWC % of revenue (helper for grids)'
c.font = Font(size=10.0)
c = ws['B12']; c.value = '=F8/F4'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['A13']; c.value = '2025 actual days: DSO 47.3 / DIO 127 / DPO 221 (ties to ~49m NWC = 4.9% rev). Bear stress: DSO +10d, DPO -30d'
c.font = Font(size=10.0)
c = ws['A14']; c.value = '(gym customers slow-pay; supplier terms tighten in downturns). 2026 change computed vs unstressed-DPO 2025 base.'
c.font = Font(size=10.0)

ws = wb.create_sheet('08 Capex & Depreciation')
ws.column_dimensions['A'].width = 46.0
ws.column_dimensions['B'].width = 12.0
ws.column_dimensions['C'].width = 12.0
ws.column_dimensions['D'].width = 12.0
ws.column_dimensions['E'].width = 12.0
ws.column_dimensions['F'].width = 12.0
ws.column_dimensions['G'].width = 12.0
ws.column_dimensions['H'].width = 12.0
ws.column_dimensions['I'].width = 12.0
ws.column_dimensions['J'].width = 12.0
ws.column_dimensions['K'].width = 12.0
ws.column_dimensions['L'].width = 12.0
ws.column_dimensions['M'].width = 12.0
ws.column_dimensions['N'].width = 12.0
ws.column_dimensions['O'].width = 12.0
ws.column_dimensions['P'].width = 12.0
ws.column_dimensions['Q'].width = 12.0
c = ws['A1']; c.value = 'CAPEX & DEPRECIATION'
c.font = Font(bold=True, color='00FFFFFF', size=12.0)
c.fill = PatternFill("solid", fgColor='001F3864')
c = ws['A3']; c.value = '€m unless noted'
c.font = Font(bold=True)
c = ws['F3']; c.value = '2026E'
c.font = Font(bold=True)
c = ws['G3']; c.value = '2027E'
c.font = Font(bold=True)
c = ws['H3']; c.value = '2028E'
c.font = Font(bold=True)
c = ws['I3']; c.value = '2029E'
c.font = Font(bold=True)
c = ws['J3']; c.value = '2030E'
c.font = Font(bold=True)
c = ws['K3']; c.value = '2031E'
c.font = Font(bold=True)
c = ws['L3']; c.value = '2032E'
c.font = Font(bold=True)
c = ws['M3']; c.value = '2033E'
c.font = Font(bold=True)
c = ws['N3']; c.value = '2034E'
c.font = Font(bold=True)
c = ws['O3']; c.value = '2035E'
c.font = Font(bold=True)
c = ws['A4']; c.value = 'Revenue'
c.font = Font(size=10.0)
c = ws['F4']; c.value = "='05 Revenue Build'!F27"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['G4']; c.value = "='05 Revenue Build'!G27"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['H4']; c.value = "='05 Revenue Build'!H27"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['I4']; c.value = "='05 Revenue Build'!I27"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['J4']; c.value = "='05 Revenue Build'!J27"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['K4']; c.value = "='05 Revenue Build'!K27"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['L4']; c.value = "='05 Revenue Build'!L27"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['M4']; c.value = "='05 Revenue Build'!M27"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['N4']; c.value = "='05 Revenue Build'!N27"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['O4']; c.value = "='05 Revenue Build'!O27"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['A5']; c.value = 'Capex % revenue (scenario)'
c.font = Font(size=10.0)
c = ws['F5']; c.value = "='03 Assumptions'!$F$14"
c.number_format = '0.0%'
c.font = Font(color='00008000', size=10.0)
c = ws['G5']; c.value = "='03 Assumptions'!$F$14"
c.number_format = '0.0%'
c.font = Font(color='00008000', size=10.0)
c = ws['H5']; c.value = "='03 Assumptions'!$F$14"
c.number_format = '0.0%'
c.font = Font(color='00008000', size=10.0)
c = ws['I5']; c.value = "='03 Assumptions'!$F$14"
c.number_format = '0.0%'
c.font = Font(color='00008000', size=10.0)
c = ws['J5']; c.value = "='03 Assumptions'!$F$14"
c.number_format = '0.0%'
c.font = Font(color='00008000', size=10.0)
c = ws['K5']; c.value = "='03 Assumptions'!$F$14"
c.number_format = '0.0%'
c.font = Font(color='00008000', size=10.0)
c = ws['L5']; c.value = "='03 Assumptions'!$F$14"
c.number_format = '0.0%'
c.font = Font(color='00008000', size=10.0)
c = ws['M5']; c.value = "='03 Assumptions'!$F$14"
c.number_format = '0.0%'
c.font = Font(color='00008000', size=10.0)
c = ws['N5']; c.value = "='03 Assumptions'!$F$14"
c.number_format = '0.0%'
c.font = Font(color='00008000', size=10.0)
c = ws['O5']; c.value = "='03 Assumptions'!$F$14"
c.number_format = '0.0%'
c.font = Font(color='00008000', size=10.0)
c = ws['A6']; c.value = 'Capex'
c.font = Font(bold=True)
c = ws['F6']; c.value = '=-F4*F5'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['G6']; c.value = '=-G4*G5'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['H6']; c.value = '=-H4*H5'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['I6']; c.value = '=-I4*I5'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['J6']; c.value = '=-J4*J5'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['K6']; c.value = '=-K4*K5'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['L6']; c.value = '=-L4*L5'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['M6']; c.value = '=-M4*M5'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['N6']; c.value = '=-N4*N5'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['O6']; c.value = '=-O4*O5'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['A7']; c.value = 'D&A (from 06)'
c.font = Font(size=10.0)
c = ws['F7']; c.value = "='06 Operating Model'!F7"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['G7']; c.value = "='06 Operating Model'!G7"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['H7']; c.value = "='06 Operating Model'!H7"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['I7']; c.value = "='06 Operating Model'!I7"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['J7']; c.value = "='06 Operating Model'!J7"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['K7']; c.value = "='06 Operating Model'!K7"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['L7']; c.value = "='06 Operating Model'!L7"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['M7']; c.value = "='06 Operating Model'!M7"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['N7']; c.value = "='06 Operating Model'!N7"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['O7']; c.value = "='06 Operating Model'!O7"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['A8']; c.value = 'Capex / D&A cover'
c.font = Font(size=10.0)
c = ws['F8']; c.value = '=F6/F7'
c.number_format = '0.0x'
c.font = Font(size=10.0)
c = ws['G8']; c.value = '=G6/G7'
c.number_format = '0.0x'
c.font = Font(size=10.0)
c = ws['H8']; c.value = '=H6/H7'
c.number_format = '0.0x'
c.font = Font(size=10.0)
c = ws['I8']; c.value = '=I6/I7'
c.number_format = '0.0x'
c.font = Font(size=10.0)
c = ws['J8']; c.value = '=J6/J7'
c.number_format = '0.0x'
c.font = Font(size=10.0)
c = ws['K8']; c.value = '=K6/K7'
c.number_format = '0.0x'
c.font = Font(size=10.0)
c = ws['L8']; c.value = '=L6/L7'
c.number_format = '0.0x'
c.font = Font(size=10.0)
c = ws['M8']; c.value = '=M6/M7'
c.number_format = '0.0x'
c.font = Font(size=10.0)
c = ws['N8']; c.value = '=N6/N7'
c.number_format = '0.0x'
c.font = Font(size=10.0)
c = ws['O8']; c.value = '=O6/O7'
c.number_format = '0.0x'
c.font = Font(size=10.0)
c = ws['A10']; c.value = '2025 actual: recurring capex 5.1% of revenue (gross 6.5% incl. one-off solar/projects). Base 5.0%.'
c.font = Font(size=10.0)

ws = wb.create_sheet('09 Free Cash Flow')
ws.column_dimensions['A'].width = 46.0
ws.column_dimensions['B'].width = 12.0
ws.column_dimensions['C'].width = 12.0
ws.column_dimensions['D'].width = 12.0
ws.column_dimensions['E'].width = 12.0
ws.column_dimensions['F'].width = 12.0
ws.column_dimensions['G'].width = 12.0
ws.column_dimensions['H'].width = 12.0
ws.column_dimensions['I'].width = 12.0
ws.column_dimensions['J'].width = 12.0
ws.column_dimensions['K'].width = 12.0
ws.column_dimensions['L'].width = 12.0
ws.column_dimensions['M'].width = 12.0
ws.column_dimensions['N'].width = 12.0
ws.column_dimensions['O'].width = 12.0
ws.column_dimensions['P'].width = 12.0
ws.column_dimensions['Q'].width = 12.0
c = ws['A1']; c.value = 'FREE CASH FLOW TO FIRM'
c.font = Font(bold=True, color='00FFFFFF', size=12.0)
c.fill = PatternFill("solid", fgColor='001F3864')
c = ws['A3']; c.value = '€m unless noted'
c.font = Font(bold=True)
c = ws['F3']; c.value = '2026E'
c.font = Font(bold=True)
c = ws['G3']; c.value = '2027E'
c.font = Font(bold=True)
c = ws['H3']; c.value = '2028E'
c.font = Font(bold=True)
c = ws['I3']; c.value = '2029E'
c.font = Font(bold=True)
c = ws['J3']; c.value = '2030E'
c.font = Font(bold=True)
c = ws['K3']; c.value = '2031E'
c.font = Font(bold=True)
c = ws['L3']; c.value = '2032E'
c.font = Font(bold=True)
c = ws['M3']; c.value = '2033E'
c.font = Font(bold=True)
c = ws['N3']; c.value = '2034E'
c.font = Font(bold=True)
c = ws['O3']; c.value = '2035E'
c.font = Font(bold=True)
c = ws['A4']; c.value = 'NOPAT'
c.font = Font(size=10.0)
c = ws['F4']; c.value = "='06 Operating Model'!F10"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['G4']; c.value = "='06 Operating Model'!G10"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['H4']; c.value = "='06 Operating Model'!H10"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['I4']; c.value = "='06 Operating Model'!I10"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['J4']; c.value = "='06 Operating Model'!J10"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['K4']; c.value = "='06 Operating Model'!K10"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['L4']; c.value = "='06 Operating Model'!L10"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['M4']; c.value = "='06 Operating Model'!M10"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['N4']; c.value = "='06 Operating Model'!N10"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['O4']; c.value = "='06 Operating Model'!O10"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['A5']; c.value = '(+) D&A'
c.font = Font(size=10.0)
c = ws['F5']; c.value = "=-'06 Operating Model'!F7"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['G5']; c.value = "=-'06 Operating Model'!G7"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['H5']; c.value = "=-'06 Operating Model'!H7"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['I5']; c.value = "=-'06 Operating Model'!I7"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['J5']; c.value = "=-'06 Operating Model'!J7"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['K5']; c.value = "=-'06 Operating Model'!K7"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['L5']; c.value = "=-'06 Operating Model'!L7"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['M5']; c.value = "=-'06 Operating Model'!M7"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['N5']; c.value = "=-'06 Operating Model'!N7"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['O5']; c.value = "=-'06 Operating Model'!O7"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['A6']; c.value = '(-) Capex'
c.font = Font(size=10.0)
c = ws['F6']; c.value = "='08 Capex & Depreciation'!F6"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['G6']; c.value = "='08 Capex & Depreciation'!G6"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['H6']; c.value = "='08 Capex & Depreciation'!H6"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['I6']; c.value = "='08 Capex & Depreciation'!I6"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['J6']; c.value = "='08 Capex & Depreciation'!J6"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['K6']; c.value = "='08 Capex & Depreciation'!K6"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['L6']; c.value = "='08 Capex & Depreciation'!L6"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['M6']; c.value = "='08 Capex & Depreciation'!M6"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['N6']; c.value = "='08 Capex & Depreciation'!N6"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['O6']; c.value = "='08 Capex & Depreciation'!O6"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['A7']; c.value = '(-/+) Change in NWC'
c.font = Font(size=10.0)
c = ws['F7']; c.value = "='07 Working Capital'!F9"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['G7']; c.value = "='07 Working Capital'!G9"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['H7']; c.value = "='07 Working Capital'!H9"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['I7']; c.value = "='07 Working Capital'!I9"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['J7']; c.value = "='07 Working Capital'!J9"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['K7']; c.value = "='07 Working Capital'!K9"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['L7']; c.value = "='07 Working Capital'!L9"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['M7']; c.value = "='07 Working Capital'!M9"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['N7']; c.value = "='07 Working Capital'!N9"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['O7']; c.value = "='07 Working Capital'!O9"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['A8']; c.value = '(-) Lease principal repayments (IFRS 16)'
c.font = Font(size=10.0)
c = ws['F8']; c.value = "=-'05 Revenue Build'!F27*'03 Assumptions'!$B$43"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['G8']; c.value = "=-'05 Revenue Build'!G27*'03 Assumptions'!$B$43"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['H8']; c.value = "=-'05 Revenue Build'!H27*'03 Assumptions'!$B$43"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['I8']; c.value = "=-'05 Revenue Build'!I27*'03 Assumptions'!$B$43"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['J8']; c.value = "=-'05 Revenue Build'!J27*'03 Assumptions'!$B$43"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['K8']; c.value = "=-'05 Revenue Build'!K27*'03 Assumptions'!$B$43"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['L8']; c.value = "=-'05 Revenue Build'!L27*'03 Assumptions'!$B$43"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['M8']; c.value = "=-'05 Revenue Build'!M27*'03 Assumptions'!$B$43"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['N8']; c.value = "=-'05 Revenue Build'!N27*'03 Assumptions'!$B$43"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['O8']; c.value = "=-'05 Revenue Build'!O27*'03 Assumptions'!$B$43"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['A9']; c.value = 'FCFF'
c.font = Font(bold=True)
c = ws['F9']; c.value = '=SUM(F4:F8)'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['G9']; c.value = '=SUM(G4:G8)'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['H9']; c.value = '=SUM(H4:H8)'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['I9']; c.value = '=SUM(I4:I8)'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['J9']; c.value = '=SUM(J4:J8)'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['K9']; c.value = '=SUM(K4:K8)'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['L9']; c.value = '=SUM(L4:L8)'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['M9']; c.value = '=SUM(M4:M8)'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['N9']; c.value = '=SUM(N4:N8)'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['O9']; c.value = '=SUM(O4:O8)'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['A10']; c.value = 'FCFF / EBITDA conversion'
c.font = Font(size=10.0)
c = ws['F10']; c.value = "=F9/'06 Operating Model'!F6"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['G10']; c.value = "=G9/'06 Operating Model'!G6"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['H10']; c.value = "=H9/'06 Operating Model'!H6"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['I10']; c.value = "=I9/'06 Operating Model'!I6"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['J10']; c.value = "=J9/'06 Operating Model'!J6"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['K10']; c.value = "=K9/'06 Operating Model'!K6"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['L10']; c.value = "=L9/'06 Operating Model'!L6"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['M10']; c.value = "=M9/'06 Operating Model'!M6"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['N10']; c.value = "=N9/'06 Operating Model'!N6"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['O10']; c.value = "=O9/'06 Operating Model'!O6"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['A12']; c.value = 'IFRS 16 v2: EBITDA keeps lease benefit; FCFF charges lease PRINCIPAL (2025: 15.1m, scaled with revenue);'
c.font = Font(size=10.0)
c = ws['A13']; c.value = 'equity bridge uses NFP EXCLUDING lease liabilities (+209.6m). Lease interest ~immaterial (net fin result ~0).'
c.font = Font(size=10.0)

ws = wb.create_sheet('10 WACC')
ws.column_dimensions['A'].width = 46.0
ws.column_dimensions['B'].width = 12.0
ws.column_dimensions['C'].width = 12.0
ws.column_dimensions['D'].width = 12.0
ws.column_dimensions['E'].width = 12.0
ws.column_dimensions['F'].width = 12.0
ws.column_dimensions['G'].width = 12.0
ws.column_dimensions['H'].width = 12.0
ws.column_dimensions['I'].width = 12.0
ws.column_dimensions['J'].width = 12.0
ws.column_dimensions['K'].width = 12.0
ws.column_dimensions['L'].width = 12.0
ws.column_dimensions['M'].width = 12.0
ws.column_dimensions['N'].width = 12.0
ws.column_dimensions['O'].width = 12.0
ws.column_dimensions['P'].width = 12.0
ws.column_dimensions['Q'].width = 12.0
c = ws['A1']; c.value = 'WACC BUILD'
c.font = Font(bold=True, color='00FFFFFF', size=12.0)
c.fill = PatternFill("solid", fgColor='001F3864')
c = ws['A3']; c.value = 'Risk-free: German Bund 10Y'
c.font = Font(size=10.0)
c = ws['B3']; c.value = 0.0313
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['D3']; c.value = 'TradingEconomics 16-Jul-26'
c.font = Font(size=10.0)
c = ws['A4']; c.value = 'Mature market ERP (Damodaran Jan-26)'
c.font = Font(size=10.0)
c = ws['B4']; c.value = 0.0423
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['D4']; c.value = 'Damodaran ctryprem'
c.font = Font(size=10.0)
c = ws['A5']; c.value = 'Levered beta (triangulated)'
c.font = Font(size=10.0)
c = ws['B5']; c.value = 0.95
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='000000FF', size=10.0)
c = ws['D5']; c.value = 'Yahoo 0.86 / TradingView 0.94'
c.font = Font(size=10.0)
c = ws['A6']; c.value = 'Weighted country risk premium'
c.font = Font(size=10.0)
c = ws['B6']; c.value = 0.006
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['D6']; c.value = 'Italy CRP 2.46% x ~11% rev + EM exposure'
c.font = Font(size=10.0)
c = ws['A8']; c.value = 'Cost of equity = rf + beta x ERP + wCRP'
c.font = Font(bold=True)
c = ws['B8']; c.value = '=B3+B5*B4+B6'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['A10']; c.value = 'Pre-tax cost of debt (hypothetical; net-cash co.)'
c.font = Font(size=10.0)
c = ws['B10']; c.value = 0.045
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['A11']; c.value = 'Tax rate'
c.font = Font(size=10.0)
c = ws['B11']; c.value = "='03 Assumptions'!$B$21"
c.number_format = '0.0%'
c.font = Font(color='00008000', size=10.0)
c = ws['A12']; c.value = 'After-tax cost of debt'
c.font = Font(size=10.0)
c = ws['B12']; c.value = '=B10*(1-B11)'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['A13']; c.value = 'Weight of debt (net cash → 0%; leases immaterial to weighting)'
c.font = Font(size=10.0)
c = ws['B13']; c.value = 0
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['A15']; c.value = 'WACC'
c.font = Font(bold=True)
c = ws['B15']; c.value = '=B8*(1-B13)+B12*B13'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c.fill = PatternFill("solid", fgColor='00FFFF00')
c = ws['A17']; c.value = '=IF(AND(B15>0.06,B15<0.10),"CHECK OK: WACC in 6-10% range","CHECK: WACC outside expected range")'
c.font = Font(bold=True, color='00FF0000', size=10.0)

ws = wb.create_sheet('11 DCF Valuation')
ws.column_dimensions['A'].width = 46.0
ws.column_dimensions['B'].width = 12.0
ws.column_dimensions['C'].width = 12.0
ws.column_dimensions['D'].width = 12.0
ws.column_dimensions['E'].width = 12.0
ws.column_dimensions['F'].width = 12.0
ws.column_dimensions['G'].width = 12.0
ws.column_dimensions['H'].width = 12.0
ws.column_dimensions['I'].width = 12.0
ws.column_dimensions['J'].width = 12.0
ws.column_dimensions['K'].width = 12.0
ws.column_dimensions['L'].width = 12.0
ws.column_dimensions['M'].width = 12.0
ws.column_dimensions['N'].width = 12.0
ws.column_dimensions['O'].width = 12.0
ws.column_dimensions['P'].width = 12.0
ws.column_dimensions['Q'].width = 12.0
c = ws['A1']; c.value = 'DCF VALUATION'
c.font = Font(bold=True, color='00FFFFFF', size=12.0)
c.fill = PatternFill("solid", fgColor='001F3864')
c = ws['A3']; c.value = '€m unless noted'
c.font = Font(bold=True)
c = ws['F3']; c.value = '2026E'
c.font = Font(bold=True)
c = ws['G3']; c.value = '2027E'
c.font = Font(bold=True)
c = ws['H3']; c.value = '2028E'
c.font = Font(bold=True)
c = ws['I3']; c.value = '2029E'
c.font = Font(bold=True)
c = ws['J3']; c.value = '2030E'
c.font = Font(bold=True)
c = ws['K3']; c.value = '2031E'
c.font = Font(bold=True)
c = ws['L3']; c.value = '2032E'
c.font = Font(bold=True)
c = ws['M3']; c.value = '2033E'
c.font = Font(bold=True)
c = ws['N3']; c.value = '2034E'
c.font = Font(bold=True)
c = ws['O3']; c.value = '2035E'
c.font = Font(bold=True)
c = ws['A4']; c.value = 'FCFF'
c.font = Font(size=10.0)
c = ws['F4']; c.value = "='09 Free Cash Flow'!F9"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['G4']; c.value = "='09 Free Cash Flow'!G9"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['H4']; c.value = "='09 Free Cash Flow'!H9"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['I4']; c.value = "='09 Free Cash Flow'!I9"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['J4']; c.value = "='09 Free Cash Flow'!J9"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['K4']; c.value = "='09 Free Cash Flow'!K9"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['L4']; c.value = "='09 Free Cash Flow'!L9"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['M4']; c.value = "='09 Free Cash Flow'!M9"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['N4']; c.value = "='09 Free Cash Flow'!N9"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['O4']; c.value = "='09 Free Cash Flow'!O9"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['A5']; c.value = 'Discount factor (year-end)'
c.font = Font(size=10.0)
c = ws['F5']; c.value = "=1/(1+'10 WACC'!$B$15)^1"
c.number_format = '0.000'
c.font = Font(size=10.0)
c = ws['G5']; c.value = "=1/(1+'10 WACC'!$B$15)^2"
c.number_format = '0.000'
c.font = Font(size=10.0)
c = ws['H5']; c.value = "=1/(1+'10 WACC'!$B$15)^3"
c.number_format = '0.000'
c.font = Font(size=10.0)
c = ws['I5']; c.value = "=1/(1+'10 WACC'!$B$15)^4"
c.number_format = '0.000'
c.font = Font(size=10.0)
c = ws['J5']; c.value = "=1/(1+'10 WACC'!$B$15)^5"
c.number_format = '0.000'
c.font = Font(size=10.0)
c = ws['K5']; c.value = "=1/(1+'10 WACC'!$B$15)^6"
c.number_format = '0.000'
c.font = Font(size=10.0)
c = ws['L5']; c.value = "=1/(1+'10 WACC'!$B$15)^7"
c.number_format = '0.000'
c.font = Font(size=10.0)
c = ws['M5']; c.value = "=1/(1+'10 WACC'!$B$15)^8"
c.number_format = '0.000'
c.font = Font(size=10.0)
c = ws['N5']; c.value = "=1/(1+'10 WACC'!$B$15)^9"
c.number_format = '0.000'
c.font = Font(size=10.0)
c = ws['O5']; c.value = "=1/(1+'10 WACC'!$B$15)^10"
c.number_format = '0.000'
c.font = Font(size=10.0)
c = ws['A6']; c.value = 'PV of FCFF'
c.font = Font(size=10.0)
c = ws['F6']; c.value = '=F4*F5'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['G6']; c.value = '=G4*G5'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['H6']; c.value = '=H4*H5'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['I6']; c.value = '=I4*I5'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['J6']; c.value = '=J4*J5'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['K6']; c.value = '=K4*K5'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['L6']; c.value = '=L4*L5'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['M6']; c.value = '=M4*M5'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['N6']; c.value = '=N4*N5'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['O6']; c.value = '=O4*O5'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['A8']; c.value = 'Sum PV explicit FCFF (2026-35)'
c.font = Font(bold=True)
c = ws['B8']; c.value = '=SUM(F6:O6)'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['A10']; c.value = 'TERMINAL VALUE — Method 1: Gordon Growth'
c.font = Font(bold=True)
c = ws['A11']; c.value = 'Terminal growth g (scenario)'
c.font = Font(size=10.0)
c = ws['B11']; c.value = "='03 Assumptions'!$F$16"
c.number_format = '0.0%'
c.font = Font(color='00008000', size=10.0)
c = ws['A12']; c.value = 'TV = FCFF2035 x (1+g) / (WACC − g)'
c.font = Font(size=10.0)
c = ws['B12']; c.value = "=O4*(1+B11)/('10 WACC'!$B$15-B11)"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['A13']; c.value = 'PV of TV'
c.font = Font(size=10.0)
c = ws['B13']; c.value = '=B12*O5'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['A14']; c.value = 'Implied exit EV/EBITDA (check vs Method 2)'
c.font = Font(size=10.0)
c = ws['B14']; c.value = "=B12/'06 Operating Model'!O6"
c.number_format = '0.0x'
c.font = Font(size=10.0)
c = ws['A16']; c.value = 'TERMINAL VALUE — Method 2: Exit multiple'
c.font = Font(bold=True)
c = ws['A17']; c.value = 'Exit multiple (scenario)'
c.font = Font(size=10.0)
c = ws['B17']; c.value = "='03 Assumptions'!$F$17"
c.number_format = '0.0x'
c.font = Font(color='00008000', size=10.0)
c = ws['A18']; c.value = 'TV = 2035 EBITDA x multiple'
c.font = Font(size=10.0)
c = ws['B18']; c.value = "='06 Operating Model'!O6*B17"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['A19']; c.value = 'PV of TV'
c.font = Font(size=10.0)
c = ws['B19']; c.value = '=B18*O5'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['A20']; c.value = 'Implied terminal growth (check)'
c.font = Font(size=10.0)
c = ws['B20']; c.value = "=('10 WACC'!$B$15*B18-O4)/(B18+O4)"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['A22']; c.value = 'ENTERPRISE → EQUITY (Gordon primary)'
c.font = Font(bold=True)
c = ws['A23']; c.value = 'Enterprise value'
c.font = Font(size=10.0)
c = ws['B23']; c.value = '=B8+B13'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['A24']; c.value = '(+) NFP excl. IFRS16 (leases charged in FCFF)'
c.font = Font(size=10.0)
c = ws['B24']; c.value = "='03 Assumptions'!$B$44"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['A25']; c.value = 'Equity value'
c.font = Font(bold=True)
c = ws['B25']; c.value = '=B23+B24'
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['A26']; c.value = 'Shares (m)'
c.font = Font(size=10.0)
c = ws['B26']; c.value = "='03 Assumptions'!$B$23"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['A27']; c.value = 'VALUE PER SHARE (€)'
c.font = Font(bold=True)
c = ws['B27']; c.value = '=B25/B26'
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c.fill = PatternFill("solid", fgColor='00FFFF00')
c = ws['A28']; c.value = 'Current price (€)'
c.font = Font(size=10.0)
c = ws['B28']; c.value = "='03 Assumptions'!$B$25"
c.number_format = '#,##0.00'
c.font = Font(color='00008000', size=10.0)
c = ws['A29']; c.value = 'Upside / (downside)'
c.font = Font(bold=True)
c = ws['B29']; c.value = '=B27/B28-1'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['A31']; c.value = 'Per share — exit-multiple method'
c.font = Font(size=10.0)
c = ws['B31']; c.value = '=(B8+B19+B24)/B26'
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['A32']; c.value = 'TV % of EV (Gordon)'
c.font = Font(size=10.0)
c = ws['B32']; c.value = '=B13/B23'
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['A34']; c.value = '=IF(\'10 WACC\'!$B$15>B11,"CHECK OK: WACC > g","ERROR: WACC <= g - Gordon invalid")'
c.font = Font(bold=True, color='00FF0000', size=10.0)
c = ws['A35']; c.value = '=IF(ABS(B27-B31)/B27<0.35,"CHECK OK: two TV methods within 35%","CHECK: TV methods diverge >35% - review exit multiple vs g")'
c.font = Font(bold=True, color='00FF0000', size=10.0)
c = ws['A37']; c.value = '5-YEAR HORIZON CROSS-CHECK (partner request)'
c.font = Font(bold=True)
c = ws['A38']; c.value = 'EV: PV FCFF 2026-30 + TV (exit mult x 2030 EBITDA)'
c.font = Font(size=10.0)
c = ws['B38']; c.value = "=SUMPRODUCT(F4:J4,F5:J5)+'06 Operating Model'!J6*'03 Assumptions'!$F$17*J5"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(size=10.0)
c = ws['A39']; c.value = '5-yr value per share'
c.font = Font(size=10.0)
c = ws['B39']; c.value = '=(B38+B24)/B26'
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['A40']; c.value = 'TV % of 5-yr EV'
c.font = Font(size=10.0)
c = ws['B40']; c.value = "='06 Operating Model'!J6*'03 Assumptions'!$F$17*J5/B38"
c.number_format = '0.0%'
c.font = Font(size=10.0)
c = ws['A42']; c.value = "Gordon-implied exit multiple (B14, currently ~9.4x) is BELOW today's ~13x trading multiple — correctly assuming a mature capital-goods re-rating, not a premium."
c = ws['A43']; c.value = 'The 12.0x exit-multiple cross-check (B17) sits above the Gordon-implied ~9.4x: the gap embeds a premium for unquantified digital optionality. Two methods agree within ~16%.'

ws = wb.create_sheet('12 Sensitivity Analysis')
ws.column_dimensions['A'].width = 46.0
ws.column_dimensions['B'].width = 12.0
ws.column_dimensions['C'].width = 12.0
ws.column_dimensions['D'].width = 12.0
ws.column_dimensions['E'].width = 12.0
ws.column_dimensions['F'].width = 12.0
ws.column_dimensions['G'].width = 12.0
ws.column_dimensions['H'].width = 12.0
ws.column_dimensions['I'].width = 12.0
ws.column_dimensions['J'].width = 12.0
ws.column_dimensions['K'].width = 12.0
ws.column_dimensions['L'].width = 12.0
ws.column_dimensions['M'].width = 12.0
ws.column_dimensions['N'].width = 12.0
ws.column_dimensions['O'].width = 12.0
ws.column_dimensions['P'].width = 12.0
ws.column_dimensions['Q'].width = 12.0
c = ws['A1']; c.value = 'SENSITIVITY ANALYSIS'
c.font = Font(bold=True, color='00FFFFFF', size=12.0)
c.fill = PatternFill("solid", fgColor='001F3864')
c = ws['A2']; c.value = 'Grids re-derive value from the live FCFF row of sheet 11; margin×growth grid uses a simplified engine (uniform growth/margin) noted below.'
c.font = Font(size=10.0)
c = ws['A4']; c.value = 'GRID 1 — WACC (rows) × terminal growth (cols)'
c.font = Font(bold=True)
c = ws['C5']; c.value = 0.015
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['D5']; c.value = 0.02
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['E5']; c.value = 0.0225
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['F5']; c.value = 0.025
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['G5']; c.value = 0.03
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['B6']; c.value = 0.07
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['C6']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B6)^'05 Revenue Build'!$F$4:$O$4)+'11 DCF Valuation'!$O$4*(1+C$5)/($B6-C$5)/(1+$B6)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['D6']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B6)^'05 Revenue Build'!$F$4:$O$4)+'11 DCF Valuation'!$O$4*(1+D$5)/($B6-D$5)/(1+$B6)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['E6']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B6)^'05 Revenue Build'!$F$4:$O$4)+'11 DCF Valuation'!$O$4*(1+E$5)/($B6-E$5)/(1+$B6)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['F6']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B6)^'05 Revenue Build'!$F$4:$O$4)+'11 DCF Valuation'!$O$4*(1+F$5)/($B6-F$5)/(1+$B6)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['G6']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B6)^'05 Revenue Build'!$F$4:$O$4)+'11 DCF Valuation'!$O$4*(1+G$5)/($B6-G$5)/(1+$B6)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['B7']; c.value = 0.0725
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['C7']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B7)^'05 Revenue Build'!$F$4:$O$4)+'11 DCF Valuation'!$O$4*(1+C$5)/($B7-C$5)/(1+$B7)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['D7']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B7)^'05 Revenue Build'!$F$4:$O$4)+'11 DCF Valuation'!$O$4*(1+D$5)/($B7-D$5)/(1+$B7)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['E7']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B7)^'05 Revenue Build'!$F$4:$O$4)+'11 DCF Valuation'!$O$4*(1+E$5)/($B7-E$5)/(1+$B7)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['F7']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B7)^'05 Revenue Build'!$F$4:$O$4)+'11 DCF Valuation'!$O$4*(1+F$5)/($B7-F$5)/(1+$B7)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['G7']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B7)^'05 Revenue Build'!$F$4:$O$4)+'11 DCF Valuation'!$O$4*(1+G$5)/($B7-G$5)/(1+$B7)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['B8']; c.value = 0.075
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['C8']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B8)^'05 Revenue Build'!$F$4:$O$4)+'11 DCF Valuation'!$O$4*(1+C$5)/($B8-C$5)/(1+$B8)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['D8']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B8)^'05 Revenue Build'!$F$4:$O$4)+'11 DCF Valuation'!$O$4*(1+D$5)/($B8-D$5)/(1+$B8)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['E8']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B8)^'05 Revenue Build'!$F$4:$O$4)+'11 DCF Valuation'!$O$4*(1+E$5)/($B8-E$5)/(1+$B8)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['F8']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B8)^'05 Revenue Build'!$F$4:$O$4)+'11 DCF Valuation'!$O$4*(1+F$5)/($B8-F$5)/(1+$B8)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['G8']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B8)^'05 Revenue Build'!$F$4:$O$4)+'11 DCF Valuation'!$O$4*(1+G$5)/($B8-G$5)/(1+$B8)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['B9']; c.value = 0.0775
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['C9']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B9)^'05 Revenue Build'!$F$4:$O$4)+'11 DCF Valuation'!$O$4*(1+C$5)/($B9-C$5)/(1+$B9)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['D9']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B9)^'05 Revenue Build'!$F$4:$O$4)+'11 DCF Valuation'!$O$4*(1+D$5)/($B9-D$5)/(1+$B9)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['E9']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B9)^'05 Revenue Build'!$F$4:$O$4)+'11 DCF Valuation'!$O$4*(1+E$5)/($B9-E$5)/(1+$B9)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['F9']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B9)^'05 Revenue Build'!$F$4:$O$4)+'11 DCF Valuation'!$O$4*(1+F$5)/($B9-F$5)/(1+$B9)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['G9']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B9)^'05 Revenue Build'!$F$4:$O$4)+'11 DCF Valuation'!$O$4*(1+G$5)/($B9-G$5)/(1+$B9)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['B10']; c.value = 0.08
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['C10']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B10)^'05 Revenue Build'!$F$4:$O$4)+'11 DCF Valuation'!$O$4*(1+C$5)/($B10-C$5)/(1+$B10)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['D10']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B10)^'05 Revenue Build'!$F$4:$O$4)+'11 DCF Valuation'!$O$4*(1+D$5)/($B10-D$5)/(1+$B10)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['E10']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B10)^'05 Revenue Build'!$F$4:$O$4)+'11 DCF Valuation'!$O$4*(1+E$5)/($B10-E$5)/(1+$B10)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['F10']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B10)^'05 Revenue Build'!$F$4:$O$4)+'11 DCF Valuation'!$O$4*(1+F$5)/($B10-F$5)/(1+$B10)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['G10']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B10)^'05 Revenue Build'!$F$4:$O$4)+'11 DCF Valuation'!$O$4*(1+G$5)/($B10-G$5)/(1+$B10)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['B11']; c.value = 0.085
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['C11']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B11)^'05 Revenue Build'!$F$4:$O$4)+'11 DCF Valuation'!$O$4*(1+C$5)/($B11-C$5)/(1+$B11)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['D11']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B11)^'05 Revenue Build'!$F$4:$O$4)+'11 DCF Valuation'!$O$4*(1+D$5)/($B11-D$5)/(1+$B11)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['E11']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B11)^'05 Revenue Build'!$F$4:$O$4)+'11 DCF Valuation'!$O$4*(1+E$5)/($B11-E$5)/(1+$B11)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['F11']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B11)^'05 Revenue Build'!$F$4:$O$4)+'11 DCF Valuation'!$O$4*(1+F$5)/($B11-F$5)/(1+$B11)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['G11']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B11)^'05 Revenue Build'!$F$4:$O$4)+'11 DCF Valuation'!$O$4*(1+G$5)/($B11-G$5)/(1+$B11)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['A14']; c.value = 'GRID 2 — WACC (rows) × exit EV/EBITDA (cols)'
c.font = Font(bold=True)
c = ws['C15']; c.value = 9
c.number_format = '0.0x'
c.font = Font(color='000000FF', size=10.0)
c = ws['D15']; c.value = 10
c.number_format = '0.0x'
c.font = Font(color='000000FF', size=10.0)
c = ws['E15']; c.value = 11
c.number_format = '0.0x'
c.font = Font(color='000000FF', size=10.0)
c = ws['F15']; c.value = 12
c.number_format = '0.0x'
c.font = Font(color='000000FF', size=10.0)
c = ws['G15']; c.value = 13
c.number_format = '0.0x'
c.font = Font(color='000000FF', size=10.0)
c = ws['H15']; c.value = 14
c.number_format = '0.0x'
c.font = Font(color='000000FF', size=10.0)
c = ws['B16']; c.value = 0.07
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['C16']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B16)^'05 Revenue Build'!$F$4:$O$4)+'06 Operating Model'!$O$6*C$15/(1+$B16)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['D16']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B16)^'05 Revenue Build'!$F$4:$O$4)+'06 Operating Model'!$O$6*D$15/(1+$B16)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['E16']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B16)^'05 Revenue Build'!$F$4:$O$4)+'06 Operating Model'!$O$6*E$15/(1+$B16)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['F16']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B16)^'05 Revenue Build'!$F$4:$O$4)+'06 Operating Model'!$O$6*F$15/(1+$B16)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['G16']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B16)^'05 Revenue Build'!$F$4:$O$4)+'06 Operating Model'!$O$6*G$15/(1+$B16)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['H16']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B16)^'05 Revenue Build'!$F$4:$O$4)+'06 Operating Model'!$O$6*H$15/(1+$B16)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['B17']; c.value = 0.0725
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['C17']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B17)^'05 Revenue Build'!$F$4:$O$4)+'06 Operating Model'!$O$6*C$15/(1+$B17)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['D17']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B17)^'05 Revenue Build'!$F$4:$O$4)+'06 Operating Model'!$O$6*D$15/(1+$B17)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['E17']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B17)^'05 Revenue Build'!$F$4:$O$4)+'06 Operating Model'!$O$6*E$15/(1+$B17)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['F17']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B17)^'05 Revenue Build'!$F$4:$O$4)+'06 Operating Model'!$O$6*F$15/(1+$B17)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['G17']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B17)^'05 Revenue Build'!$F$4:$O$4)+'06 Operating Model'!$O$6*G$15/(1+$B17)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['H17']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B17)^'05 Revenue Build'!$F$4:$O$4)+'06 Operating Model'!$O$6*H$15/(1+$B17)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['B18']; c.value = 0.075
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['C18']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B18)^'05 Revenue Build'!$F$4:$O$4)+'06 Operating Model'!$O$6*C$15/(1+$B18)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['D18']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B18)^'05 Revenue Build'!$F$4:$O$4)+'06 Operating Model'!$O$6*D$15/(1+$B18)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['E18']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B18)^'05 Revenue Build'!$F$4:$O$4)+'06 Operating Model'!$O$6*E$15/(1+$B18)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['F18']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B18)^'05 Revenue Build'!$F$4:$O$4)+'06 Operating Model'!$O$6*F$15/(1+$B18)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['G18']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B18)^'05 Revenue Build'!$F$4:$O$4)+'06 Operating Model'!$O$6*G$15/(1+$B18)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['H18']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B18)^'05 Revenue Build'!$F$4:$O$4)+'06 Operating Model'!$O$6*H$15/(1+$B18)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['B19']; c.value = 0.0775
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['C19']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B19)^'05 Revenue Build'!$F$4:$O$4)+'06 Operating Model'!$O$6*C$15/(1+$B19)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['D19']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B19)^'05 Revenue Build'!$F$4:$O$4)+'06 Operating Model'!$O$6*D$15/(1+$B19)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['E19']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B19)^'05 Revenue Build'!$F$4:$O$4)+'06 Operating Model'!$O$6*E$15/(1+$B19)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['F19']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B19)^'05 Revenue Build'!$F$4:$O$4)+'06 Operating Model'!$O$6*F$15/(1+$B19)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['G19']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B19)^'05 Revenue Build'!$F$4:$O$4)+'06 Operating Model'!$O$6*G$15/(1+$B19)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['H19']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B19)^'05 Revenue Build'!$F$4:$O$4)+'06 Operating Model'!$O$6*H$15/(1+$B19)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['B20']; c.value = 0.08
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['C20']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B20)^'05 Revenue Build'!$F$4:$O$4)+'06 Operating Model'!$O$6*C$15/(1+$B20)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['D20']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B20)^'05 Revenue Build'!$F$4:$O$4)+'06 Operating Model'!$O$6*D$15/(1+$B20)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['E20']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B20)^'05 Revenue Build'!$F$4:$O$4)+'06 Operating Model'!$O$6*E$15/(1+$B20)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['F20']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B20)^'05 Revenue Build'!$F$4:$O$4)+'06 Operating Model'!$O$6*F$15/(1+$B20)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['G20']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B20)^'05 Revenue Build'!$F$4:$O$4)+'06 Operating Model'!$O$6*G$15/(1+$B20)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['H20']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B20)^'05 Revenue Build'!$F$4:$O$4)+'06 Operating Model'!$O$6*H$15/(1+$B20)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['B21']; c.value = 0.085
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['C21']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B21)^'05 Revenue Build'!$F$4:$O$4)+'06 Operating Model'!$O$6*C$15/(1+$B21)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['D21']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B21)^'05 Revenue Build'!$F$4:$O$4)+'06 Operating Model'!$O$6*D$15/(1+$B21)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['E21']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B21)^'05 Revenue Build'!$F$4:$O$4)+'06 Operating Model'!$O$6*E$15/(1+$B21)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['F21']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B21)^'05 Revenue Build'!$F$4:$O$4)+'06 Operating Model'!$O$6*F$15/(1+$B21)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['G21']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B21)^'05 Revenue Build'!$F$4:$O$4)+'06 Operating Model'!$O$6*G$15/(1+$B21)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['H21']; c.value = "=(SUMPRODUCT('11 DCF Valuation'!$F$4:$O$4,1/(1+$B21)^'05 Revenue Build'!$F$4:$O$4)+'06 Operating Model'!$O$6*H$15/(1+$B21)^10+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['A24']; c.value = 'GRID 3 - EBITDA margin endpoint (rows) x revenue CAGR 2026-35 (cols) - EXACT for stated axes'
c.font = Font(bold=True)
c = ws['A25']; c.value = 'Engine: Rev_t=1019.3x(1+CAGR)^t; margin linear 21.6% to endpoint by yr7; FCFF=EBITDA(1-t)+t*DA-capex-lease-dNWC; Gordon TV.'
c.font = Font(size=10.0)
c = ws['C26']; c.value = 0.03
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['D26']; c.value = 0.05
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['E26']; c.value = 0.07
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['F26']; c.value = 0.09
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['G26']; c.value = 0.11
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['B27']; c.value = 0.215
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['C27']; c.value = "=(SUMPRODUCT( (1019.3*(1+C$26)^'05 Revenue Build'!$F$4:$O$4) * ( ((0.216+($B27-0.216)*(('05 Revenue Build'!$F$4:$O$4)*('05 Revenue Build'!$F$4:$O$4<=7)+7*('05 Revenue Build'!$F$4:$O$4>7))/7)-'03 Assumptions'!$B$22)*(1-'03 Assumptions'!$B$21) + '03 Assumptions'!$B$22 - '03 Assumptions'!$F$14 - '03 Assumptions'!$B$43 ) / (1+'10 WACC'!$B$15)^'05 Revenue Build'!$F$4:$O$4 )- SUMPRODUCT( '07 Working Capital'!$B$12*1019.3*((1+C$26)^'05 Revenue Build'!$F$4:$O$4-(1+C$26)^('05 Revenue Build'!$F$4:$O$4-1)) / (1+'10 WACC'!$B$15)^'05 Revenue Build'!$F$4:$O$4 )+ (1019.3*(1+C$26)^10*(($B27-'03 Assumptions'!$B$22)*(1-'03 Assumptions'!$B$21)+'03 Assumptions'!$B$22-'03 Assumptions'!$F$14-'03 Assumptions'!$B$43-'07 Working Capital'!$B$12*C$26)*(1+'03 Assumptions'!$F$16)/('10 WACC'!$B$15-'03 Assumptions'!$F$16))/(1+'10 WACC'!$B$15)^10+ '03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['D27']; c.value = "=(SUMPRODUCT( (1019.3*(1+D$26)^'05 Revenue Build'!$F$4:$O$4) * ( ((0.216+($B27-0.216)*(('05 Revenue Build'!$F$4:$O$4)*('05 Revenue Build'!$F$4:$O$4<=7)+7*('05 Revenue Build'!$F$4:$O$4>7))/7)-'03 Assumptions'!$B$22)*(1-'03 Assumptions'!$B$21) + '03 Assumptions'!$B$22 - '03 Assumptions'!$F$14 - '03 Assumptions'!$B$43 ) / (1+'10 WACC'!$B$15)^'05 Revenue Build'!$F$4:$O$4 )- SUMPRODUCT( '07 Working Capital'!$B$12*1019.3*((1+D$26)^'05 Revenue Build'!$F$4:$O$4-(1+D$26)^('05 Revenue Build'!$F$4:$O$4-1)) / (1+'10 WACC'!$B$15)^'05 Revenue Build'!$F$4:$O$4 )+ (1019.3*(1+D$26)^10*(($B27-'03 Assumptions'!$B$22)*(1-'03 Assumptions'!$B$21)+'03 Assumptions'!$B$22-'03 Assumptions'!$F$14-'03 Assumptions'!$B$43-'07 Working Capital'!$B$12*D$26)*(1+'03 Assumptions'!$F$16)/('10 WACC'!$B$15-'03 Assumptions'!$F$16))/(1+'10 WACC'!$B$15)^10+ '03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['E27']; c.value = "=(SUMPRODUCT( (1019.3*(1+E$26)^'05 Revenue Build'!$F$4:$O$4) * ( ((0.216+($B27-0.216)*(('05 Revenue Build'!$F$4:$O$4)*('05 Revenue Build'!$F$4:$O$4<=7)+7*('05 Revenue Build'!$F$4:$O$4>7))/7)-'03 Assumptions'!$B$22)*(1-'03 Assumptions'!$B$21) + '03 Assumptions'!$B$22 - '03 Assumptions'!$F$14 - '03 Assumptions'!$B$43 ) / (1+'10 WACC'!$B$15)^'05 Revenue Build'!$F$4:$O$4 )- SUMPRODUCT( '07 Working Capital'!$B$12*1019.3*((1+E$26)^'05 Revenue Build'!$F$4:$O$4-(1+E$26)^('05 Revenue Build'!$F$4:$O$4-1)) / (1+'10 WACC'!$B$15)^'05 Revenue Build'!$F$4:$O$4 )+ (1019.3*(1+E$26)^10*(($B27-'03 Assumptions'!$B$22)*(1-'03 Assumptions'!$B$21)+'03 Assumptions'!$B$22-'03 Assumptions'!$F$14-'03 Assumptions'!$B$43-'07 Working Capital'!$B$12*E$26)*(1+'03 Assumptions'!$F$16)/('10 WACC'!$B$15-'03 Assumptions'!$F$16))/(1+'10 WACC'!$B$15)^10+ '03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['F27']; c.value = "=(SUMPRODUCT( (1019.3*(1+F$26)^'05 Revenue Build'!$F$4:$O$4) * ( ((0.216+($B27-0.216)*(('05 Revenue Build'!$F$4:$O$4)*('05 Revenue Build'!$F$4:$O$4<=7)+7*('05 Revenue Build'!$F$4:$O$4>7))/7)-'03 Assumptions'!$B$22)*(1-'03 Assumptions'!$B$21) + '03 Assumptions'!$B$22 - '03 Assumptions'!$F$14 - '03 Assumptions'!$B$43 ) / (1+'10 WACC'!$B$15)^'05 Revenue Build'!$F$4:$O$4 )- SUMPRODUCT( '07 Working Capital'!$B$12*1019.3*((1+F$26)^'05 Revenue Build'!$F$4:$O$4-(1+F$26)^('05 Revenue Build'!$F$4:$O$4-1)) / (1+'10 WACC'!$B$15)^'05 Revenue Build'!$F$4:$O$4 )+ (1019.3*(1+F$26)^10*(($B27-'03 Assumptions'!$B$22)*(1-'03 Assumptions'!$B$21)+'03 Assumptions'!$B$22-'03 Assumptions'!$F$14-'03 Assumptions'!$B$43-'07 Working Capital'!$B$12*F$26)*(1+'03 Assumptions'!$F$16)/('10 WACC'!$B$15-'03 Assumptions'!$F$16))/(1+'10 WACC'!$B$15)^10+ '03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['G27']; c.value = "=(SUMPRODUCT( (1019.3*(1+G$26)^'05 Revenue Build'!$F$4:$O$4) * ( ((0.216+($B27-0.216)*(('05 Revenue Build'!$F$4:$O$4)*('05 Revenue Build'!$F$4:$O$4<=7)+7*('05 Revenue Build'!$F$4:$O$4>7))/7)-'03 Assumptions'!$B$22)*(1-'03 Assumptions'!$B$21) + '03 Assumptions'!$B$22 - '03 Assumptions'!$F$14 - '03 Assumptions'!$B$43 ) / (1+'10 WACC'!$B$15)^'05 Revenue Build'!$F$4:$O$4 )- SUMPRODUCT( '07 Working Capital'!$B$12*1019.3*((1+G$26)^'05 Revenue Build'!$F$4:$O$4-(1+G$26)^('05 Revenue Build'!$F$4:$O$4-1)) / (1+'10 WACC'!$B$15)^'05 Revenue Build'!$F$4:$O$4 )+ (1019.3*(1+G$26)^10*(($B27-'03 Assumptions'!$B$22)*(1-'03 Assumptions'!$B$21)+'03 Assumptions'!$B$22-'03 Assumptions'!$F$14-'03 Assumptions'!$B$43-'07 Working Capital'!$B$12*G$26)*(1+'03 Assumptions'!$F$16)/('10 WACC'!$B$15-'03 Assumptions'!$F$16))/(1+'10 WACC'!$B$15)^10+ '03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['B28']; c.value = 0.225
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['C28']; c.value = "=(SUMPRODUCT( (1019.3*(1+C$26)^'05 Revenue Build'!$F$4:$O$4) * ( ((0.216+($B28-0.216)*(('05 Revenue Build'!$F$4:$O$4)*('05 Revenue Build'!$F$4:$O$4<=7)+7*('05 Revenue Build'!$F$4:$O$4>7))/7)-'03 Assumptions'!$B$22)*(1-'03 Assumptions'!$B$21) + '03 Assumptions'!$B$22 - '03 Assumptions'!$F$14 - '03 Assumptions'!$B$43 ) / (1+'10 WACC'!$B$15)^'05 Revenue Build'!$F$4:$O$4 )- SUMPRODUCT( '07 Working Capital'!$B$12*1019.3*((1+C$26)^'05 Revenue Build'!$F$4:$O$4-(1+C$26)^('05 Revenue Build'!$F$4:$O$4-1)) / (1+'10 WACC'!$B$15)^'05 Revenue Build'!$F$4:$O$4 )+ (1019.3*(1+C$26)^10*(($B28-'03 Assumptions'!$B$22)*(1-'03 Assumptions'!$B$21)+'03 Assumptions'!$B$22-'03 Assumptions'!$F$14-'03 Assumptions'!$B$43-'07 Working Capital'!$B$12*C$26)*(1+'03 Assumptions'!$F$16)/('10 WACC'!$B$15-'03 Assumptions'!$F$16))/(1+'10 WACC'!$B$15)^10+ '03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['D28']; c.value = "=(SUMPRODUCT( (1019.3*(1+D$26)^'05 Revenue Build'!$F$4:$O$4) * ( ((0.216+($B28-0.216)*(('05 Revenue Build'!$F$4:$O$4)*('05 Revenue Build'!$F$4:$O$4<=7)+7*('05 Revenue Build'!$F$4:$O$4>7))/7)-'03 Assumptions'!$B$22)*(1-'03 Assumptions'!$B$21) + '03 Assumptions'!$B$22 - '03 Assumptions'!$F$14 - '03 Assumptions'!$B$43 ) / (1+'10 WACC'!$B$15)^'05 Revenue Build'!$F$4:$O$4 )- SUMPRODUCT( '07 Working Capital'!$B$12*1019.3*((1+D$26)^'05 Revenue Build'!$F$4:$O$4-(1+D$26)^('05 Revenue Build'!$F$4:$O$4-1)) / (1+'10 WACC'!$B$15)^'05 Revenue Build'!$F$4:$O$4 )+ (1019.3*(1+D$26)^10*(($B28-'03 Assumptions'!$B$22)*(1-'03 Assumptions'!$B$21)+'03 Assumptions'!$B$22-'03 Assumptions'!$F$14-'03 Assumptions'!$B$43-'07 Working Capital'!$B$12*D$26)*(1+'03 Assumptions'!$F$16)/('10 WACC'!$B$15-'03 Assumptions'!$F$16))/(1+'10 WACC'!$B$15)^10+ '03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['E28']; c.value = "=(SUMPRODUCT( (1019.3*(1+E$26)^'05 Revenue Build'!$F$4:$O$4) * ( ((0.216+($B28-0.216)*(('05 Revenue Build'!$F$4:$O$4)*('05 Revenue Build'!$F$4:$O$4<=7)+7*('05 Revenue Build'!$F$4:$O$4>7))/7)-'03 Assumptions'!$B$22)*(1-'03 Assumptions'!$B$21) + '03 Assumptions'!$B$22 - '03 Assumptions'!$F$14 - '03 Assumptions'!$B$43 ) / (1+'10 WACC'!$B$15)^'05 Revenue Build'!$F$4:$O$4 )- SUMPRODUCT( '07 Working Capital'!$B$12*1019.3*((1+E$26)^'05 Revenue Build'!$F$4:$O$4-(1+E$26)^('05 Revenue Build'!$F$4:$O$4-1)) / (1+'10 WACC'!$B$15)^'05 Revenue Build'!$F$4:$O$4 )+ (1019.3*(1+E$26)^10*(($B28-'03 Assumptions'!$B$22)*(1-'03 Assumptions'!$B$21)+'03 Assumptions'!$B$22-'03 Assumptions'!$F$14-'03 Assumptions'!$B$43-'07 Working Capital'!$B$12*E$26)*(1+'03 Assumptions'!$F$16)/('10 WACC'!$B$15-'03 Assumptions'!$F$16))/(1+'10 WACC'!$B$15)^10+ '03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['F28']; c.value = "=(SUMPRODUCT( (1019.3*(1+F$26)^'05 Revenue Build'!$F$4:$O$4) * ( ((0.216+($B28-0.216)*(('05 Revenue Build'!$F$4:$O$4)*('05 Revenue Build'!$F$4:$O$4<=7)+7*('05 Revenue Build'!$F$4:$O$4>7))/7)-'03 Assumptions'!$B$22)*(1-'03 Assumptions'!$B$21) + '03 Assumptions'!$B$22 - '03 Assumptions'!$F$14 - '03 Assumptions'!$B$43 ) / (1+'10 WACC'!$B$15)^'05 Revenue Build'!$F$4:$O$4 )- SUMPRODUCT( '07 Working Capital'!$B$12*1019.3*((1+F$26)^'05 Revenue Build'!$F$4:$O$4-(1+F$26)^('05 Revenue Build'!$F$4:$O$4-1)) / (1+'10 WACC'!$B$15)^'05 Revenue Build'!$F$4:$O$4 )+ (1019.3*(1+F$26)^10*(($B28-'03 Assumptions'!$B$22)*(1-'03 Assumptions'!$B$21)+'03 Assumptions'!$B$22-'03 Assumptions'!$F$14-'03 Assumptions'!$B$43-'07 Working Capital'!$B$12*F$26)*(1+'03 Assumptions'!$F$16)/('10 WACC'!$B$15-'03 Assumptions'!$F$16))/(1+'10 WACC'!$B$15)^10+ '03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['G28']; c.value = "=(SUMPRODUCT( (1019.3*(1+G$26)^'05 Revenue Build'!$F$4:$O$4) * ( ((0.216+($B28-0.216)*(('05 Revenue Build'!$F$4:$O$4)*('05 Revenue Build'!$F$4:$O$4<=7)+7*('05 Revenue Build'!$F$4:$O$4>7))/7)-'03 Assumptions'!$B$22)*(1-'03 Assumptions'!$B$21) + '03 Assumptions'!$B$22 - '03 Assumptions'!$F$14 - '03 Assumptions'!$B$43 ) / (1+'10 WACC'!$B$15)^'05 Revenue Build'!$F$4:$O$4 )- SUMPRODUCT( '07 Working Capital'!$B$12*1019.3*((1+G$26)^'05 Revenue Build'!$F$4:$O$4-(1+G$26)^('05 Revenue Build'!$F$4:$O$4-1)) / (1+'10 WACC'!$B$15)^'05 Revenue Build'!$F$4:$O$4 )+ (1019.3*(1+G$26)^10*(($B28-'03 Assumptions'!$B$22)*(1-'03 Assumptions'!$B$21)+'03 Assumptions'!$B$22-'03 Assumptions'!$F$14-'03 Assumptions'!$B$43-'07 Working Capital'!$B$12*G$26)*(1+'03 Assumptions'!$F$16)/('10 WACC'!$B$15-'03 Assumptions'!$F$16))/(1+'10 WACC'!$B$15)^10+ '03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['B29']; c.value = 0.235
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['C29']; c.value = "=(SUMPRODUCT( (1019.3*(1+C$26)^'05 Revenue Build'!$F$4:$O$4) * ( ((0.216+($B29-0.216)*(('05 Revenue Build'!$F$4:$O$4)*('05 Revenue Build'!$F$4:$O$4<=7)+7*('05 Revenue Build'!$F$4:$O$4>7))/7)-'03 Assumptions'!$B$22)*(1-'03 Assumptions'!$B$21) + '03 Assumptions'!$B$22 - '03 Assumptions'!$F$14 - '03 Assumptions'!$B$43 ) / (1+'10 WACC'!$B$15)^'05 Revenue Build'!$F$4:$O$4 )- SUMPRODUCT( '07 Working Capital'!$B$12*1019.3*((1+C$26)^'05 Revenue Build'!$F$4:$O$4-(1+C$26)^('05 Revenue Build'!$F$4:$O$4-1)) / (1+'10 WACC'!$B$15)^'05 Revenue Build'!$F$4:$O$4 )+ (1019.3*(1+C$26)^10*(($B29-'03 Assumptions'!$B$22)*(1-'03 Assumptions'!$B$21)+'03 Assumptions'!$B$22-'03 Assumptions'!$F$14-'03 Assumptions'!$B$43-'07 Working Capital'!$B$12*C$26)*(1+'03 Assumptions'!$F$16)/('10 WACC'!$B$15-'03 Assumptions'!$F$16))/(1+'10 WACC'!$B$15)^10+ '03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['D29']; c.value = "=(SUMPRODUCT( (1019.3*(1+D$26)^'05 Revenue Build'!$F$4:$O$4) * ( ((0.216+($B29-0.216)*(('05 Revenue Build'!$F$4:$O$4)*('05 Revenue Build'!$F$4:$O$4<=7)+7*('05 Revenue Build'!$F$4:$O$4>7))/7)-'03 Assumptions'!$B$22)*(1-'03 Assumptions'!$B$21) + '03 Assumptions'!$B$22 - '03 Assumptions'!$F$14 - '03 Assumptions'!$B$43 ) / (1+'10 WACC'!$B$15)^'05 Revenue Build'!$F$4:$O$4 )- SUMPRODUCT( '07 Working Capital'!$B$12*1019.3*((1+D$26)^'05 Revenue Build'!$F$4:$O$4-(1+D$26)^('05 Revenue Build'!$F$4:$O$4-1)) / (1+'10 WACC'!$B$15)^'05 Revenue Build'!$F$4:$O$4 )+ (1019.3*(1+D$26)^10*(($B29-'03 Assumptions'!$B$22)*(1-'03 Assumptions'!$B$21)+'03 Assumptions'!$B$22-'03 Assumptions'!$F$14-'03 Assumptions'!$B$43-'07 Working Capital'!$B$12*D$26)*(1+'03 Assumptions'!$F$16)/('10 WACC'!$B$15-'03 Assumptions'!$F$16))/(1+'10 WACC'!$B$15)^10+ '03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['E29']; c.value = "=(SUMPRODUCT( (1019.3*(1+E$26)^'05 Revenue Build'!$F$4:$O$4) * ( ((0.216+($B29-0.216)*(('05 Revenue Build'!$F$4:$O$4)*('05 Revenue Build'!$F$4:$O$4<=7)+7*('05 Revenue Build'!$F$4:$O$4>7))/7)-'03 Assumptions'!$B$22)*(1-'03 Assumptions'!$B$21) + '03 Assumptions'!$B$22 - '03 Assumptions'!$F$14 - '03 Assumptions'!$B$43 ) / (1+'10 WACC'!$B$15)^'05 Revenue Build'!$F$4:$O$4 )- SUMPRODUCT( '07 Working Capital'!$B$12*1019.3*((1+E$26)^'05 Revenue Build'!$F$4:$O$4-(1+E$26)^('05 Revenue Build'!$F$4:$O$4-1)) / (1+'10 WACC'!$B$15)^'05 Revenue Build'!$F$4:$O$4 )+ (1019.3*(1+E$26)^10*(($B29-'03 Assumptions'!$B$22)*(1-'03 Assumptions'!$B$21)+'03 Assumptions'!$B$22-'03 Assumptions'!$F$14-'03 Assumptions'!$B$43-'07 Working Capital'!$B$12*E$26)*(1+'03 Assumptions'!$F$16)/('10 WACC'!$B$15-'03 Assumptions'!$F$16))/(1+'10 WACC'!$B$15)^10+ '03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['F29']; c.value = "=(SUMPRODUCT( (1019.3*(1+F$26)^'05 Revenue Build'!$F$4:$O$4) * ( ((0.216+($B29-0.216)*(('05 Revenue Build'!$F$4:$O$4)*('05 Revenue Build'!$F$4:$O$4<=7)+7*('05 Revenue Build'!$F$4:$O$4>7))/7)-'03 Assumptions'!$B$22)*(1-'03 Assumptions'!$B$21) + '03 Assumptions'!$B$22 - '03 Assumptions'!$F$14 - '03 Assumptions'!$B$43 ) / (1+'10 WACC'!$B$15)^'05 Revenue Build'!$F$4:$O$4 )- SUMPRODUCT( '07 Working Capital'!$B$12*1019.3*((1+F$26)^'05 Revenue Build'!$F$4:$O$4-(1+F$26)^('05 Revenue Build'!$F$4:$O$4-1)) / (1+'10 WACC'!$B$15)^'05 Revenue Build'!$F$4:$O$4 )+ (1019.3*(1+F$26)^10*(($B29-'03 Assumptions'!$B$22)*(1-'03 Assumptions'!$B$21)+'03 Assumptions'!$B$22-'03 Assumptions'!$F$14-'03 Assumptions'!$B$43-'07 Working Capital'!$B$12*F$26)*(1+'03 Assumptions'!$F$16)/('10 WACC'!$B$15-'03 Assumptions'!$F$16))/(1+'10 WACC'!$B$15)^10+ '03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['G29']; c.value = "=(SUMPRODUCT( (1019.3*(1+G$26)^'05 Revenue Build'!$F$4:$O$4) * ( ((0.216+($B29-0.216)*(('05 Revenue Build'!$F$4:$O$4)*('05 Revenue Build'!$F$4:$O$4<=7)+7*('05 Revenue Build'!$F$4:$O$4>7))/7)-'03 Assumptions'!$B$22)*(1-'03 Assumptions'!$B$21) + '03 Assumptions'!$B$22 - '03 Assumptions'!$F$14 - '03 Assumptions'!$B$43 ) / (1+'10 WACC'!$B$15)^'05 Revenue Build'!$F$4:$O$4 )- SUMPRODUCT( '07 Working Capital'!$B$12*1019.3*((1+G$26)^'05 Revenue Build'!$F$4:$O$4-(1+G$26)^('05 Revenue Build'!$F$4:$O$4-1)) / (1+'10 WACC'!$B$15)^'05 Revenue Build'!$F$4:$O$4 )+ (1019.3*(1+G$26)^10*(($B29-'03 Assumptions'!$B$22)*(1-'03 Assumptions'!$B$21)+'03 Assumptions'!$B$22-'03 Assumptions'!$F$14-'03 Assumptions'!$B$43-'07 Working Capital'!$B$12*G$26)*(1+'03 Assumptions'!$F$16)/('10 WACC'!$B$15-'03 Assumptions'!$F$16))/(1+'10 WACC'!$B$15)^10+ '03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['B30']; c.value = 0.245
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['C30']; c.value = "=(SUMPRODUCT( (1019.3*(1+C$26)^'05 Revenue Build'!$F$4:$O$4) * ( ((0.216+($B30-0.216)*(('05 Revenue Build'!$F$4:$O$4)*('05 Revenue Build'!$F$4:$O$4<=7)+7*('05 Revenue Build'!$F$4:$O$4>7))/7)-'03 Assumptions'!$B$22)*(1-'03 Assumptions'!$B$21) + '03 Assumptions'!$B$22 - '03 Assumptions'!$F$14 - '03 Assumptions'!$B$43 ) / (1+'10 WACC'!$B$15)^'05 Revenue Build'!$F$4:$O$4 )- SUMPRODUCT( '07 Working Capital'!$B$12*1019.3*((1+C$26)^'05 Revenue Build'!$F$4:$O$4-(1+C$26)^('05 Revenue Build'!$F$4:$O$4-1)) / (1+'10 WACC'!$B$15)^'05 Revenue Build'!$F$4:$O$4 )+ (1019.3*(1+C$26)^10*(($B30-'03 Assumptions'!$B$22)*(1-'03 Assumptions'!$B$21)+'03 Assumptions'!$B$22-'03 Assumptions'!$F$14-'03 Assumptions'!$B$43-'07 Working Capital'!$B$12*C$26)*(1+'03 Assumptions'!$F$16)/('10 WACC'!$B$15-'03 Assumptions'!$F$16))/(1+'10 WACC'!$B$15)^10+ '03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['D30']; c.value = "=(SUMPRODUCT( (1019.3*(1+D$26)^'05 Revenue Build'!$F$4:$O$4) * ( ((0.216+($B30-0.216)*(('05 Revenue Build'!$F$4:$O$4)*('05 Revenue Build'!$F$4:$O$4<=7)+7*('05 Revenue Build'!$F$4:$O$4>7))/7)-'03 Assumptions'!$B$22)*(1-'03 Assumptions'!$B$21) + '03 Assumptions'!$B$22 - '03 Assumptions'!$F$14 - '03 Assumptions'!$B$43 ) / (1+'10 WACC'!$B$15)^'05 Revenue Build'!$F$4:$O$4 )- SUMPRODUCT( '07 Working Capital'!$B$12*1019.3*((1+D$26)^'05 Revenue Build'!$F$4:$O$4-(1+D$26)^('05 Revenue Build'!$F$4:$O$4-1)) / (1+'10 WACC'!$B$15)^'05 Revenue Build'!$F$4:$O$4 )+ (1019.3*(1+D$26)^10*(($B30-'03 Assumptions'!$B$22)*(1-'03 Assumptions'!$B$21)+'03 Assumptions'!$B$22-'03 Assumptions'!$F$14-'03 Assumptions'!$B$43-'07 Working Capital'!$B$12*D$26)*(1+'03 Assumptions'!$F$16)/('10 WACC'!$B$15-'03 Assumptions'!$F$16))/(1+'10 WACC'!$B$15)^10+ '03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['E30']; c.value = "=(SUMPRODUCT( (1019.3*(1+E$26)^'05 Revenue Build'!$F$4:$O$4) * ( ((0.216+($B30-0.216)*(('05 Revenue Build'!$F$4:$O$4)*('05 Revenue Build'!$F$4:$O$4<=7)+7*('05 Revenue Build'!$F$4:$O$4>7))/7)-'03 Assumptions'!$B$22)*(1-'03 Assumptions'!$B$21) + '03 Assumptions'!$B$22 - '03 Assumptions'!$F$14 - '03 Assumptions'!$B$43 ) / (1+'10 WACC'!$B$15)^'05 Revenue Build'!$F$4:$O$4 )- SUMPRODUCT( '07 Working Capital'!$B$12*1019.3*((1+E$26)^'05 Revenue Build'!$F$4:$O$4-(1+E$26)^('05 Revenue Build'!$F$4:$O$4-1)) / (1+'10 WACC'!$B$15)^'05 Revenue Build'!$F$4:$O$4 )+ (1019.3*(1+E$26)^10*(($B30-'03 Assumptions'!$B$22)*(1-'03 Assumptions'!$B$21)+'03 Assumptions'!$B$22-'03 Assumptions'!$F$14-'03 Assumptions'!$B$43-'07 Working Capital'!$B$12*E$26)*(1+'03 Assumptions'!$F$16)/('10 WACC'!$B$15-'03 Assumptions'!$F$16))/(1+'10 WACC'!$B$15)^10+ '03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['F30']; c.value = "=(SUMPRODUCT( (1019.3*(1+F$26)^'05 Revenue Build'!$F$4:$O$4) * ( ((0.216+($B30-0.216)*(('05 Revenue Build'!$F$4:$O$4)*('05 Revenue Build'!$F$4:$O$4<=7)+7*('05 Revenue Build'!$F$4:$O$4>7))/7)-'03 Assumptions'!$B$22)*(1-'03 Assumptions'!$B$21) + '03 Assumptions'!$B$22 - '03 Assumptions'!$F$14 - '03 Assumptions'!$B$43 ) / (1+'10 WACC'!$B$15)^'05 Revenue Build'!$F$4:$O$4 )- SUMPRODUCT( '07 Working Capital'!$B$12*1019.3*((1+F$26)^'05 Revenue Build'!$F$4:$O$4-(1+F$26)^('05 Revenue Build'!$F$4:$O$4-1)) / (1+'10 WACC'!$B$15)^'05 Revenue Build'!$F$4:$O$4 )+ (1019.3*(1+F$26)^10*(($B30-'03 Assumptions'!$B$22)*(1-'03 Assumptions'!$B$21)+'03 Assumptions'!$B$22-'03 Assumptions'!$F$14-'03 Assumptions'!$B$43-'07 Working Capital'!$B$12*F$26)*(1+'03 Assumptions'!$F$16)/('10 WACC'!$B$15-'03 Assumptions'!$F$16))/(1+'10 WACC'!$B$15)^10+ '03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['G30']; c.value = "=(SUMPRODUCT( (1019.3*(1+G$26)^'05 Revenue Build'!$F$4:$O$4) * ( ((0.216+($B30-0.216)*(('05 Revenue Build'!$F$4:$O$4)*('05 Revenue Build'!$F$4:$O$4<=7)+7*('05 Revenue Build'!$F$4:$O$4>7))/7)-'03 Assumptions'!$B$22)*(1-'03 Assumptions'!$B$21) + '03 Assumptions'!$B$22 - '03 Assumptions'!$F$14 - '03 Assumptions'!$B$43 ) / (1+'10 WACC'!$B$15)^'05 Revenue Build'!$F$4:$O$4 )- SUMPRODUCT( '07 Working Capital'!$B$12*1019.3*((1+G$26)^'05 Revenue Build'!$F$4:$O$4-(1+G$26)^('05 Revenue Build'!$F$4:$O$4-1)) / (1+'10 WACC'!$B$15)^'05 Revenue Build'!$F$4:$O$4 )+ (1019.3*(1+G$26)^10*(($B30-'03 Assumptions'!$B$22)*(1-'03 Assumptions'!$B$21)+'03 Assumptions'!$B$22-'03 Assumptions'!$F$14-'03 Assumptions'!$B$43-'07 Working Capital'!$B$12*G$26)*(1+'03 Assumptions'!$F$16)/('10 WACC'!$B$15-'03 Assumptions'!$F$16))/(1+'10 WACC'!$B$15)^10+ '03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['B31']; c.value = 0.255
c.number_format = '0.0%'
c.font = Font(color='000000FF', size=10.0)
c = ws['C31']; c.value = "=(SUMPRODUCT( (1019.3*(1+C$26)^'05 Revenue Build'!$F$4:$O$4) * ( ((0.216+($B31-0.216)*(('05 Revenue Build'!$F$4:$O$4)*('05 Revenue Build'!$F$4:$O$4<=7)+7*('05 Revenue Build'!$F$4:$O$4>7))/7)-'03 Assumptions'!$B$22)*(1-'03 Assumptions'!$B$21) + '03 Assumptions'!$B$22 - '03 Assumptions'!$F$14 - '03 Assumptions'!$B$43 ) / (1+'10 WACC'!$B$15)^'05 Revenue Build'!$F$4:$O$4 )- SUMPRODUCT( '07 Working Capital'!$B$12*1019.3*((1+C$26)^'05 Revenue Build'!$F$4:$O$4-(1+C$26)^('05 Revenue Build'!$F$4:$O$4-1)) / (1+'10 WACC'!$B$15)^'05 Revenue Build'!$F$4:$O$4 )+ (1019.3*(1+C$26)^10*(($B31-'03 Assumptions'!$B$22)*(1-'03 Assumptions'!$B$21)+'03 Assumptions'!$B$22-'03 Assumptions'!$F$14-'03 Assumptions'!$B$43-'07 Working Capital'!$B$12*C$26)*(1+'03 Assumptions'!$F$16)/('10 WACC'!$B$15-'03 Assumptions'!$F$16))/(1+'10 WACC'!$B$15)^10+ '03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['D31']; c.value = "=(SUMPRODUCT( (1019.3*(1+D$26)^'05 Revenue Build'!$F$4:$O$4) * ( ((0.216+($B31-0.216)*(('05 Revenue Build'!$F$4:$O$4)*('05 Revenue Build'!$F$4:$O$4<=7)+7*('05 Revenue Build'!$F$4:$O$4>7))/7)-'03 Assumptions'!$B$22)*(1-'03 Assumptions'!$B$21) + '03 Assumptions'!$B$22 - '03 Assumptions'!$F$14 - '03 Assumptions'!$B$43 ) / (1+'10 WACC'!$B$15)^'05 Revenue Build'!$F$4:$O$4 )- SUMPRODUCT( '07 Working Capital'!$B$12*1019.3*((1+D$26)^'05 Revenue Build'!$F$4:$O$4-(1+D$26)^('05 Revenue Build'!$F$4:$O$4-1)) / (1+'10 WACC'!$B$15)^'05 Revenue Build'!$F$4:$O$4 )+ (1019.3*(1+D$26)^10*(($B31-'03 Assumptions'!$B$22)*(1-'03 Assumptions'!$B$21)+'03 Assumptions'!$B$22-'03 Assumptions'!$F$14-'03 Assumptions'!$B$43-'07 Working Capital'!$B$12*D$26)*(1+'03 Assumptions'!$F$16)/('10 WACC'!$B$15-'03 Assumptions'!$F$16))/(1+'10 WACC'!$B$15)^10+ '03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['E31']; c.value = "=(SUMPRODUCT( (1019.3*(1+E$26)^'05 Revenue Build'!$F$4:$O$4) * ( ((0.216+($B31-0.216)*(('05 Revenue Build'!$F$4:$O$4)*('05 Revenue Build'!$F$4:$O$4<=7)+7*('05 Revenue Build'!$F$4:$O$4>7))/7)-'03 Assumptions'!$B$22)*(1-'03 Assumptions'!$B$21) + '03 Assumptions'!$B$22 - '03 Assumptions'!$F$14 - '03 Assumptions'!$B$43 ) / (1+'10 WACC'!$B$15)^'05 Revenue Build'!$F$4:$O$4 )- SUMPRODUCT( '07 Working Capital'!$B$12*1019.3*((1+E$26)^'05 Revenue Build'!$F$4:$O$4-(1+E$26)^('05 Revenue Build'!$F$4:$O$4-1)) / (1+'10 WACC'!$B$15)^'05 Revenue Build'!$F$4:$O$4 )+ (1019.3*(1+E$26)^10*(($B31-'03 Assumptions'!$B$22)*(1-'03 Assumptions'!$B$21)+'03 Assumptions'!$B$22-'03 Assumptions'!$F$14-'03 Assumptions'!$B$43-'07 Working Capital'!$B$12*E$26)*(1+'03 Assumptions'!$F$16)/('10 WACC'!$B$15-'03 Assumptions'!$F$16))/(1+'10 WACC'!$B$15)^10+ '03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['F31']; c.value = "=(SUMPRODUCT( (1019.3*(1+F$26)^'05 Revenue Build'!$F$4:$O$4) * ( ((0.216+($B31-0.216)*(('05 Revenue Build'!$F$4:$O$4)*('05 Revenue Build'!$F$4:$O$4<=7)+7*('05 Revenue Build'!$F$4:$O$4>7))/7)-'03 Assumptions'!$B$22)*(1-'03 Assumptions'!$B$21) + '03 Assumptions'!$B$22 - '03 Assumptions'!$F$14 - '03 Assumptions'!$B$43 ) / (1+'10 WACC'!$B$15)^'05 Revenue Build'!$F$4:$O$4 )- SUMPRODUCT( '07 Working Capital'!$B$12*1019.3*((1+F$26)^'05 Revenue Build'!$F$4:$O$4-(1+F$26)^('05 Revenue Build'!$F$4:$O$4-1)) / (1+'10 WACC'!$B$15)^'05 Revenue Build'!$F$4:$O$4 )+ (1019.3*(1+F$26)^10*(($B31-'03 Assumptions'!$B$22)*(1-'03 Assumptions'!$B$21)+'03 Assumptions'!$B$22-'03 Assumptions'!$F$14-'03 Assumptions'!$B$43-'07 Working Capital'!$B$12*F$26)*(1+'03 Assumptions'!$F$16)/('10 WACC'!$B$15-'03 Assumptions'!$F$16))/(1+'10 WACC'!$B$15)^10+ '03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['G31']; c.value = "=(SUMPRODUCT( (1019.3*(1+G$26)^'05 Revenue Build'!$F$4:$O$4) * ( ((0.216+($B31-0.216)*(('05 Revenue Build'!$F$4:$O$4)*('05 Revenue Build'!$F$4:$O$4<=7)+7*('05 Revenue Build'!$F$4:$O$4>7))/7)-'03 Assumptions'!$B$22)*(1-'03 Assumptions'!$B$21) + '03 Assumptions'!$B$22 - '03 Assumptions'!$F$14 - '03 Assumptions'!$B$43 ) / (1+'10 WACC'!$B$15)^'05 Revenue Build'!$F$4:$O$4 )- SUMPRODUCT( '07 Working Capital'!$B$12*1019.3*((1+G$26)^'05 Revenue Build'!$F$4:$O$4-(1+G$26)^('05 Revenue Build'!$F$4:$O$4-1)) / (1+'10 WACC'!$B$15)^'05 Revenue Build'!$F$4:$O$4 )+ (1019.3*(1+G$26)^10*(($B31-'03 Assumptions'!$B$22)*(1-'03 Assumptions'!$B$21)+'03 Assumptions'!$B$22-'03 Assumptions'!$F$14-'03 Assumptions'!$B$43-'07 Working Capital'!$B$12*G$26)*(1+'03 Assumptions'!$F$16)/('10 WACC'!$B$15-'03 Assumptions'!$F$16))/(1+'10 WACC'!$B$15)^10+ '03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['A33']; c.value = 'Grid 3 is a full re-derivation, EXACT given its axis definitions (total-revenue CAGR, margin endpoint); digital'
c.font = Font(size=10.0)
c = ws['A34']; c.value = 'mix and bear path-shape definitionally held at base. Grids 1-2 discount the LIVE model FCFF row - no approximation.'
c.font = Font(size=10.0)

ws = wb.create_sheet('13 Scenarios')
ws.column_dimensions['A'].width = 46.0
ws.column_dimensions['B'].width = 12.0
ws.column_dimensions['C'].width = 12.0
ws.column_dimensions['D'].width = 12.0
ws.column_dimensions['E'].width = 12.0
ws.column_dimensions['F'].width = 12.0
ws.column_dimensions['G'].width = 12.0
ws.column_dimensions['H'].width = 12.0
ws.column_dimensions['I'].width = 12.0
ws.column_dimensions['J'].width = 12.0
ws.column_dimensions['K'].width = 12.0
ws.column_dimensions['L'].width = 12.0
ws.column_dimensions['M'].width = 12.0
ws.column_dimensions['N'].width = 12.0
ws.column_dimensions['O'].width = 12.0
ws.column_dimensions['P'].width = 12.0
ws.column_dimensions['Q'].width = 12.0
c = ws['A1']; c.value = 'SCENARIOS'
c.font = Font(bold=True, color='00FFFFFF', size=12.0)
c.fill = PatternFill("solid", fgColor='001F3864')
c = ws['A3']; c.value = 'To view a scenario through the FULL model, set the switch on 03 Assumptions (B3).'
c.font = Font(size=10.0)
c = ws['A4']; c.value = 'This sheet lists all three assumption sets; the ACTIVE column shows full-model outputs for the current switch.'
c.font = Font(size=10.0)
c = ws['A6']; c.value = 'Assumption'
c.font = Font(bold=True)
c = ws['B6']; c.value = 'Bear'
c.font = Font(bold=True)
c = ws['C6']; c.value = 'Base'
c.font = Font(bold=True)
c = ws['D6']; c.value = 'Bull'
c.font = Font(bold=True)
c = ws['A7']; c.value = "='03 Assumptions'!A6"
c.font = Font(color='00008000', size=10.0)
c = ws['B7']; c.value = "='03 Assumptions'!C6"
c.number_format = '0.0%'
c.font = Font(color='00008000', size=10.0)
c = ws['C7']; c.value = "='03 Assumptions'!D6"
c.number_format = '0.0%'
c.font = Font(color='00008000', size=10.0)
c = ws['D7']; c.value = "='03 Assumptions'!E6"
c.number_format = '0.0%'
c.font = Font(color='00008000', size=10.0)
c = ws['A8']; c.value = "='03 Assumptions'!A7"
c.font = Font(color='00008000', size=10.0)
c = ws['B8']; c.value = "='03 Assumptions'!C7"
c.number_format = '0.0%'
c.font = Font(color='00008000', size=10.0)
c = ws['C8']; c.value = "='03 Assumptions'!D7"
c.number_format = '0.0%'
c.font = Font(color='00008000', size=10.0)
c = ws['D8']; c.value = "='03 Assumptions'!E7"
c.number_format = '0.0%'
c.font = Font(color='00008000', size=10.0)
c = ws['A9']; c.value = "='03 Assumptions'!A8"
c.font = Font(color='00008000', size=10.0)
c = ws['B9']; c.value = "='03 Assumptions'!C8"
c.number_format = '0.0%'
c.font = Font(color='00008000', size=10.0)
c = ws['C9']; c.value = "='03 Assumptions'!D8"
c.number_format = '0.0%'
c.font = Font(color='00008000', size=10.0)
c = ws['D9']; c.value = "='03 Assumptions'!E8"
c.number_format = '0.0%'
c.font = Font(color='00008000', size=10.0)
c = ws['A10']; c.value = "='03 Assumptions'!A9"
c.font = Font(color='00008000', size=10.0)
c = ws['B10']; c.value = "='03 Assumptions'!C9"
c.number_format = '0.0%'
c.font = Font(color='00008000', size=10.0)
c = ws['C10']; c.value = "='03 Assumptions'!D9"
c.number_format = '0.0%'
c.font = Font(color='00008000', size=10.0)
c = ws['D10']; c.value = "='03 Assumptions'!E9"
c.number_format = '0.0%'
c.font = Font(color='00008000', size=10.0)
c = ws['A11']; c.value = "='03 Assumptions'!A10"
c.font = Font(color='00008000', size=10.0)
c = ws['B11']; c.value = "='03 Assumptions'!C10"
c.number_format = '0.0%'
c.font = Font(color='00008000', size=10.0)
c = ws['C11']; c.value = "='03 Assumptions'!D10"
c.number_format = '0.0%'
c.font = Font(color='00008000', size=10.0)
c = ws['D11']; c.value = "='03 Assumptions'!E10"
c.number_format = '0.0%'
c.font = Font(color='00008000', size=10.0)
c = ws['A12']; c.value = "='03 Assumptions'!A11"
c.font = Font(color='00008000', size=10.0)
c = ws['B12']; c.value = "='03 Assumptions'!C11"
c.number_format = '0.0%'
c.font = Font(color='00008000', size=10.0)
c = ws['C12']; c.value = "='03 Assumptions'!D11"
c.number_format = '0.0%'
c.font = Font(color='00008000', size=10.0)
c = ws['D12']; c.value = "='03 Assumptions'!E11"
c.number_format = '0.0%'
c.font = Font(color='00008000', size=10.0)
c = ws['A13']; c.value = "='03 Assumptions'!A12"
c.font = Font(color='00008000', size=10.0)
c = ws['B13']; c.value = "='03 Assumptions'!C12"
c.number_format = '0.0%'
c.font = Font(color='00008000', size=10.0)
c = ws['C13']; c.value = "='03 Assumptions'!D12"
c.number_format = '0.0%'
c.font = Font(color='00008000', size=10.0)
c = ws['D13']; c.value = "='03 Assumptions'!E12"
c.number_format = '0.0%'
c.font = Font(color='00008000', size=10.0)
c = ws['A14']; c.value = "='03 Assumptions'!A13"
c.font = Font(color='00008000', size=10.0)
c = ws['B14']; c.value = "='03 Assumptions'!C13"
c.number_format = '0.0%'
c.font = Font(color='00008000', size=10.0)
c = ws['C14']; c.value = "='03 Assumptions'!D13"
c.number_format = '0.0%'
c.font = Font(color='00008000', size=10.0)
c = ws['D14']; c.value = "='03 Assumptions'!E13"
c.number_format = '0.0%'
c.font = Font(color='00008000', size=10.0)
c = ws['A15']; c.value = "='03 Assumptions'!A14"
c.font = Font(color='00008000', size=10.0)
c = ws['B15']; c.value = "='03 Assumptions'!C14"
c.number_format = '0.0%'
c.font = Font(color='00008000', size=10.0)
c = ws['C15']; c.value = "='03 Assumptions'!D14"
c.number_format = '0.0%'
c.font = Font(color='00008000', size=10.0)
c = ws['D15']; c.value = "='03 Assumptions'!E14"
c.number_format = '0.0%'
c.font = Font(color='00008000', size=10.0)
c = ws['A16']; c.value = "='03 Assumptions'!A15"
c.font = Font(color='00008000', size=10.0)
c = ws['B16']; c.value = "='03 Assumptions'!C15"
c.number_format = '0.0%'
c.font = Font(color='00008000', size=10.0)
c = ws['C16']; c.value = "='03 Assumptions'!D15"
c.number_format = '0.0%'
c.font = Font(color='00008000', size=10.0)
c = ws['D16']; c.value = "='03 Assumptions'!E15"
c.number_format = '0.0%'
c.font = Font(color='00008000', size=10.0)
c = ws['A17']; c.value = "='03 Assumptions'!A16"
c.font = Font(color='00008000', size=10.0)
c = ws['B17']; c.value = "='03 Assumptions'!C16"
c.number_format = '0.0%'
c.font = Font(color='00008000', size=10.0)
c = ws['C17']; c.value = "='03 Assumptions'!D16"
c.number_format = '0.0%'
c.font = Font(color='00008000', size=10.0)
c = ws['D17']; c.value = "='03 Assumptions'!E16"
c.number_format = '0.0%'
c.font = Font(color='00008000', size=10.0)
c = ws['A18']; c.value = "='03 Assumptions'!A17"
c.font = Font(color='00008000', size=10.0)
c = ws['B18']; c.value = "='03 Assumptions'!C17"
c.number_format = '0.0x'
c.font = Font(color='00008000', size=10.0)
c = ws['C18']; c.value = "='03 Assumptions'!D17"
c.number_format = '0.0x'
c.font = Font(color='00008000', size=10.0)
c = ws['D18']; c.value = "='03 Assumptions'!E17"
c.number_format = '0.0x'
c.font = Font(color='00008000', size=10.0)
c = ws['A20']; c.value = 'ACTIVE scenario full-model outputs'
c.font = Font(bold=True)
c = ws['A21']; c.value = 'Active scenario (1/2/3)'
c.font = Font(size=10.0)
c = ws['B21']; c.value = "='03 Assumptions'!$B$3"
c.number_format = '#,##0;(#,##0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['A22']; c.value = 'Value per share (Gordon)'
c.font = Font(size=10.0)
c = ws['B22']; c.value = "='11 DCF Valuation'!B27"
c.number_format = '#,##0.00'
c.font = Font(color='00008000', size=10.0)
c = ws['A23']; c.value = 'Value per share (exit multiple)'
c.font = Font(size=10.0)
c = ws['B23']; c.value = "='11 DCF Valuation'!B31"
c.number_format = '#,##0.00'
c.font = Font(color='00008000', size=10.0)
c = ws['A24']; c.value = 'Upside vs price'
c.font = Font(size=10.0)
c = ws['B24']; c.value = "='11 DCF Valuation'!B29"
c.number_format = '0.0%'
c.font = Font(color='00008000', size=10.0)
c = ws['A26']; c.value = 'Why each scenario exists'
c.font = Font(bold=True)
c = ws['A27']; c.value = 'BEAR: BtoB capex cycle turns 2026-27; margin expansion stalls (mean reversion); digital never monetises. Tests cyclical downside.'
c.font = Font(size=10.0)
c = ws['A28']; c.value = 'BASE: growth normalises toward GDP+, margins fade to 23.5% endpoint as procurement gains annualise; digital modest. Most-likely path.'
c.font = Font(size=10.0)
c = ws['A29']; c.value = 'BULL: platform thesis works — attach doubles, ARPU compounds, margins reach 25.5% on mix. Tests what consensus TP requires.'
c.font = Font(size=10.0)
c = ws['A30']; c.value = 'Most valuable assumptions (sensitivity ranking): BtoB growth path > margin endpoint > WACC > terminal g > digital attach > capex.'
c.font = Font(size=10.0)

ws = wb.create_sheet('14 Charts & Dashboard')
ws.column_dimensions['A'].width = 46.0
ws.column_dimensions['B'].width = 12.0
ws.column_dimensions['C'].width = 12.0
ws.column_dimensions['D'].width = 12.0
ws.column_dimensions['E'].width = 12.0
ws.column_dimensions['F'].width = 12.0
ws.column_dimensions['G'].width = 12.0
ws.column_dimensions['H'].width = 12.0
ws.column_dimensions['I'].width = 12.0
ws.column_dimensions['J'].width = 12.0
ws.column_dimensions['K'].width = 12.0
ws.column_dimensions['L'].width = 12.0
ws.column_dimensions['M'].width = 12.0
ws.column_dimensions['N'].width = 12.0
ws.column_dimensions['O'].width = 12.0
ws.column_dimensions['P'].width = 12.0
ws.column_dimensions['Q'].width = 12.0
c = ws['A1']; c.value = 'CHARTS & DASHBOARD'
c.font = Font(bold=True, color='00FFFFFF', size=12.0)
c.fill = PatternFill("solid", fgColor='001F3864')
c = ws['A3']; c.value = 'KEY OUTPUTS (active scenario)'
c.font = Font(bold=True)
c = ws['A4']; c.value = 'Value per share — Gordon (€)'
c.font = Font(size=10.0)
c = ws['B4']; c.value = "='11 DCF Valuation'!B27"
c.number_format = '#,##0.00'
c.font = Font(color='00008000', size=10.0)
c = ws['A5']; c.value = 'Value per share — exit multiple (€)'
c.font = Font(size=10.0)
c = ws['B5']; c.value = "='11 DCF Valuation'!B31"
c.number_format = '#,##0.00'
c.font = Font(color='00008000', size=10.0)
c = ws['A6']; c.value = 'Current price (€)'
c.font = Font(size=10.0)
c = ws['B6']; c.value = "='11 DCF Valuation'!B28"
c.number_format = '#,##0.00'
c.font = Font(color='00008000', size=10.0)
c = ws['A7']; c.value = 'Upside / (downside)'
c.font = Font(size=10.0)
c = ws['B7']; c.value = "='11 DCF Valuation'!B29"
c.number_format = '0.0%'
c.font = Font(color='00008000', size=10.0)
c = ws['A8']; c.value = 'WACC'
c.font = Font(size=10.0)
c = ws['B8']; c.value = "='10 WACC'!B15"
c.number_format = '0.0%'
c.font = Font(color='00008000', size=10.0)
c = ws['A9']; c.value = 'Enterprise value (€m)'
c.font = Font(size=10.0)
c = ws['B9']; c.value = "='11 DCF Valuation'!B23"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['A10']; c.value = 'TV % of EV'
c.font = Font(size=10.0)
c = ws['B10']; c.value = "='11 DCF Valuation'!B32"
c.number_format = '0.0%'
c.font = Font(color='00008000', size=10.0)
c = ws['A11']; c.value = '2035 revenue (€m)'
c.font = Font(size=10.0)
c = ws['B11']; c.value = "='05 Revenue Build'!O27"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['A12']; c.value = '2035 EBITDA margin'
c.font = Font(size=10.0)
c = ws['B12']; c.value = "='06 Operating Model'!O5"
c.number_format = '0.0%'
c.font = Font(color='00008000', size=10.0)
c = ws['A13']; c.value = '2035 digital % of revenue'
c.font = Font(size=10.0)
c = ws['B13']; c.value = "='05 Revenue Build'!O29"
c.number_format = '0.0%'
c.font = Font(color='00008000', size=10.0)
c = ws['A16']; c.value = 'Chart data (live links)'
c.font = Font(bold=True)
c = ws['A17']; c.value = 'Year'
c.font = Font(size=10.0)
c = ws['F17']; c.value = 2026
c.number_format = '#,##0;(#,##0);-'
c.font = Font(size=10.0)
c = ws['G17']; c.value = 2027
c.number_format = '#,##0;(#,##0);-'
c.font = Font(size=10.0)
c = ws['H17']; c.value = 2028
c.number_format = '#,##0;(#,##0);-'
c.font = Font(size=10.0)
c = ws['I17']; c.value = 2029
c.number_format = '#,##0;(#,##0);-'
c.font = Font(size=10.0)
c = ws['J17']; c.value = 2030
c.number_format = '#,##0;(#,##0);-'
c.font = Font(size=10.0)
c = ws['K17']; c.value = 2031
c.number_format = '#,##0;(#,##0);-'
c.font = Font(size=10.0)
c = ws['L17']; c.value = 2032
c.number_format = '#,##0;(#,##0);-'
c.font = Font(size=10.0)
c = ws['M17']; c.value = 2033
c.number_format = '#,##0;(#,##0);-'
c.font = Font(size=10.0)
c = ws['N17']; c.value = 2034
c.number_format = '#,##0;(#,##0);-'
c.font = Font(size=10.0)
c = ws['O17']; c.value = 2035
c.number_format = '#,##0;(#,##0);-'
c.font = Font(size=10.0)
c = ws['A18']; c.value = 'Revenue'
c.font = Font(size=10.0)
c = ws['F18']; c.value = "='05 Revenue Build'!F27"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['G18']; c.value = "='05 Revenue Build'!G27"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['H18']; c.value = "='05 Revenue Build'!H27"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['I18']; c.value = "='05 Revenue Build'!I27"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['J18']; c.value = "='05 Revenue Build'!J27"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['K18']; c.value = "='05 Revenue Build'!K27"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['L18']; c.value = "='05 Revenue Build'!L27"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['M18']; c.value = "='05 Revenue Build'!M27"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['N18']; c.value = "='05 Revenue Build'!N27"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['O18']; c.value = "='05 Revenue Build'!O27"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['A19']; c.value = 'EBITDA'
c.font = Font(size=10.0)
c = ws['F19']; c.value = "='06 Operating Model'!F6"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['G19']; c.value = "='06 Operating Model'!G6"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['H19']; c.value = "='06 Operating Model'!H6"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['I19']; c.value = "='06 Operating Model'!I6"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['J19']; c.value = "='06 Operating Model'!J6"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['K19']; c.value = "='06 Operating Model'!K6"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['L19']; c.value = "='06 Operating Model'!L6"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['M19']; c.value = "='06 Operating Model'!M6"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['N19']; c.value = "='06 Operating Model'!N6"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['O19']; c.value = "='06 Operating Model'!O6"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['A20']; c.value = 'FCFF'
c.font = Font(size=10.0)
c = ws['F20']; c.value = "='09 Free Cash Flow'!F9"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['G20']; c.value = "='09 Free Cash Flow'!G9"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['H20']; c.value = "='09 Free Cash Flow'!H9"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['I20']; c.value = "='09 Free Cash Flow'!I9"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['J20']; c.value = "='09 Free Cash Flow'!J9"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['K20']; c.value = "='09 Free Cash Flow'!K9"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['L20']; c.value = "='09 Free Cash Flow'!L9"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['M20']; c.value = "='09 Free Cash Flow'!M9"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['N20']; c.value = "='09 Free Cash Flow'!N9"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['O20']; c.value = "='09 Free Cash Flow'!O9"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['A21']; c.value = 'BtoB rev'
c.font = Font(size=10.0)
c = ws['F21']; c.value = "='05 Revenue Build'!F8"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['G21']; c.value = "='05 Revenue Build'!G8"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['H21']; c.value = "='05 Revenue Build'!H8"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['I21']; c.value = "='05 Revenue Build'!I8"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['J21']; c.value = "='05 Revenue Build'!J8"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['K21']; c.value = "='05 Revenue Build'!K8"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['L21']; c.value = "='05 Revenue Build'!L8"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['M21']; c.value = "='05 Revenue Build'!M8"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['N21']; c.value = "='05 Revenue Build'!N8"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['O21']; c.value = "='05 Revenue Build'!O8"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['A22']; c.value = 'BtoC rev'
c.font = Font(size=10.0)
c = ws['F22']; c.value = "='05 Revenue Build'!F12"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['G22']; c.value = "='05 Revenue Build'!G12"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['H22']; c.value = "='05 Revenue Build'!H12"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['I22']; c.value = "='05 Revenue Build'!I12"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['J22']; c.value = "='05 Revenue Build'!J12"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['K22']; c.value = "='05 Revenue Build'!K12"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['L22']; c.value = "='05 Revenue Build'!L12"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['M22']; c.value = "='05 Revenue Build'!M12"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['N22']; c.value = "='05 Revenue Build'!N12"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['O22']; c.value = "='05 Revenue Build'!O12"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['A23']; c.value = 'Digital rev'
c.font = Font(size=10.0)
c = ws['F23']; c.value = "='05 Revenue Build'!F25"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['G23']; c.value = "='05 Revenue Build'!G25"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['H23']; c.value = "='05 Revenue Build'!H25"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['I23']; c.value = "='05 Revenue Build'!I25"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['J23']; c.value = "='05 Revenue Build'!J25"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['K23']; c.value = "='05 Revenue Build'!K25"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['L23']; c.value = "='05 Revenue Build'!L25"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['M23']; c.value = "='05 Revenue Build'!M25"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['N23']; c.value = "='05 Revenue Build'!N25"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['O23']; c.value = "='05 Revenue Build'!O25"
c.number_format = '#,##0.0;(#,##0.0);-'
c.font = Font(color='00008000', size=10.0)
c = ws['A25']; c.value = 'Football field (per share, €)'
c.font = Font(bold=True)
c = ws['A26']; c.value = 'Method'
c.font = Font(size=10.0)
c = ws['B26']; c.value = 'Low'
c.font = Font(size=10.0)
c = ws['C26']; c.value = 'High'
c.font = Font(size=10.0)
c = ws['A27']; c.value = '52-week range'
c.font = Font(size=10.0)
c = ws['B27']; c.value = 11
c.number_format = '#,##0.00'
c.font = Font(color='000000FF', size=10.0)
c = ws['C27']; c.value = 21.82
c.number_format = '#,##0.00'
c.font = Font(color='000000FF', size=10.0)
c = ws['A28']; c.value = 'Comps (10.5x/14x 2026E EBITDA)'
c.font = Font(size=10.0)
c = ws['B28']; c.value = "=(10.5*'06 Operating Model'!F6+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['C28']; c.value = "=(14*'06 Operating Model'!F6+'03 Assumptions'!$B$44)/'03 Assumptions'!$B$23"
c.number_format = '#,##0.00'
c.font = Font(size=10.0)
c = ws['A29']; c.value = 'DCF (grid1 min/max)'
c.font = Font(size=10.0)
c = ws['B29']; c.value = "='12 Sensitivity Analysis'!C11"
c.number_format = '#,##0.00'
c.font = Font(color='00008000', size=10.0)
c = ws['C29']; c.value = "='12 Sensitivity Analysis'!G6"
c.number_format = '#,##0.00'
c.font = Font(color='00008000', size=10.0)
c = ws['A30']; c.value = 'Consensus TP'
c.font = Font(size=10.0)
c = ws['B30']; c.value = 19.3
c.number_format = '#,##0.00'
c.font = Font(color='000000FF', size=10.0)
c = ws['C30']; c.value = 20.6
c.number_format = '#,##0.00'
c.font = Font(color='000000FF', size=10.0)
c = ws['A33']; c.value = 'ILLUSTRATIVE PROBABILITY-WEIGHTED FAIR VALUE (subjective weights - NOT a headline valuation)'
c = ws['A34']; c.value = 'Scenario'
c = ws['B34']; c.value = 'Value (EUR)'
c = ws['C34']; c.value = 'Weight'
c = ws['A35']; c.value = 'Bear'
c = ws['B35']; c.value = 8.82
c = ws['C35']; c.value = 0.4
c = ws['A36']; c.value = 'Base'
c = ws['B36']; c.value = 15.72
c = ws['C36']; c.value = 0.4
c = ws['A37']; c.value = 'Bull'
c = ws['B37']; c.value = 24.05
c = ws['C37']; c.value = 0.2
c = ws['A38']; c.value = 'Prob-weighted FV'
c = ws['B38']; c.value = '=SUMPRODUCT(B35:B37,C35:C37)'
c = ws['C38']; c.value = '=SUM(C35:C37)'
c = ws['A39']; c.value = 'Values recorded from scenario runs 18-Jul-26 (re-run switch 03!B3 to refresh). Weights illustrative (40/40/20); 40% bear reflects cyclical-capex + Americas risk. Headline remains base-case EUR 15.72 / HOLD.'
c = ws['A40']; c.value = '=IF(ABS(C38-1)<0.001,"CHECK OK: weights sum to 100%","ERROR: weights must sum to 1")'

wb.save('Technogym_DCF_Model.xlsx')
print("wrote", 'Technogym_DCF_Model.xlsx')